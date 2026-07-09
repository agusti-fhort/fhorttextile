"""pom/dictionary_service.py — diccionari de nomenclatura del client (setup, un sol cop).

Carrega l'Excel del client (codi_client + descripció EN + descripció local + idioma), i per
cada fila PROPOSA un POM via find_pom_master (mai desa automàticament). La revisió humana
(Montse) confirma la taula sencera abans del commit. Stateless: no hi ha taula de staging;
el preview retorna les propostes i el commit rep la taula ja validada per l'usuari.

Precedent openpyxl: models_app/bulk_import_service.py. Llibreria: openpyxl (ja instal·lada).
"""
import datetime
import io

# Columnes del full Plantilla (en ordre). Només codi_client és obligatori.
COLUMNS = ['codi_client', 'descripcio_en', 'descripcio_local', 'idioma']
META_SHEET = '_meta'
PLANTILLA_SHEET = 'Diccionari'


def _norm(s):
    return str(s).strip() if s is not None else ''


# ───────────────────────────── Plantilla descarregable ─────────────────────────────

def generate_template_bytes(customer):
    """Workbook de plantilla del diccionari per a un Customer, com a bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = PLANTILLA_SHEET
    bold = Font(bold=True)
    for ci, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=name)
        if name == 'codi_client':
            cell.font = bold

    inst = wb.create_sheet('Instruccions', 0)
    inst['A1'] = f"Diccionari de nomenclatura per a: {customer.codi} — {customer.nom}"
    inst['A3'] = "Una fila per cada nomenclatura del client. Només 'codi_client' és obligatori."
    inst['A5'] = "codi_client: com anomena el client la mesura (ex: H.11, CHEST, B)."
    inst['A6'] = "descripcio_en: descripció canònica internacional (anglès)."
    inst['A7'] = "descripcio_local: descripció en l'idioma de l'empresa."
    inst['A8'] = "idioma: codi ISO 639-1 de la descripció local (ex: ca, es, fr). Opcional."
    inst['A10'] = "En carregar-lo, el sistema PROPOSA un POM per fila; tu revises i confirmes."
    inst['A11'] = "RES de toleràncies ni graduació: això és només nomenclatura."
    wb.active = 0

    meta = wb.create_sheet(META_SHEET)
    meta.sheet_state = 'hidden'
    meta['A1'] = customer.codi
    meta['A2'] = customer.id

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ───────────────────────────── Parse del fitxer pujat ─────────────────────────────

def parse_upload(file_bytes):
    """Llegeix el xlsx. Retorna (detected_customer_codi, raw_rows) amb raw_rows =
    llista de dicts {row_num, cells:{columna: valor}} (files buides ignorades)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    detected = None
    if META_SHEET in wb.sheetnames:
        detected = _norm(wb[META_SHEET]['A1'].value) or None

    ws = wb[PLANTILLA_SHEET] if PLANTILLA_SHEET in wb.sheetnames else wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not all_rows:
        return detected, []
    header = [_norm(c).lower() for c in all_rows[0]]

    raw_rows = []
    for excel_idx, row in enumerate(all_rows[1:], start=2):
        cells, any_val = {}, False
        for ci, col in enumerate(header):
            if not col:
                continue
            val = _norm(row[ci]) if ci < len(row) else ''
            cells[col] = val
            if val:
                any_val = True
        if any_val:
            raw_rows.append({'row_num': excel_idx, 'cells': cells})
    return detected, raw_rows


# ───────────────────────────── Proposta (find_pom_master) ─────────────────────────────

def _pom_payload(pm):
    """Display canònic d'un POM proposat (codi POM-XXX + abreviatura + nom EN)."""
    if pm is None:
        return None
    pg = pm.pom_global
    return {
        'pom_master_id': pm.id,
        'codi_global': (pg.codi if pg else None),
        'abbreviation': (pg.abbreviation if pg else None) or pm.codi_client,
        'nom_en': (pg.nom_en if pg else None) or pm.nom_client,
        'codi_client': pm.codi_client,
    }


def _count_candidates(desc, poms_cache):
    """#match informatiu: quants POMs actius casa la descripció per substring (no bloquejant;
    substitueix el criteri de 0031 com a AVÍS d'ambigüitat, no com a resolutor)."""
    d = (desc or '').lower().strip()
    if not d or len(d) <= 3:
        return 0
    n = 0
    for pm in poms_cache:
        nom = (pm.nom_client or '').lower()
        nom_en = (pm.pom_global.nom_en or '').lower() if pm.pom_global_id else ''
        if (nom and (d in nom or nom in d)) or (nom_en and (d in nom_en or nom_en in d)):
            n += 1
    return n


def build_preview(customer, raw_rows):
    """Per cada fila: proposta de POM (find_pom_master), badge de confiança, #match, i DIFF
    respecte de l'àlies existent (marca CORRECCIÓ HUMANA si l'existent és origen=MANUAL).
    Cap escriptura. Retorna (rows, resum)."""
    from fhort.pom.models import POMMaster, CustomerPOMAlias
    from fhort.models_app.extraction_views import find_pom_master

    existing = {
        (a.client_code or '').strip().lower(): a
        for a in CustomerPOMAlias.objects.filter(customer=customer).select_related('pom', 'pom__pom_global')
    }
    poms_cache = list(POMMaster.objects.select_related('pom_global').filter(actiu=True))

    rows, n_auto, n_diff, n_nomatch = [], 0, 0, 0
    for r in raw_rows:
        c = r['cells']
        code = (c.get('codi_client') or '').strip()
        if not code:
            continue
        desc_en = (c.get('descripcio_en') or '').strip()
        desc_local = (c.get('descripcio_local') or '').strip()
        idioma = (c.get('idioma') or '').strip().lower()[:2]

        pm, mtype, conf = find_pom_master(code, desc_en or desc_local, customer=customer)
        n_match = _count_candidates(desc_en or desc_local, poms_cache)
        ex = existing.get(code.lower())
        ex_payload = None
        diff = False
        preserve_manual = False
        if ex is not None:
            ex_payload = {**_pom_payload(ex.pom), 'origen': ex.origen}
            diff = (pm is None) or (ex.pom_id != pm.id)
            preserve_manual = (ex.origen == 'MANUAL')

        if conf in ('HIGH', 'MEDIUM') and pm is not None:
            n_auto += 1
        if pm is None:
            n_nomatch += 1
        if diff:
            n_diff += 1

        rows.append({
            'row_num': r['row_num'],
            'codi_client': code,
            'descripcio_en': desc_en,
            'descripcio_local': desc_local,
            'idioma': idioma,
            'proposal': _pom_payload(pm),
            'match_type': mtype,
            'confidence': conf,
            'n_match': n_match,
            'existing': ex_payload,
            'diff': diff,
            'preserve_manual': preserve_manual,
        })

    resum = {
        'total': len(rows),
        'auto': n_auto,
        'diff': n_diff,
        'no_match': n_nomatch,
    }
    return rows, resum
