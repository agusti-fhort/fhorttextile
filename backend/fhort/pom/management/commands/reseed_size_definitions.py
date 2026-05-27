"""
Reseed SizeDefinitions des de FHORT_Master_Data_Reference_v2.xlsx (full Size_Definitions).

Llegeix les files de l'Excel i fa get_or_create per (size_system, etiqueta).
Si la SizeDefinition existeix → actualitza els camps numèrics.
Si no existeix → la crea.

Mapping de columnes Excel → camps del model:
  size_label     → etiqueta
  display_order  → ordre
  body_height_cm → body_height_cm
  body_bust_cm   → body_bust_cm
  body_waist_cm  → body_waist_cm
  body_hip_cm    → body_hip_cm
  age_months_min → age_months_min
  age_months_max → age_months_max
  notes          → (ignorat, no existeix al model)

Executa sobre el tenant 'fhort' via schema_context.
"""

from pathlib import Path

from django.core.management.base import BaseCommand
from django_tenants.utils import schema_context

EXCEL_PATH = Path('/root/fhort-sessions/FHORT_Master_Data_Reference_v2.xlsx')
SHEET_NAME = 'Size_Definitions'
TENANT_SCHEMA = 'fhort'
HEADER_ROW = 3  # capçalera real (files 1-2 són títol/descripció)


def _to_value(cell):
    """Convert 'NULL' (string) i strings buits a None; tot la resta es retorna tal qual."""
    if cell is None:
        return None
    if isinstance(cell, str):
        s = cell.strip()
        if s == '' or s.upper() == 'NULL':
            return None
        return s
    return cell


def _to_float(cell):
    v = _to_value(cell)
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_int(cell):
    v = _to_value(cell)
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


class Command(BaseCommand):
    help = 'Reseed SizeDefinitions des del full Size_Definitions de l\'Excel master.'

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR('openpyxl no està instal·lat al venv'))
            return

        if not EXCEL_PATH.exists():
            self.stderr.write(self.style.ERROR(f'No trobo l\'Excel: {EXCEL_PATH}'))
            return

        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            self.stderr.write(self.style.ERROR(f'No trobo el full {SHEET_NAME!r}'))
            return
        ws = wb[SHEET_NAME]

        # Llegim totes les files; la capçalera real és a HEADER_ROW.
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < HEADER_ROW + 1:
            self.stderr.write(self.style.ERROR('Full sense dades'))
            return

        header = [c.strip() if isinstance(c, str) else c for c in rows[HEADER_ROW - 1]]
        try:
            idx = {name: header.index(name) for name in (
                'size_system', 'size_label', 'display_order',
                'body_height_cm', 'body_bust_cm', 'body_waist_cm', 'body_hip_cm',
                'age_months_min', 'age_months_max',
            )}
        except ValueError as e:
            self.stderr.write(self.style.ERROR(f'Capçalera incompleta: {e}'))
            return

        data_rows = rows[HEADER_ROW:]  # files després de la capçalera

        with schema_context(TENANT_SCHEMA):
            from fhort.pom.models import SizeSystem, SizeDefinition

            created = updated = skipped = 0
            unknown_systems = set()
            missing_label = 0

            for row_num, row in enumerate(data_rows, start=HEADER_ROW + 1):
                if not row or all(c is None or (isinstance(c, str) and c.strip() == '') for c in row):
                    continue  # fila completament buida

                codi_sys = _to_value(row[idx['size_system']])
                etiqueta = _to_value(row[idx['size_label']])

                if not codi_sys:
                    continue
                if not etiqueta:
                    missing_label += 1
                    self.stdout.write(self.style.WARNING(
                        f'  fila {row_num}: size_label buit per {codi_sys!r} — skip'
                    ))
                    skipped += 1
                    continue

                ss = SizeSystem.objects.filter(codi=codi_sys).first()
                if not ss:
                    unknown_systems.add(codi_sys)
                    skipped += 1
                    continue

                defaults = {
                    'ordre':          _to_int(row[idx['display_order']])   or 0,
                    'body_height_cm': _to_float(row[idx['body_height_cm']]),
                    'body_bust_cm':   _to_float(row[idx['body_bust_cm']]),
                    'body_waist_cm':  _to_float(row[idx['body_waist_cm']]),
                    'body_hip_cm':    _to_float(row[idx['body_hip_cm']]),
                    'age_months_min': _to_int(row[idx['age_months_min']]),
                    'age_months_max': _to_int(row[idx['age_months_max']]),
                }

                sd, was_created = SizeDefinition.objects.get_or_create(
                    size_system=ss,
                    etiqueta=etiqueta,
                    defaults=defaults,
                )
                if was_created:
                    created += 1
                else:
                    changed = False
                    for k, v in defaults.items():
                        if getattr(sd, k) != v:
                            setattr(sd, k, v)
                            changed = True
                    if changed:
                        sd.save(update_fields=list(defaults.keys()))
                    updated += 1

            self.stdout.write('')
            self.stdout.write(self.style.SUCCESS(f'Creades:      {created}'))
            self.stdout.write(self.style.SUCCESS(f'Actualitzades: {updated}'))
            self.stdout.write(self.style.WARNING(f'Saltades:     {skipped}'))
            if unknown_systems:
                self.stdout.write(self.style.WARNING(
                    f'SizeSystem desconeguts: {sorted(unknown_systems)}'
                ))
            if missing_label:
                self.stdout.write(self.style.WARNING(
                    f'Files sense size_label: {missing_label}'
                ))
