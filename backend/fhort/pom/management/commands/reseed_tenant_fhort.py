"""
S14-A · Reseed complet del tenant 'fhort' després de la neteja.

Pas A — POMMaster (106, un per POMGlobal)
Pas B — GarmentPOMMap (~85, des d'Excel)
Pas C — GradingRuleSet (14) + GradingRule (~59), des d'Excel
Pas D — SizingProfile (expansió 1 Excel row → N GarmentType del grup)

Adaptacions vs plan original:
- GarmentPOMMap usa `ordre` (no `display_order`).
- GradingRule.pom és FK a POMMaster (no POMGlobal); requereix talla_base i valor_base.
- GradingRule no té camp `increment_above_xl`: es desa a `valors_step['above_xl']`.
- SizingProfile no té `garment_group`: expandim per GarmentType del grup.
- size_system del RuleSet es resol des dels SizingProfiles que el referencien (l'Excel no ho dóna directament al sheet RuleSets).
"""
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django_tenants.utils import schema_context


EXCEL_PATH = '/root/fhort-sessions/FHORT_Master_Data_Reference_v2.xlsx'
GRADING_EXCEL_PATH = '/root/fhort-sessions/FHORT_GradingRules_Dataset_v1.xlsx'


def _find_header(rows, key='id', second=None):
    for i, r in enumerate(rows):
        if r and r[0] == key and (second is None or (len(r) > 1 and r[1] == second)):
            return i
    return None


def _str(v):
    return str(v).strip() if v not in (None, '') else ''


def _to_float(v):
    if v in (None, ''):
        return 0.0
    s = str(v).replace('—', '').replace(',', '.').strip()
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _to_int(v):
    if v in (None, ''):
        return 0
    try:
        return int(str(v).strip())
    except ValueError:
        return 0


def _parse_bool(v, default=False):
    if v in (None, ''):
        return default
    return _str(v).lower() in ('true', '✓', '✓ key', 'yes', '1')


class Command(BaseCommand):
    help = 'S14-A · Reseed complet del tenant fhort des de l\'Excel mestre'

    def add_arguments(self, parser):
        parser.add_argument('--tenant', default='fhort')
        parser.add_argument('--excel', default=EXCEL_PATH,
                            help='Excel mestre (POMs, GarmentPOMMap, SizingProfiles)')
        parser.add_argument('--grading-excel', default=GRADING_EXCEL_PATH,
                            help='Excel ampliat de GradingRuleSets + GradingRules (S14-A)')

    def handle(self, *args, **opts):
        tenant = opts['tenant']
        excel_path = opts['excel']
        grading_excel_path = opts['grading_excel']

        # ── Carregar Excels ─────────────────────────────────────────────
        self.stdout.write(self.style.WARNING(f'Carregant Excel mestre: {excel_path}'))
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        self.stdout.write(self.style.WARNING(f'Carregant Excel grading: {grading_excel_path}'))
        wb_grad = openpyxl.load_workbook(grading_excel_path, read_only=True, data_only=True)

        # ── Garment_POM_Map ─────────────────────────────────────────────
        ws = wb['Garment_POM_Map']
        rows = list(ws.iter_rows(values_only=True))
        h = _find_header(rows, 'id', 'garment_type')
        assert h is not None, 'No s\'ha trobat capçalera de Garment_POM_Map'
        gpm_data = []
        for r in rows[h+1:]:
            if not r or len(r) < 7 or not r[0] or not r[1]:
                continue
            if not _str(r[0]).isdigit():
                continue
            gpm_data.append({
                'garment_type_codi': _str(r[1]),
                'pom_code': _str(r[2]),
                'ordre': _to_int(r[4]),
                'is_key': _str(r[5]) == '✓ KEY',
                'obligatori': _str(r[6]).upper() == 'TRUE',
            })

        # ── GradingRuleSets (NOU dataset v1, 18 RuleSets) ───────────────
        # Capçalera a fila idx 1: codi_sistema, nom_en, nom_ca, target,
        # construction, fit_type, base_size, grade_increment_ref,
        # is_system_default, norma_ref, notes
        # Camps norma_ref/notes ignorats: el model GradingRuleSet no els té.
        ws = wb_grad['GradingRuleSets']
        rows = list(ws.iter_rows(values_only=True))
        h = _find_header(rows, 'codi_sistema')
        assert h is not None, 'No s\'ha trobat capçalera de GradingRuleSets'
        rs_data = []
        for r in rows[h+1:]:
            if not r or not r[0]:
                continue
            codi = _str(r[0])
            # Línies de secció/títol començarien per text descriptiu; les
            # files de dades són identificades pel codi_sistema (no buit)
            # i un nom_en a la cel·la r[1].
            if not _str(r[1]):
                continue
            rs_data.append({
                'codi_sistema': codi,
                'nom_en': _str(r[1]),
                'nom_ca': _str(r[2]),
                'target': _str(r[3]),
                'construction': _str(r[4]),
                'fit_type': _str(r[5]),
                'base_size': _str(r[6]),
                'is_system_default': _parse_bool(r[8]),
            })

        # ── GradingRules (NOU dataset v1, ~618 regles + files de secció) ─
        # Capçalera a fila idx 1: ruleset_codi, pom_code, logica,
        # increment_cm, increment_above_xl_cm, actiu, notes_en, norma_ref
        # Files de secció: r[1] (pom_code) és None i r[0] conté "—".
        # Camp notes_en/norma_ref ignorats: el model GradingRule no els té.
        ws = wb_grad['GradingRules']
        rows = list(ws.iter_rows(values_only=True))
        h = _find_header(rows, 'ruleset_codi')
        assert h is not None, 'No s\'ha trobat capçalera de GradingRules'
        rules_data = []
        for r in rows[h+1:]:
            if not r or not r[0] or not r[1]:
                # Saltar files de secció (pom_code None) i files buides.
                continue
            rules_data.append({
                'ruleset_codi': _str(r[0]),
                'pom_code': _str(r[1]),
                'logica': _str(r[2]) or 'LINEAR',
                'increment': _to_float(r[3]),
                'increment_above_xl': _to_float(r[4]) if r[4] not in (None, '', '—') else None,
                'actiu': _parse_bool(r[5], default=True),
                'notes': '',
            })

        # ── Sizing_Profiles ─────────────────────────────────────────────
        ws = wb['Sizing_Profiles']
        rows = list(ws.iter_rows(values_only=True))
        h = _find_header(rows, 'id')
        sp_data = []
        for r in rows[h+1:]:
            if not r or not r[0] or not _str(r[0]).isdigit():
                continue
            sp_data.append({
                'target': _str(r[1]),
                'garment_group': _str(r[2]),
                'construction': _str(r[3]),
                'fit_type': _str(r[4]),
                'size_system': _str(r[5]),
                'grading_ruleset': _str(r[6]),
                'is_default': _parse_bool(r[7]),
                'notes': _str(r[8]),
            })

        self.stdout.write(
            f'  Excel: GarmentPOMMap={len(gpm_data)}, '
            f'RuleSets={len(rs_data)}, Rules={len(rules_data)}, '
            f'SizingProfiles={len(sp_data)}'
        )

        # ── Mapa rs_codi → size_system_codi (des dels SizingProfiles) ──
        rs_size_system = {}
        rs_target = {}
        for sp in sp_data:
            rsk = sp['grading_ruleset']
            if rsk and rsk not in rs_size_system:
                rs_size_system[rsk] = sp['size_system']
                rs_target[rsk] = sp['target']

        # ── Operacions al tenant ────────────────────────────────────────
        with schema_context(tenant):
            from fhort.pom.models import (
                POMGlobal, POMMaster, POMCategory,
                GarmentPOMMap, GarmentType, GarmentGroup,
                GradingRuleSet, GradingRule,
                SizingProfile, SizeSystem, SizeDefinition,
                Target, FitType, ConstructionType,
            )

            with transaction.atomic():
                # =============================================
                # PAS 0 · Neteja en ordre invers de dependencia
                # (idempotència: l'ordre Pas A→D borraria POMMaster
                # primer, però GradingRule/GarmentPOMMap el PROTECT-en
                # després d'una primera execució)
                # =============================================
                self.stdout.write(self.style.WARNING('Pas 0 · Neteja prèvia'))
                from fhort.pom.models import (
                    GradingException as _GradingException,
                    ClientMesuraPerfil as _CMP,
                )
                for label, model in [
                    ('GradingException', _GradingException),
                    ('GradingRule', GradingRule),
                    ('SizingProfile', SizingProfile),
                    ('GradingRuleSet', GradingRuleSet),
                    ('ClientMesuraPerfil', _CMP),
                    ('GarmentPOMMap', GarmentPOMMap),
                    ('POMMaster', POMMaster),
                ]:
                    n, _ = model.objects.all().delete()
                    if n:
                        self.stdout.write(f'  {label}: {n} esborrats')

                # =============================================
                # PAS A · POMMaster (1 per POMGlobal)
                # =============================================
                self.stdout.write(self.style.WARNING('\nPas A · POMMaster'))
                # POMCategory: agafem només les "noves" (codi == nom de categoria
                # que coincideix amb POMGlobal.categoria string). Si hi ha
                # duplicat (CAT-UB + Upper body), preferim la que té display_order > 0
                # i actiu.
                cat_map = {}
                for c in POMCategory.objects.filter(actiu=True).order_by('-display_order'):
                    cat_map.setdefault(c.codi, c)

                masters = []
                cat_missing = set()
                for pg in POMGlobal.objects.all().order_by('codi'):
                    cat = cat_map.get(pg.categoria)
                    if pg.categoria and not cat:
                        cat_missing.add(pg.categoria)
                    masters.append(POMMaster(
                        pom_global=pg,
                        codi_client=pg.abbreviation or pg.codi,
                        nom_client=pg.nom_en,
                        actiu=pg.actiu,
                        categoria=cat,
                        notes=pg.notes or '',
                    ))
                POMMaster.objects.bulk_create(masters)
                total_pm = POMMaster.objects.count()
                self.stdout.write(f'  POMMaster creats: {total_pm}')
                if cat_missing:
                    self.stdout.write(self.style.WARNING(
                        f'  Categories sense POMCategory match: {sorted(cat_missing)}'
                    ))

                # =============================================
                # PAS B · GarmentPOMMap
                # =============================================
                self.stdout.write(self.style.WARNING('\nPas B · GarmentPOMMap'))
                GarmentPOMMap.objects.all().delete()
                gt_map = {gt.codi_client: gt for gt in GarmentType.objects.all()}
                # Map POMMaster per POMGlobal.codi (POM-001, POM-002, ...)
                pm_by_pgcodi = {
                    pm.pom_global.codi: pm
                    for pm in POMMaster.objects.select_related('pom_global').all()
                    if pm.pom_global_id
                }
                creats_b = 0
                saltats_b = []
                maps = []
                for d in gpm_data:
                    gt = gt_map.get(d['garment_type_codi'])
                    pm = pm_by_pgcodi.get(d['pom_code'])
                    if not gt:
                        saltats_b.append(f"GT no trobat: {d['garment_type_codi']}")
                        continue
                    if not pm:
                        saltats_b.append(f"POM no trobat: {d['pom_code']}")
                        continue
                    maps.append(GarmentPOMMap(
                        garment_type=gt, pom=pm,
                        ordre=d['ordre'],
                        is_key=d['is_key'],
                        obligatori=d['obligatori'],
                    ))
                    creats_b += 1
                GarmentPOMMap.objects.bulk_create(maps, ignore_conflicts=True)
                self.stdout.write(f'  GarmentPOMMap creats: {creats_b}')
                if saltats_b:
                    self.stdout.write(f'  Saltats: {len(saltats_b)} (mostra: {saltats_b[:5]})')

                # =============================================
                # PAS C · GradingRuleSet + GradingRule
                # =============================================
                self.stdout.write(self.style.WARNING('\nPas C · GradingRuleSet + GradingRule'))
                GradingRule.objects.all().delete()
                GradingRuleSet.objects.all().delete()

                target_map = {t.codi: t for t in Target.objects.all()}
                fit_map = {f.codi: f for f in FitType.objects.all()}
                constr_map = {c.codi: c for c in ConstructionType.objects.all()}
                ss_map = {ss.codi: ss for ss in SizeSystem.objects.all()}

                rs_map = {}
                rs_base_size = {}  # ruleset_codi → 'M'/'S'/...
                rs_size_system_obj = {}
                for d in rs_data:
                    ss_codi = rs_size_system.get(d['codi_sistema'], '')
                    ss = ss_map.get(ss_codi)
                    rs = GradingRuleSet.objects.create(
                        nom=d['nom_en'],
                        codi_sistema=d['codi_sistema'],
                        target=target_map.get(d['target']),
                        construction=constr_map.get(d['construction']),
                        fit_type=fit_map.get(d['fit_type']),
                        size_system=ss,
                        is_system_default=d['is_system_default'],
                        version_number=1,
                        actiu=True,
                    )
                    rs_map[d['codi_sistema']] = rs
                    rs_base_size[d['codi_sistema']] = d['base_size']
                    rs_size_system_obj[d['codi_sistema']] = ss
                self.stdout.write(f'  GradingRuleSets creats: {len(rs_map)}')

                # Per a cada rule: resoldre talla_base via (size_system de rs, etiqueta = base_size)
                # Fallback: primer SizeDefinition disponible globalment
                pm_by_codi_or_pgcodi = {}
                for pm in POMMaster.objects.select_related('pom_global').all():
                    pm_by_codi_or_pgcodi[pm.codi_client] = pm
                    if pm.pom_global_id:
                        pm_by_codi_or_pgcodi[pm.pom_global.codi] = pm

                fallback_sd = SizeDefinition.objects.first()
                if not fallback_sd:
                    raise RuntimeError('No hi ha cap SizeDefinition al tenant!')

                # Cache (size_system_id, etiqueta) → SizeDefinition
                sd_cache = {}
                for sd in SizeDefinition.objects.select_related('size_system').all():
                    sd_cache[(sd.size_system_id, sd.etiqueta)] = sd

                rules_objs = []
                saltats_r = []
                for d in rules_data:
                    rs = rs_map.get(d['ruleset_codi'])
                    pm = pm_by_codi_or_pgcodi.get(d['pom_code'])
                    if not rs:
                        saltats_r.append(f"RS no trobat: {d['ruleset_codi']}")
                        continue
                    if not pm:
                        saltats_r.append(f"POM no trobat: {d['pom_code']}")
                        continue
                    ss_id = rs.size_system_id
                    base_label = rs_base_size.get(d['ruleset_codi'], '')
                    sd = sd_cache.get((ss_id, base_label)) if ss_id else None
                    if not sd:
                        sd = fallback_sd
                    valors_step = None
                    if d['increment_above_xl'] is not None:
                        valors_step = {'above_xl': d['increment_above_xl']}
                    rules_objs.append(GradingRule(
                        rule_set=rs,
                        pom=pm,
                        talla_base=sd,
                        logica=d['logica'],
                        valor_base=0,
                        increment=d['increment'],
                        valors_step=valors_step,
                        actiu=d['actiu'],
                    ))
                GradingRule.objects.bulk_create(rules_objs, ignore_conflicts=True)
                self.stdout.write(f'  GradingRules creades: {GradingRule.objects.count()}')
                if saltats_r:
                    self.stdout.write(f'  Saltats: {len(saltats_r)} (mostra: {saltats_r[:5]})')

                # =============================================
                # PAS D · SizingProfile (expansió per grup)
                # =============================================
                self.stdout.write(self.style.WARNING('\nPas D · SizingProfile'))
                SizingProfile.objects.all().delete()
                # GarmentType per grup
                gt_by_group = {}
                for gt in GarmentType.objects.all():
                    gt_by_group.setdefault(gt.grup, []).append(gt)

                profiles = []
                saltats_p = []
                expandits = 0
                for d in sp_data:
                    rs = rs_map.get(d['grading_ruleset'])
                    ss = ss_map.get(d['size_system'])
                    target = target_map.get(d['target'])
                    constr = constr_map.get(d['construction'])
                    fit = fit_map.get(d['fit_type'])
                    grup_gts = gt_by_group.get(d['garment_group'], [])
                    if not rs:
                        saltats_p.append(f"RS no trobat: {d['grading_ruleset']}")
                        continue
                    if not target or not constr or not fit or not ss:
                        saltats_p.append(
                            f"Lookup falla per {d['grading_ruleset']}: "
                            f"target={bool(target)} constr={bool(constr)} "
                            f"fit={bool(fit)} ss={bool(ss)}"
                        )
                        continue
                    if not grup_gts:
                        saltats_p.append(f"Cap GarmentType al grup: {d['garment_group']}")
                        continue
                    for gt in grup_gts:
                        profiles.append(SizingProfile(
                            target=target,
                            garment_type=gt,
                            construction=constr,
                            fit_type=fit,
                            size_system=ss,
                            grading_rule_set=rs,
                            is_default=d['is_default'],
                            version=1,
                            notes=d['notes'],
                        ))
                        expandits += 1
                # Pot haver-hi duplicats per (target, gt, constr, fit) — fem dedup
                seen = set()
                unique_profiles = []
                for p in profiles:
                    k = (p.target_id, p.garment_type_id, p.construction_id, p.fit_type_id)
                    if k in seen:
                        continue
                    seen.add(k)
                    unique_profiles.append(p)
                SizingProfile.objects.bulk_create(unique_profiles)
                self.stdout.write(
                    f'  SizingProfiles creats: {len(unique_profiles)} '
                    f'(expandits={expandits}, deduplicats={expandits - len(unique_profiles)})'
                )
                if saltats_p:
                    self.stdout.write(f'  Saltats: {len(saltats_p)} (mostra: {saltats_p[:5]})')

        # ── Resum final ─────────────────────────────────────────────────
        with schema_context(tenant):
            from fhort.pom.models import (
                POMMaster, GarmentPOMMap, GradingRuleSet,
                GradingRule, SizingProfile,
            )
            self.stdout.write(self.style.SUCCESS(
                f'\n✓ Reseed tenant {tenant} complet:'
                f'\n  POMMaster:       {POMMaster.objects.count()}'
                f'\n  GarmentPOMMap:   {GarmentPOMMap.objects.count()}'
                f'\n  GradingRuleSet:  {GradingRuleSet.objects.count()}'
                f'\n  GradingRule:     {GradingRule.objects.count()}'
                f'\n  SizingProfile:   {SizingProfile.objects.count()}'
            ))
