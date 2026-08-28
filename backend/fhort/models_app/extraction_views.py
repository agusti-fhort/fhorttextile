# fhort/models_app/extraction_views.py
import base64 as _base64
import datetime as _dt
import io as _io
import logging as _logging
import re as _re

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status

from fhort.pom.models import MeasurementLayer
from fhort.pom.size_labels import canonical_size_label
from fhort.pom.grading_utils import normalitza_cm
from fhort.models_app.extraction_utils import registra_us_ia


# ═══════════════════════════════════════════════════════════════════════════════════════
# SET-2/T8 · LA PEÇA DE DESTÍ D'UN IMPORT — EL PUNT ÚNIC DEL PIPELINE
#
# Decisió Agus (Patró C): **un import = una prenda**, i s'inicia DES DE LA PEÇA. El garment
# es dedueix del context i no es pregunta mai; viu a `ImportSession.garment` des de la
# iniciació i totes les vistes del pipeline el llegeixen d'AQUÍ, mai del cos de la petició.
#
# ⚠️ I EL SEU VALOR EFECTIU NO ES RE-DERIVA ENLLOC. Run, talla base, sistema i joc de regles
# d'una peça són `garment.X or model.X` amb `is None` per predicat, i això viu en UN sol
# mòdul (`services_garment.valor_efectiu`) des de T2-bis. Aquestes dues funcions són la
# porta d'aquest fitxer cap allà: qui necessiti l'eix o el valor efectiu d'un import passa
# per aquí i no escriu cap `or` propi.
# ═══════════════════════════════════════════════════════════════════════════════════════

def _peca_de(session):
    """La prenda de destí d'una sessió d'import: `ModelGarment` o `None` (= la MARE).

    `None` no és un error tolerat: la mare ÉS el model i no té fila (D3). Una sessió amb un
    codi que ja no existeix (peça esborrada enmig d'un import) també torna `None` i escriu a
    la mare —que és el comportament d'abans d'aquest tram— en comptes de petar a mig camí.
    """
    from fhort.models_app.models import ModelGarment
    codi = (getattr(session, 'garment', '') or '').strip()
    if not codi or not session.model_id:
        return None
    return ModelGarment.objects.filter(model_id=session.model_id, codi=codi).first()


def _efectiu(session, camp):
    """El valor EFECTIU d'un camp heretable per a la peça de destí de la sessió."""
    from fhort.models_app.services_garment import valor_efectiu
    return valor_efectiu(session.model, _peca_de(session), camp)


def _efectiu_de_peca(session, codi, camp):
    """El valor EFECTIU d'un camp heretable per a UNA peça qualsevol d'aquest import.

    SET-2/T8-ter — `_efectiu` pregunta per la peça de LA SESSIÓ, i amb el garment a la fila la
    sessió ja no té una sola peça de destí. Aquest germà pren el codi com a argument.

    No duplica cap regla d'herència: totes dues passen per `valor_efectiu`, que segueix sent el
    punt únic (la llei de T2-bis és que `garment.X or model.X` viu en UN sol lloc).
    """
    from fhort.models_app.models import ModelGarment
    from fhort.models_app.services_garment import valor_efectiu
    codi = (codi or '').strip()
    peca = (ModelGarment.objects.filter(model_id=session.model_id, codi=codi).first()
            if (codi and session.model_id) else None)
    return valor_efectiu(session.model, peca, camp)


def _peces_amb_metrica_propia(session, codis):
    """Les peces de `codis` que declaren run o talla base PRÒPIS (override no NULL).

    ── SET-2/T8-ter · PER QUÈ EL PAS 1 ES RESOL CONTRA LA MARE (decisió Agus, §3) ──────────
    El pas 1 aparella les columnes del DOCUMENT amb les talles del model, i el document és UN:
    porta un sol joc de columnes per a totes les peces que hi surten. Un aparellament per peça
    voldria dir un pas 1 per peça, que és un wizard diferent i no el que aquest tram construeix.

    La referència neutra, doncs, és LA MARE — el model mateix (D3), que és de qui les peces
    hereten mentre no diguin el contrari. I com que NULL vol dir «hereta», el cas normal és que
    totes les peces hi coincideixin: el 100% del corpus d'avui (la `02 Short` del 1323 té
    `base_size_label=None`).

    El cas rar és que una peça declari metrica pròpia. Llavors el pas 1 NO pot parlar per ella i
    **s'ha de DIR** en comptes de decidir en silenci: aparellar contra la mare i escriure a una
    peça amb un altre run seria fabricar un aparellament que ningú ha validat. Es torna la
    llista perquè el wizard l'avisi; no barra, perquè qui sap si el document parla d'aquella
    peça és la persona que reparteix les files.
    """
    from fhort.models_app.models import ModelGarment
    if not session.model_id:
        return []
    codis_reals = [c for c in {(c or '').strip() for c in (codis or [])} if c]
    if not codis_reals:
        return []
    fora = []
    for peca in ModelGarment.objects.filter(model_id=session.model_id, codi__in=codis_reals):
        propis = [camp for camp in ('size_run_model', 'base_size_label', 'size_system')
                  if getattr(peca, camp if camp != 'size_system' else 'size_system_id') is not None]
        if propis:
            fora.append({'garment': peca.codi, 'garment_nom': peca.nom or peca.codi,
                         'camps': propis})
    return fora


def _nom_de_peca(model, codi):
    """Com s'anomena la peça `codi` d'aquest model, per DIR-LA a l'usuari.

    SET-2/T8-ter — germà de `_nom_de_la_peca`, que pregunta per la peça de la SESSIÓ. Amb el
    garment a la fila cal poder anomenar-ne una qualsevol: una llista de poda que barreja files
    de dues peces ha de dir de quina és cadascuna, i `_nom_de_la_peca` no ho sap fer perquè
    només coneix la de la sessió.
    """
    from fhort.models_app.models import ModelGarment
    codi = (codi or '').strip()
    if not codi:
        return (getattr(model, 'nom_prenda', '') or getattr(model, 'codi_intern', '') or '') if model else ''
    peca = ModelGarment.objects.filter(model=model, codi=codi).first() if model else None
    return (peca.nom or peca.codi) if peca is not None else codi


def _nom_de_la_peca(session):
    """Com s'anomena la prenda de destí, per DIR-LA a l'usuari.

    L'avís del multi-prenda ha de dir el NOM («s'importarà tot a Llaçada»), no una
    genèrica: un avís que no diu on va la feina obliga a anar-ho a comprovar. Si ningú ha
    batejat la peça encara, el codi és el millor nom que en tenim i és millor que el silenci.
    """
    peca = _peca_de(session)
    if peca is not None:
        return peca.nom or peca.codi
    model = session.model
    return (getattr(model, 'nom_prenda', '') or getattr(model, 'codi_intern', '') or '') if model else ''


def normalize_size_run(raw):
    """Convert any size_run format to 'XXS·XS·S·M·L·XL'."""
    if not raw:
        return ''
    if isinstance(raw, list):
        sizes = [str(s).strip() for s in raw if str(s).strip()]
    elif isinstance(raw, str):
        # Can be "['XXS', 'XS', 'S']" or "XXS,XS,S" or "XXS XS S"
        sizes = _re.findall(r'[A-Z0-9]+', raw.upper())
        # Filter out tokens that do not look like sizes
        sizes = [s for s in sizes if 1 <= len(s) <= 5]
    else:
        return ''
    return '·'.join(sizes)


def parse_any(raw):
    """Normalize the year to a 4-digit integer."""
    if not raw:
        return _dt.date.today().year
    try:
        y = int(str(raw).strip())
        if y < 100:
            y += 2000
        return y
    except (ValueError, TypeError):
        return _dt.date.today().year


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_model_view(request, model_id):
    """
    DELETE /api/v1/models/<id>/delete/
    Delete the model and all associated data in cascade:
    BaseMeasurements, SizeFittings, GradingVersions, GradedSpecs,
    ModelFitxers (physical files included), POMAlerts, ModelTasques.
    """
    from django.core.files.storage import default_storage
    from fhort.models_app.models import Model, ModelFitxer

    try:
        model = Model.objects.get(id=model_id)
    except Model.DoesNotExist:
        return Response({'error': 'Model no trobat'}, status=404)

    nom = model.nom_prenda
    codi = model.codi_intern

    # Delete associated physical files (do not block if it fails)
    try:
        for fitxer in ModelFitxer.objects.filter(model=model):
            if fitxer.fitxer and default_storage.exists(fitxer.fitxer.name):
                default_storage.delete(fitxer.fitxer.name)
    except Exception:
        pass

    # Delete the model (DB cascade)
    model.delete()

    return Response({
        'deleted': True,
        'model_id': model_id,
        'nom': nom,
        'codi': codi,
        'message': f'Model "{nom}" ({codi}) esborrat correctament.',
    })


# =====================================================================
# F2.2/F2.3 — Importació guiada per sessió (ImportSession)
# Crida 1: cribratge barat (tipologia, nº models, run de talles).
# =====================================================================

CRIBRATGE_MODEL = 'claude-opus-4-7'

# RETURN ONLY VALID JSON — cap prosa, cap markdown. Visió barata, sense thinking.
CRIBRATGE_PROMPT = """You are a fast triage system for fashion tech-sheet documents.
Look at the document and return ONLY a single valid JSON object. No prose, no markdown fences.

Detect, at a glance (do NOT extract measurements):
- How many INDEPENDENT garment models/styles the document contains (distinct style names or codes).
- The garment typology, in English (dress, trousers, shirt, skirt, jacket, pyjama...).
- The target gender/age segment.
- The size-run labels exactly as printed, in order.
- Which size system those labels belong to.

Return EXACTLY this shape:
{
  "num_models": <int>,
  "models_detectats": [{"nom": "<style name or code>", "pagina": <int>, "descripcio": "<short>"}],
  "tipologia_detectada": "<dress|trousers|shirt|skirt|...>",
  "genere_detectat": "<woman|man|unisex|baby|kids>",
  "run_talles_document": ["<label1>", "<label2>", "..."],
  "sistema_talles": "<letters|age_months|age_years|numeric|height_cm|unknown>"
}

Rules:
- num_models counts distinct styles/patterns. Two patterns on the same page = 2.
- Use the EXACT size labels printed in the document, preserving their order.
- letters = XS/S/M/L..., numeric = 34/36/38..., age_months = 0M/3M/6M..., age_years = 6Y/8Y...,
  height_cm = 50/56/62... (cm body height for baby/kids). If unsure, "unknown".
- Output ONLY the JSON object, nothing else."""


def _excel_to_text(file_bytes: bytes) -> str:
    """Converteix un .xlsx/.xls a text tabulat perquè la IA en llegeixi el contingut."""
    import openpyxl
    wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True, read_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f'### Full: {ws.title}')
        for row in ws.iter_rows(values_only=True):
            cells = ['' if c is None else str(c) for c in row]
            if any(cells):
                lines.append('\t'.join(cells))
    wb.close()
    return '\n'.join(lines)


# ═══════════════ Parser determinista d'Excel — perfil "spec sheet" ═══════════════
# QA-S8 · FIX C (DIAGNOSI_QA_S8_IMPORT §D1c). El parser anterior cercava el codi a la
# COLUMNA A i abdicava sempre que la taula no hi comencés — que és el cas de totes les
# fitxes reals que tenim (la taula Brownie viu de la B a la H i la columna A és buida de
# dalt a baix). Aquest perfil ancora la taula pel CONTINGUT de la capçalera i mapa les
# columnes per ETIQUETA, no per índex.

#: Etiquetes que ANCOREN la capçalera. La columna on viuen no importa: és el contingut
#: qui mana (D1c·1). Calen totes dues famílies a la mateixa fila per considerar-la capçalera.
_ETIQ_CODI = {'CODE', 'CODI', 'POM', 'POM CODE'}
_ETIQ_DESC = {'DESCRIPTION', 'DESCRIPCIO', 'DESCRIPCIÓ', 'DESC', 'ENGLISH'}

#: Columnes de SERVEI: tenen etiqueta però NO són talles (D1c·6). Sense aquesta llista,
#: 'SAMPLE', 'ADJUSTMENTS' i 'COMMENTS' entrarien com si fossin tres talles més.
_ETIQ_SERVEI = _ETIQ_CODI | _ETIQ_DESC | {
    'GRADING', 'DIM', 'SAMPLE', 'SAMPLE SIZE', 'ADJUSTMENTS', 'ADJUSTMENT',
    'COMMENTS', 'COMMENT', 'NOTES', 'NOTE', 'REMARKS', 'MEASUREMENT', 'MEASUREMENTS',
}

#: Una etiqueta de TALLA: lletres (S/M/L/XL/XXL/2XL), numèrica (34/36), edat (6M/8Y) o T2.
_RE_TALLA = _re.compile(r'^(?:X*[SL]|M|\d+X[SL]|\d{1,3}(?:[.,]\d{1,2})?|\d{1,2}\s*[MYA]|T\d{1,2})$',
                        _re.I)

#: Un CODI de POM: curt i sense espais interns (A, D, G1, EK2, U2, LZ1, SF). El que no hi
#: encaixa i seu a la columna del codi és un BANNER ('SKETCH WITH CODES'), no un POM (D1c·5).
_RE_CODI = _re.compile(r'^[A-Za-z0-9][A-Za-z0-9.\-/]{0,7}$')

#: Metadades del bloc superior (B2:B7 a la fitxa Brownie) → claus del `header`, les MATEIXES
#: que retorna la via Opus (extraction_prompt.py), perquè els dos camins parlin igual.
_META_HEADER = {
    'BRAND': 'brand',
    'NAME STYLE': 'style_name',
    'STYLE NAME': 'style_name',
    'STYLE': 'style_name',
    'COLOR': 'color',
    'COLOUR': 'color',
    'SEASON': 'season',
    'DATE': 'date',
    'STYLE NO': 'style_reference',
    'REF': 'style_reference',
    'REFERENCE': 'style_reference',
}

#: Files amb codi I valor a la talla base que calen per donar la taula per ENTESA. Per sota
#: d'això el parser abdica: tres files coherents són la prova mínima que hem trobat una taula
#: de mesures de debò i no un bloc qualsevol amb text a sobre.
_MIN_FILES_ENTESA = 3


def _num(v):
    """Valor numèric d'una cel·la d'xlsx, ja normalitzat a 0,1 mm; None altrament.

    D3 · porta d'entrada del camí XLSX. `openpyxl` retorna el float binari cru de la
    cel·la (16.749999999999999 per un 16,75 escrit), i aquest era el valor que arribava
    fins a `detect_grading`. Delega a `normalitza_cm` perquè els tres camins d'entrada
    normalitzin al MATEIX lloc i amb la mateixa regla.
    """
    return normalitza_cm(v)


def _etiqueta(v):
    """Text d'una cel·la de capçalera, normalitzat per COMPARAR (majúscules, espais collapsats)."""
    if v is None:
        return ''
    return ' '.join(str(v).split()).upper()


def _valor_meta(v):
    """Valor d'una cel·la del bloc de metadades, com a text net."""
    if v is None:
        return ''
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    return str(v).strip()


def _files_banner(ws, ci_codi):
    """Files (1-indexades) tapades per un bloc FUSIONAT ample — sketch, comentaris, peus.

    D1c·5 i bandera 4: la fitxa Brownie té `B39:H39` (el rètol 'SKETCH WITH CODES') i tres
    blocs grans (`B40:H67`, `B70:H97`, `B100:H127`). Un parser que no els talli recorre fins
    a la fila 127 i s'empassa el sketch com si fossin POMs. Un merge d'UNA columna (les
    capçaleres verticals `B9:B10`) NO és un banner: només compten els que travessen la taula.
    """
    banner = set()
    for rang in ws.merged_cells.ranges:
        ample = rang.max_col - rang.min_col + 1
        if ample >= 3 and rang.min_col <= ci_codi + 1 <= rang.max_col:
            banner.update(range(rang.min_row, rang.max_row + 1))
    return banner


def _parse_excel_poms(file_bytes: bytes, base_hint=None, run_hint=None, full_seleccionat=None):
    """Parse determinista d'una fitxa Excel de POMs (via ràpida del wizard).

    Retorna `(poms, talles, meta)`:
      poms  = [{'codi_fitxa', 'descripcio', 'dim', 'values': {talla: float}, 'tol_*', 'seccio'}]
      talles = [etiquetes de talla, en ordre de columna]
      meta  = {'header', 'base_size', 'full', 'n_files_amb_codi', 'motiu', 'fulls',
               'full_seleccionat_ignorat'}

    ⚠️ **PORTA D'ABDICACIÓ — la llei d'aquest parser** (DIAGNOSI_QA_S8_IMPORT, risc de D1c).
    El contracte del wizard és "si el parser no en treu res, cau a la IA". Un parser més
    llest però EQUIVOCAT ja no cau: substitueix la IA **en silenci** i escriu dades dolentes.
    Això és pitjor que el defecte que arregla. Per tant aquesta funció només retorna files
    quan pot DEMOSTRAR que ha entès la taula:

      1. capçalera ancorada per CONTINGUT — una fila amb una etiqueta de CODI *i* una de
         DESCRIPCIÓ (a la columna que sigui);
      2. columna de TALLA BASE identificada — el `SAMPLE SIZE` de les metadades (o el
         `base_hint` del model) ha de correspondre a una columna de talla REAL; i
      3. almenys `_MIN_FILES_ENTESA` files amb codi *i* valor numèric a aquella talla base.

    Si qualsevol de les tres falla, retorna `([], [], meta)` amb el motiu, i el caller cau
    a la IA **com fins ara**. La prova es fa full per full: en un llibre de diverses pestanyes
    guanya la primera que la passa (a la fitxa Rosalia això descarta 'PROTO COMMENTS' —que té
    la columna de la talla base BUIDA— i tria 'RECTI 1 COMMENTS', que és on hi ha les mesures).

    **F5 · TOTS ELS FULLS S'AVALUEN** (DIAGNOSI_MULTIPECA_DALIA, taula final §15). Fins ara la
    funció retornava DINS del bucle al primer full que passava la porta i els següents no es
    miraven mai: un llibre amb una peça per pestanya perdia els fulls 2..N sencers i **sense
    cap avís de contingut** — l'únic que arribava era l'avís genèric de multi-model del
    cribratge. Ara es recorre el llibre sencer i `meta['fulls']` porta l'informe
    `[{nom, n_files_amb_codi, passa_porta}]` de cada pestanya.

    El COMPORTAMENT PER DEFECTE NO CANVIA: sense `full_seleccionat` guanya el primer full que
    passa la porta, exactament com abans. `full_seleccionat` (el que el tècnic tria al wizard)
    el força; si el full triat no existeix o no passa la porta, es cau al default i el nom
    ignorat queda a `meta['full_seleccionat_ignorat']` en comptes d'abdicar en silenci.

    Cost acceptat: ara es llegeixen totes les pestanyes i no només fins a la primera bona. És
    el preu de poder DIR què hi ha al llibre; les fitxes reals en tenen dues o tres.

    `meta['n_files_amb_codi']` és el nombre de files de POM que el document conté de debò.
    Serveix per al Fix D encara que s'abdiqui: si la IA en retorna menys, algú ho ha de dir
    (la fitxa del Tate té 26 POMs i la IA en va perdre un, 'JJ', sense cap avís).
    """
    import openpyxl

    meta = {'header': {}, 'base_size': None, 'full': None,
            'n_files_amb_codi': 0, 'motiu': 'cap full amb capçalera de POMs reconeixible',
            'fulls': [], 'full_seleccionat_ignorat': None}
    run_canonic = {canonical_size_label(t) for t in (run_hint or []) if str(t).strip()}

    #: Fulls que HAN passat la porta, en ordre de llibre: títol → (poms, talles, header, base).
    #: Un dict i no una llista perquè la tria per nom sigui directa; conserva l'ordre d'inserció.
    candidats = {}

    def _apunta(nom, n_files, passa):
        meta['fulls'].append({'nom': nom, 'n_files_amb_codi': n_files, 'passa_porta': passa})

    # read_only=False: els merges (D1c·3 i ·5) NO es poblen en mode read_only, i sense ells
    # no es poden ni compondre les capçaleres dobles ni tallar els blocs del sketch.
    wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True)
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))

            # ── 1. Ancorar la capçalera pel CONTINGUT (D1c·1): la fila que porta alhora una
            # etiqueta de codi i una de descripció. La columna on caiguin és la que sigui.
            header_idx = ci_codi = ci_desc = None
            for idx, row in enumerate(rows):
                codi_ci = desc_ci = None
                for ci, cell in enumerate(row):
                    et = _etiqueta(cell)
                    if codi_ci is None and et in _ETIQ_CODI:
                        codi_ci = ci
                    elif desc_ci is None and et in _ETIQ_DESC:
                        desc_ci = ci
                if codi_ci is not None and desc_ci is not None:
                    header_idx, ci_codi, ci_desc = idx, codi_ci, desc_ci
                    break
            if header_idx is None:
                _apunta(ws.title, 0, False)
                continue

            # ── 2. Capçalera DOBLE amb merges (D1c·3). Els merges verticals (B9:B10) només
            # porten valor a la cel·la de dalt; les etiquetes NO fusionades de la segona fila
            # (C10='ENGLISH', F10='RECTI 1') són pròpies. Regla: mana la primera fila, la
            # segona només omple els buits.
            etiquetes = {ci: _etiqueta(c) for ci, c in enumerate(rows[header_idx])}
            if header_idx + 1 < len(rows):
                seguent = rows[header_idx + 1]
                if not _etiqueta(seguent[ci_codi] if ci_codi < len(seguent) else None):
                    for ci, cell in enumerate(seguent):
                        if not etiquetes.get(ci):
                            etiquetes[ci] = _etiqueta(cell)

            # ── 3. Mapa de columnes per ETIQUETA (D1c·2 i ·6). Una columna és de TALLA si té
            # etiqueta, no és de servei, no és de tolerància, i sembla una talla (o és al run
            # del model). Així 'SAMPLE', 'ADJUSTMENTS' i 'COMMENTS' es queden fora.
            size_cols, dim_ci = [], None
            tol_minus_ci = tol_plus_ci = tol_single_ci = None
            for ci, et in sorted(etiquetes.items()):
                if not et or ci in (ci_codi, ci_desc):
                    continue
                if 'TOL' in et:
                    # B2: les columnes de tolerància es capturen (tol_minus/tol_plus). Una sola
                    # columna 'Tol' sense signe → mateix valor als dos costats (simètrica).
                    if '-' in et or 'MIN' in et:
                        tol_minus_ci = ci
                    elif '+' in et or 'PLUS' in et or 'MAX' in et:
                        tol_plus_ci = ci
                    else:
                        tol_single_ci = ci
                    continue
                if et == 'DIM':
                    dim_ci = ci
                    continue
                if et in _ETIQ_SERVEI:
                    continue
                if _RE_TALLA.match(et) or canonical_size_label(et) in run_canonic:
                    # L'etiqueta que es desa és la del document tal com hi surt (l'etiqueta
                    # del tenant la posa el reconcile de W5); `et` és només per comparar.
                    crua = rows[header_idx][ci] if ci < len(rows[header_idx]) else None
                    size_cols.append((ci, str(crua).strip() if crua is not None else et))
            if not size_cols:
                meta['motiu'] = f"full '{ws.title}': cap columna de talla reconeguda"
                _apunta(ws.title, 0, False)
                continue

            # ── 4. Bloc de metadades (B2:B7) — el bonus barat de D1c. Etiqueta a la columna del
            # codi, valor a la de la descripció. D'aquí surt el SAMPLE SIZE, que és qui diu quina
            # és la TALLA BASE del document (D1c·6): mai "la primera columna".
            header_meta = {}
            sample_size = None
            for row in rows[:header_idx]:
                clau = _etiqueta(row[ci_codi] if ci_codi < len(row) else None)
                valor = _valor_meta(row[ci_desc] if ci_desc < len(row) else None)
                if not clau or not valor:
                    continue
                if clau == 'SAMPLE SIZE':
                    sample_size = valor
                elif clau in _META_HEADER:
                    header_meta[_META_HEADER[clau]] = valor

            # ── 5. La TALLA BASE. Si el document (o el model) la declara, ha de correspondre a
            # una columna de talla real: si no hi és, hem entès malament la taula → abdicar.
            # Si ningú no la declara, la base és la primera talla (contracte del parser antic).
            #
            # ⚠️ `base_label` entra com l'etiqueta de QUI LA DECLARA (el `SAMPLE SIZE` del
            # document o el `base_hint` del MODEL) i, un cop trobada la columna, passa a ser
            # l'etiqueta CRUA D'AQUELLA COLUMNA. No és cosmètica: `values` està indexat per
            # l'etiqueta crua del document (:394), i la porta de :419 compara per igualtat
            # literal de cadena. Amb `base_hint='00/01'` i la columna '0M-1M' —el cas real de
            # LOS-SS27-0834— la canonicalització SÍ que trobava la columna (les dues donen
            # '0/1'), però `base_label` es quedava com '00/01' i llavors `'00/01' in values`
            # era fals a cada fila → amb_base=0 → el parser abdicava i queia a la IA sobre un
            # xlsx que havia entès perfectament (DIAGNOSI_MULTIPECA_DALIA Q5).
            base_label = (sample_size or base_hint or '').strip()
            base_ci = None
            if base_label:
                canon = canonical_size_label(base_label)
                trobada = next(((ci, lbl) for ci, lbl in size_cols
                                if canonical_size_label(lbl) == canon), None)
                if trobada is None:
                    meta['motiu'] = (f"full '{ws.title}': la talla base '{base_label}' no té "
                                     f"columna a la taula")
                    _apunta(ws.title, 0, False)
                    continue
                base_ci, base_label = trobada
            else:
                base_ci, base_label = size_cols[0][0], size_cols[0][1]

            # ── 6. Files de dades. Tres menes de fila que NO són POMs:
            #   · SECCIÓ ('Bodice:', 'Cord:') → codi buit + descripció plena. SALTAR, mai `break`
            #     (D1c·4: el parser antic hi feia `break` i es quedava amb zero files).
            #   · BANNER ('SKETCH WITH CODES') i blocs fusionats → FI DE TAULA (D1c·5).
            #   · fila buida → saltar.
            banner = _files_banner(ws, ci_codi)

            def _cell(row, ci):
                return _num(row[ci]) if (ci is not None and ci < len(row)) else None

            #: R4 · l'última fila que ÉS un POM. El "fi de taula" no és una forma de fila: és una
            #: POSICIÓ — per sota d'aquí ja no hi ha taula. Amb la fitxa DALIA (PROD, tenant
            #: `los`) el rètol de peça viu a la columna del CODI i la seva fila va fusionada de
            #: banda a banda: les dues coses feien `break` a la PRIMERA secció, just sota la
            #: capçalera, i el full sencer es perdia (0 files, abdicació i caiguda a la IA).
            def _fila_de_pom(row):
                c = (str(row[ci_codi]).strip()
                     if (ci_codi < len(row) and row[ci_codi] is not None) else '')
                return (bool(c) and bool(_RE_CODI.match(c))
                        and any(_cell(row, ci) is not None for ci, _ in size_cols))

            ultima_pom = max((i for i in range(header_idx + 1, len(rows)) if _fila_de_pom(rows[i])),
                             default=header_idx)

            poms = []
            #: Secció vigent (F4). El text de les files de secció es LLEGIA i es llençava; ara
            #: es recorda i cada POM se n'endú una còpia. La convenció d'arrel per peça
            #: (01./02./03.) no es podia automatitzar sense això: no hi havia d'on treure la
            #: peça (DIAGNOSI_MULTIPECA_DALIA · Q3, taula final §14).
            #: Es reinicia a cada full: una secció no travessa pestanyes.
            seccio_vigent = None
            for idx in range(header_idx + 1, len(rows)):
                row = rows[idx]
                if (idx + 1) in banner and idx > ultima_pom:
                    break                         # bloc del sketch, peus: fi de taula de debò
                codi = str(row[ci_codi]).strip() if (ci_codi < len(row)
                                                     and row[ci_codi] is not None) else ''
                if not codi:
                    # Codi buit: secció, capçalera-2 o fila buida. La distingeix la DESCRIPCIÓ,
                    # no l'absència de valors: a la fitxa Rosalia la secció 'Chest piece:' (f23)
                    # porta zeros a les columnes de talla i la fila de soroll f22 en porta
                    # sense cap text. Amb text → és una secció; sense → segueix sent soroll.
                    text_seccio = (str(row[ci_desc]).strip()
                                   if (ci_desc < len(row) and row[ci_desc] is not None) else '')
                    if text_seccio:
                        seccio_vigent = text_seccio
                    continue
                if not _RE_CODI.match(codi):
                    if idx > ultima_pom:
                        break                     # rètol ('SKETCH WITH CODES') → fi de taula
                    seccio_vigent = codi          # títol de peça a la columna del codi (LOSAN)
                    continue
                desc = (str(row[ci_desc]).strip()
                        if (ci_desc < len(row) and row[ci_desc] is not None) else '')
                values = {}
                for ci, lbl in size_cols:
                    nv = _cell(row, ci)
                    if nv is not None:
                        values[lbl] = nv
                if not desc and not values:
                    continue                      # codi solt sense res: soroll, no un POM
                tm, tp = _cell(row, tol_minus_ci), _cell(row, tol_plus_ci)
                ts = _cell(row, tol_single_ci)
                if ts is not None:
                    tm = ts if tm is None else tm
                    tp = ts if tp is None else tp
                poms.append({
                    'codi_fitxa': codi,           # D1c·7: 'D ' → 'D' (strip)
                    'descripcio': desc,
                    'dim': _cell(row, dim_ci),
                    'values': values,
                    'tol_minus': tm,
                    'tol_plus': tp,
                    # F4 · secció d'origen. None quan el document no en té cap (la majoria de
                    # fitxes d'una sola peça): és DADA, no comportament — ningú no hi decideix
                    # res encara.
                    'seccio': seccio_vigent,
                })

            # ── 7. LA PORTA. ¿Podem demostrar que hem entès la taula? Files amb codi I valor a
            # la talla base. Si no arriben a _MIN_FILES_ENTESA, aquest full no és una taula de
            # mesures que sapiguem llegir: abdiquem i que la IA hi digui la seva.
            # El recompte de files sobreviu a l'abdicació a PROPÒSIT: si aquest full s'ha
            # entès prou per comptar-ne les files però no per fiar-se'n, la IA se n'ocuparà
            # — i el Fix D encara ha de poder dir "el document en tenia N i n'has tret menys".
            meta['n_files_amb_codi'] = max(meta['n_files_amb_codi'], len(poms))

            amb_base = sum(1 for p in poms if base_label in p['values'])
            passa = amb_base >= _MIN_FILES_ENTESA
            _apunta(ws.title, len(poms), passa)
            if not passa:
                meta['motiu'] = (f"full '{ws.title}': només {amb_base} fila(es) amb valor a la "
                                 f"talla base '{base_label}' (en calen {_MIN_FILES_ENTESA})")
                continue

            candidats[ws.title] = (poms, [lbl for _, lbl in size_cols], header_meta, base_label)
    finally:
        wb.close()

    # ── 8. LA TRIA. Per defecte, el primer full que ha passat la porta (comportament de
    # sempre). `full_seleccionat` el força; si no és triable, es diu i es cau al default.
    if not candidats:
        return [], [], meta
    tria = next(iter(candidats))
    if full_seleccionat:
        if full_seleccionat in candidats:
            tria = full_seleccionat
        else:
            meta['full_seleccionat_ignorat'] = full_seleccionat

    poms, talles, header_meta, base_label = candidats[tria]
    meta.update({
        'header': header_meta,
        'base_size': base_label,
        'full': tria,
        # Files de POM que el document conté DE DEBÒ (Fix D). Aquí és igual a len(poms): el
        # parser no en deixa caure cap — una fila sense valor a la talla base ('JJ') és un POM
        # legítim (BaseMeasurement.base_value_cm és null=True), no un descart.
        'n_files_amb_codi': len(poms),
        'motiu': None,
    })
    return poms, talles, meta


def _cribratge_content_block(file_bytes: bytes, filename: str, content_type: str) -> dict:
    """Bloc de contingut per a la API segons el tipus de fitxa origen (PDF/imatge/Excel)."""
    name = (filename or '').lower()
    ct = (content_type or '').lower()

    if ct == 'application/pdf' or name.endswith('.pdf'):
        return {
            'type': 'document',
            'source': {
                'type': 'base64',
                'media_type': 'application/pdf',
                'data': _base64.standard_b64encode(file_bytes).decode(),
            },
        }
    if name.endswith(('.xlsx', '.xls')) or 'spreadsheet' in ct or 'excel' in ct:
        text = _excel_to_text(file_bytes)
        return {'type': 'text', 'text': f'Contingut del full de càlcul (fitxa Excel):\n{text[:12000]}'}

    # Imatge (jpg/png/webp)
    if ct in ('image/jpeg', 'image/png', 'image/webp'):
        media = ct
    elif name.endswith('.png'):
        media = 'image/png'
    elif name.endswith('.webp'):
        media = 'image/webp'
    else:
        media = 'image/jpeg'
    return {
        'type': 'image',
        'source': {
            'type': 'base64',
            'media_type': media,
            'data': _base64.standard_b64encode(file_bytes).decode(),
        },
    }


def _cribratge_determinista(file_name, file_bytes, base_hint, run_model):
    """Cribratge SENSE IA per a un .xlsx que el parser determinista entén — o `None`.

    Decisió Agus 2026-07-22: «IA només quan el determinista no pot; un xlsx parsejable no ha
    de costar ni un cèntim de token.»

    El cribratge existeix per saber QUANTS models porta el document i QUIN run de talles té.
    Per a un xlsx que el parser entén, tots dos ja els sabem sense preguntar res: el parser
    només retorna files quan pot DEMOSTRAR que ha entès la taula (capçalera ancorada per
    contingut + columna de talla base identificada + mínim de files coherents), i les seves
    talles surten de les columnes reals, no d'una lectura. Fins ara el fitxer s'enviava a
    Opus SEMPRE, també quan tot seguit el parser el resoldria sol al pas 2: pagàvem una
    lectura de visió per confirmar una cosa que teníem a la mà.

    Els altres camps que retornava el cribratge (tipologia, gènere, `pot_continuar`) no els
    llegeix ningú al wizard —només consumeix `run_talles_document` i `num_models`—, o sigui
    que saltar la crida no li treu res.

    Retorna `None` (→ cau a Opus, com sempre) per a PDF, imatge, i xlsx on el parser abdica.
    Que el parser peti compta com abdicar: davant del dubte, IA.
    """
    if not (file_name or '').lower().endswith(('.xlsx', '.xls')):
        return None
    # SET-2/T8 — les pistes són les de la PEÇA DE DESTÍ (valors efectius), no les de la mare:
    # una peça amb run propi ha de fer llegir el document amb el SEU run, o el parser ancora
    # les columnes contra una escala que aquella prenda no fa servir.
    run_hint = [s.strip() for s in (run_model or '').replace(';', '·').split('·')
                if s.strip()]
    try:
        poms, talles, meta = _parse_excel_poms(
            file_bytes, base_hint=base_hint, run_hint=run_hint)
    except Exception:
        _logging.getLogger(__name__).exception(
            'Cribratge: el parser determinista ha petat; es cau a la IA')
        return None
    if not poms:
        return None
    return {
        'num_models': 1,               # el parser ha entès UNA taula de mesures
        'models_detectats': [],
        # Tipologia i gènere els decideix l'humà al pas següent; el cribratge només els
        # proposava, i el wizard no els llegeix.
        'tipologia_detectada': '',
        'genere_detectat': '',
        'run_talles_document': talles,
        'sistema_talles': 'unknown',
        'origen': 'parser_determinista',
        'n_files_amb_codi': meta.get('n_files_amb_codi') or len(poms),
    }


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_session_cribratge_view(request):
    """
    POST /api/v1/import-sessions/cribratge/
    multipart: document (fitxer), model_id, garment_type_item_code, garment

    Crida 1 — cribratge barat (visió, Opus, tokens baixos, sense thinking): detecta nº de
    models al document, tipologia, gènere i el run de talles. SEMPRE retorna resultats; el
    gating (bloqueig de talles, confirmació multi-model) és el pas F2.3.

    SET-2/T8 — `garment` és el CONTEXT de la peça des d'on s'ha obert l'import ('' = mare).
    És l'ÚNICA porta on entra: a partir d'aquí el pipeline el llegeix de la sessió.
    """
    import anthropic
    from django.conf import settings
    from django.core.files.base import ContentFile

    from fhort.accounts.models import UserProfile
    from fhort.models_app.models import ImportSession, Model, ModelGarment
    from fhort.models_app.extraction_utils import safe_json_parse
    from fhort.tasks.models import GarmentTypeItem

    file_obj = request.FILES.get('document')
    if not file_obj:
        return Response({'error': 'Cal adjuntar un fitxer (camp "document")'}, status=400)

    model_id = request.data.get('model_id')
    if not model_id:
        return Response({'error': 'Cal indicar model_id'}, status=400)
    try:
        model = Model.objects.get(id=model_id)
    except Model.DoesNotExist:
        return Response({'error': f'Model {model_id} no trobat'}, status=404)

    item_code = (request.data.get('garment_type_item_code') or '').strip()
    item = None
    if item_code:
        # Prefer the item already on the model if its code matches; else look up by code.
        if model.garment_type_item_id and model.garment_type_item.code == item_code:
            item = model.garment_type_item
        else:
            item = GarmentTypeItem.objects.filter(code=item_code).first()

    # ── SET-2/T8 · L'EIX DE LA PEÇA. Ve del CONTEXT (el contenidor des d'on s'ha premut
    # «Importar taula»), mai d'una pregunta del wizard. Un codi que no és cap peça d'aquest
    # model és un 400 i no un silenci: escriure a la mare quan el client ha dit '02' seria
    # exactament el dany que aquest tram tanca.
    garment = (request.data.get('garment') or '').strip()
    if garment and not ModelGarment.objects.filter(model=model, codi=garment).exists():
        return Response({'error': f"El model {model.id} no té cap prenda «{garment}»."},
                        status=400)

    profile = UserProfile.objects.filter(user=request.user).first()

    api_key = getattr(settings, 'ANTHROPIC_API_KEY', '')
    if not api_key:
        return Response({'error': 'ANTHROPIC_API_KEY no configurada al backend'}, status=500)

    # Crea la sessió i desa el document origen.
    session = ImportSession.objects.create(
        estat='CRIBRATGE', creat_per=profile, model=model, tipologia_confirmada=item,
        garment=garment,
    )
    file_bytes = file_obj.read()
    session.document.save(file_obj.name, ContentFile(file_bytes), save=True)

    # ── CRIBRATGE DETERMINISTA PRIMER (Agus 2026-07-22) ───────────────────────────────
    # «IA només quan el determinista no pot; un xlsx parsejable no ha de costar ni un
    # cèntim de token.»
    #
    # El cribratge existeix per saber QUANTS models porta el document i QUIN run de talles
    # té. Per a un .xlsx que el parser determinista entén, això ja ho sabem sense preguntar
    # res a ningú: el parser només retorna files quan pot DEMOSTRAR que ha entès la taula
    # (capçalera ancorada per contingut + columna de talla base + mínim de files coherents),
    # i les seves talles surten de les columnes reals, no d'una lectura.
    #
    # Fins ara el fitxer s'enviava a Opus SEMPRE, també quan tot seguit el parser el
    # resoldria sol al pas 2 — o sigui que pagàvem una lectura de visió per confirmar una
    # cosa que teníem a la mà.
    #
    # Els altres camps del cribratge (tipologia, gènere, `pot_continuar`) no els llegeix
    # ningú al wizard: només consumeix `run_talles_document` i `num_models`. Per això
    # saltar la crida no li treu res. PDF, imatge i xlsx on el parser abdica segueixen
    # passant per Opus exactament com abans.
    resultat = _cribratge_determinista(
        file_obj.name, file_bytes,
        _efectiu(session, 'base_size_label'), _efectiu(session, 'size_run_model'))
    cribratge_ia = resultat is None
    if not cribratge_ia:
        _logging.getLogger(__name__).info(
            f"Cribratge SENSE IA (sessió {session.token}): parser determinista, "
            f"{resultat['n_files_amb_codi']} POM(s), talles={resultat['run_talles_document']}")

    if cribratge_ia:
        content_block = _cribratge_content_block(file_bytes, file_obj.name, file_obj.content_type)

        try:
            client = anthropic.Anthropic(api_key=api_key)
            response = client.messages.create(
                model=CRIBRATGE_MODEL,
                max_tokens=900,
                system=CRIBRATGE_PROMPT,
                messages=[{'role': 'user', 'content': [content_block]}],
            )
            raw = ''.join(
                b.text for b in response.content if getattr(b, 'type', None) == 'text'
            ).strip()
            registra_us_ia(cami='cribratge', model_ia=CRIBRATGE_MODEL,
                           usage=getattr(response, 'usage', None),
                           import_session=session, model=model, created_by=profile)
        except Exception as e:
            _logging.getLogger(__name__).exception('Cribratge: error a la crida Claude')
            # Una crida que peta també s'ha pagat: es registra igual (ok=False).
            registra_us_ia(cami='cribratge', model_ia=CRIBRATGE_MODEL,
                           import_session=session, model=model, created_by=profile,
                           ok=False, error=str(e))
            return Response({'error': f'Error a la crida de cribratge: {e}', 'token': str(session.token)},
                            status=502)

        # Parse tolerant (Fase 1).
        try:
            resultat = safe_json_parse(raw)
        except ValueError as e:
            session.avisos = (session.avisos or []) + [f'Cribratge: JSON invàlid ({e})']
            session.save(update_fields=['avisos', 'actualitzat_at'])
            return Response({'error': f'Cribratge: resposta no parsejable ({e})',
                             'token': str(session.token), 'raw': raw[:500]}, status=422)

    num_models = resultat.get('num_models') or len(resultat.get('models_detectats') or []) or 0
    models_detectats = resultat.get('models_detectats') or []
    tipologia = resultat.get('tipologia_detectada') or ''
    genere = resultat.get('genere_detectat') or ''
    run_document = resultat.get('run_talles_document') or []
    sistema = resultat.get('sistema_talles') or 'unknown'

    # El run CONFIGURAT és el de la PEÇA DE DESTÍ (efectiu), no el de la mare: és el que la
    # pantalla de talles compara contra el document i el que el confirm acabarà escrivint.
    run_configurat = [
        s.strip() for s in (_efectiu(session, 'size_run_model') or '').replace(';', '·').split('·')
        if s.strip()
    ]

    # ── SET-2/T8 · MÉS D'UN PATRÓ AL DOCUMENT = AVÍS, MAI BARRERA ────────────────────
    # Amb «un import = una prenda», que el document en porti dues deixa de ser una condició
    # d'error: la peça de destí ja està decidida pel context i tot el que s'importi hi anirà.
    # El cribratge, doncs, EMET LA SENYAL i no barra; el gest de barrar-lo era el
    # `num_models == 1` de `pot_continuar`, que se'n va d'aquell predicat i es queda com a
    # dada. La senyal es PERSISTEIX (el confirm la torna al front, que és on es pinta).
    mes_duna_prenda = num_models > 1

    # Desa a la sessió (no fa gating; només cribratge).
    session.model_detectat = models_detectats
    session.run_conciliat = {
        'document': run_document,
        'sistema': sistema,
        'configurat': run_configurat,
        'estat': 'PENDENT',
    }
    # Persisteix el cribratge cru per a F2.3 (gènere/tipologia) sense tocar `resultat` definitiu.
    session.resultat = {**(session.resultat or {}), 'cribratge': resultat,
                        'mes_duna_prenda': mes_duna_prenda}
    session.estat = 'CRIBRATGE'
    session.save()

    plausible_genere = genere in ('woman', 'man', 'unisex', 'baby', 'kids')
    # Pel camí determinista no hi ha tipologia ni gènere per validar (ningú no els demanava
    # al parser): el que fa continuable el document és que s'hagi entès UNA taula de mesures.
    pot_continuar = (not cribratge_ia) or bool(
        tipologia and tipologia != 'unknown' and plausible_genere)

    return Response({
        'token': str(session.token),
        'estat': session.estat,
        'num_models': num_models,
        'model_detectat': models_detectats,
        'tipologia_detectada': tipologia,
        'genere_detectat': genere,
        'run_talles_document': run_document,
        'sistema_talles': sistema,
        'run_configurat': run_configurat,
        'pot_continuar': pot_continuar,
        # SET-2/T8 — la peça de destí, com a FET (el wizard la MOSTRA, no la pregunta) i la
        # senyal del document multi-prenda, que és avís i no barrera.
        'garment': session.garment,
        'garment_nom': _nom_de_la_peca(session),
        'mes_duna_prenda': mes_duna_prenda,
        # Traça del routing: qui ha resolt aquest cribratge. Visible a la resposta perquè
        # «quantes crides d'IA ha fet aquest import?» tingui resposta sense mirar logs.
        'cribratge_origen': ('ia' if cribratge_ia else 'parser_determinista'),
    }, status=200)


def _norm_label(s):
    return (s or '').strip().upper()


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def import_session_talles_view(request, token):
    """
    PATCH /api/v1/import-sessions/<token>/talles/  (Pas W1 — APARELLAMENT de talles)

    Rep:
      - talles_seleccionades: labels del DOCUMENT que el tècnic manté com a columnes.
      - talla_mapping: [{document, model}] editat per l'humà (opcional). Si NO ve, el backend
        auto-proposa l'aparellament per forma canònica (dialecte mesos inclòs).

    Retorna la proposta/validació (talla_mapping + no_aparellades + errors), les etiquetes REALS
    del model (system_labels, del SizeSystem) i base_size_label. El resultat és LA LLEI de la
    sessió: es desa a run_conciliat.talla_mapping i el confirm el consumeix en exclusiva (la clau
    `mapeig` antiga es retira). Validació: aparellament UNÍVOC (1↔1), model del system, sense dups.
    `alinear` RETIRAT: el run del model parla SEMPRE en etiquetes tenant.

    SET-2/T8 — TOT el que aquesta vista llegeix i escriu és de la PEÇA DE DESTÍ: el sistema
    de talles i la base són els EFECTIUS, i el canvi de talla base (B5) aterra a l'override
    de la prenda, no al model. Sense això, obrir el pas 1 d'un import a la 02 i triar-hi una
    base canviava la base de la MARE — el dany d'aquest tram, fet des del pas 1.
    """
    from fhort.models_app.models import ImportSession, ModelGarment

    session = ImportSession.objects.filter(token=token).select_related(
        'model', 'model__size_system',
    ).first()
    if not session:
        return Response({'error': 'Sessió no trobada'}, status=404)
    model = session.model
    if not model:
        return Response({'error': 'La sessió no té model associat'}, status=400)

    # La peça de destí i qui ESCRIU els seus camps heretables: la prenda si n'hi ha, el
    # model si som a la mare (que és el model mateix, D3).
    peca = _peca_de(session)
    destinatari = peca if peca is not None else model
    size_system = _efectiu(session, 'size_system')

    talles_sel = [str(t).strip() for t in (request.data.get('talles_seleccionades') or []) if str(t).strip()]
    mapping_in = request.data.get('talla_mapping')   # [{document, model}] editat per l'humà (opcional)
    base_in = request.data.get('base_size_label')    # B5: canvi de talla base (etiqueta model)
    # F5: full del llibre que el tècnic vol importar (xlsx multi-pestanya). Opcional; absent =
    # no s'hi toca. CAP endpoint nou i CAP camp nou: viu a `run_conciliat`, el JSONField que ja
    # és la casa de l'estat d'aquest pas. Qui re-extreu és el wizard, tornant a cridar
    # /extraccio/ — aquesta vista només desa la tria.
    full_in = request.data.get('full_seleccionat')

    # Run configurat actual de la PEÇA (etiquetes tenant; només informatiu).
    configurat = [
        s.strip() for s in (_efectiu(session, 'size_run_model') or '').replace(';', '·').split('·')
        if s.strip()
    ]

    # Etiquetes REALS de la peça (SizeDefinition del seu system, ordenades) — LA veritat del
    # panell dret.
    system_labels = []
    if size_system is not None:
        system_labels = list(size_system.talles.order_by('ordre').values_list('etiqueta', flat=True))
    canon_to_tenant = {}
    for _e in system_labels:
        canon_to_tenant.setdefault(canonical_size_label(_e), _e)

    def _propose(doc_labels):
        """Auto-proposta document→model per forma canònica (dialecte mesos inclòs). 1↔1."""
        pairs, no_ap, used = [], [], set()
        for d in doc_labels:
            tgt = canon_to_tenant.get(canonical_size_label(d))
            if tgt and tgt not in used:
                pairs.append({'document': d, 'model': tgt})
                used.add(tgt)
            else:
                no_ap.append(d)
        return pairs, no_ap

    errors = []
    if mapping_in is not None:
        # Validació de l'aparellament editat per l'humà: UNÍVOC (1↔1), model del system, sense dups.
        talla_mapping, no_aparellades, seen_doc, seen_model = [], [], set(), set()
        sys_set = set(system_labels)
        for pair in mapping_in:
            d = str((pair or {}).get('document') or '').strip()
            mdl = str((pair or {}).get('model') or '').strip()
            if not d:
                continue
            if not mdl:
                no_aparellades.append(d)
                continue
            if mdl not in sys_set:
                errors.append(f"La talla model «{mdl}» no és del sistema de talles del model.")
            if d in seen_doc:
                errors.append(f"La talla del document «{d}» surt aparellada més d'un cop.")
            if mdl in seen_model:
                errors.append(f"La talla del model «{mdl}» s'aparella dues vegades (ha de ser 1↔1).")
            seen_doc.add(d)
            seen_model.add(mdl)
            talla_mapping.append({'document': d, 'model': mdl})
    else:
        talla_mapping, no_aparellades = _propose(talles_sel)

    # ── B5 · TALLA BASE. Canvi opcional (limitat a les SizeDefinition del system) → escriu a
    # la PEÇA DE DESTÍ: a la seva fila d'override si és una prenda, al model si és la mare.
    if base_in is not None:
        base_in = str(base_in).strip()
        if base_in and base_in in set(system_labels) and base_in != (_efectiu(session, 'base_size_label') or ''):
            destinatari.base_size_label = base_in
            destinatari.save(update_fields=['base_size_label'])
        elif base_in and base_in not in set(system_labels):
            errors.append(f"La talla base «{base_in}» no és del sistema de talles del model.")
    base_label = (_efectiu(session, 'base_size_label') or '').strip()

    # Guard BLOQUEJANT: la talla base ha de tenir una columna del document aparellada (si no, l'import
    # no pot escriure el valor base → seria el 422 del confirm). Es bloqueja ja al pas 1.
    base_paired = any(p.get('model') == base_label for p in talla_mapping)
    base_avisos = []
    if base_label and not base_paired:
        errors.append(f"La talla base «{base_label}» no té cap columna del document aparellada.")

    # Avís NO bloquejant: base divergent de la convenció (mínima del run · S/38 dona · M/42 home)
    # o de l'àncora del ruleset del model.
    def _conventional_base():
        tgt = (model.target or '').upper()
        if any(k in tgt for k in ('WOMAN', 'WOMEN')):
            for c in ('S', '38'):
                if c in system_labels:
                    return c
        if 'MAN' in tgt or 'MEN' in tgt:
            for c in ('M', '42'):
                if c in system_labels:
                    return c
        return system_labels[0] if system_labels else None   # mínima del run
    conv = _conventional_base()
    if base_label and conv and base_label != conv:
        base_avisos.append(f"La talla base «{base_label}» divergeix de la convenció del segment (esperada «{conv}»).")
    _joc = _efectiu(session, 'grading_rule_set')
    if base_label and _joc is not None:
        anchor = (_joc.regles.values_list('talla_base__etiqueta', flat=True).first())
        if anchor and anchor != base_label:
            base_avisos.append(f"La talla base «{base_label}» divergeix de l'àncora del ruleset «{anchor}».")

    ready = bool(talla_mapping) and not errors

    rc = dict(session.run_conciliat or {})
    rc.update({
        'configurat': configurat,
        'seleccionades': talles_sel,
        'talla_mapping': talla_mapping,       # B1: LA LLEI de la sessió (document→model tenant).
        'no_aparellades': no_aparellades,
        'sense_desti': no_aparellades,        # compat: lectors antics.
        'estat': 'RESOLT' if ready else 'PENDENT',
    })
    if full_in is not None:
        nou_full = str(full_in).strip()
        rc['full_seleccionat'] = nou_full or None
    rc.pop('mapeig', None)                     # la clau `mapeig` MOR: una sola font de veritat.
    session.run_conciliat = rc
    if ready:
        session.estat = 'TALLES'
    session.save(update_fields=['run_conciliat', 'estat', 'actualitzat_at'])

    # Columnes del document sense parella → oferim pre-omplir el Size Map Setup (run de client nou).
    size_map_prefill = None
    if not ready and no_aparellades:
        target_codi = model.target or ''
        if not target_codi and size_system is not None:
            _ss_target = size_system.targets.first()
            if _ss_target:
                target_codi = _ss_target.codi
        size_map_prefill = {
            'target_codi': target_codi or None,
            'labels': no_aparellades,
            'base_size': base_label or None,
            'import_session_token': str(session.token),
            'model_id': model.id,
        }

    return Response({
        'ready': ready,
        'estat': session.estat,
        'run_conciliat': rc,
        'talla_mapping': talla_mapping,
        'no_aparellades': no_aparellades,
        'system_labels': system_labels,       # etiquetes REALS del model (selectors + panell dret).
        'base_size_label': base_label,
        'base_paired': base_paired,
        'base_avisos': base_avisos,           # B5: divergències no bloquejants de la talla base.
        'conventional_base': conv,
        'size_run_model': _efectiu(session, 'size_run_model'),
        'errors': errors,
        'size_map_prefill': size_map_prefill,
    }, status=200)


# ─────────────────────── Matching POMMaster (compartit) ───────────────────────
# Extret de create_from_extraction_view perquè l'extracció per sessió (W2) i la creació
# directa des d'extracció comparteixin EXACTAMENT la mateixa lògica de matching.
_POM_SYNONYMS = {
    # Existing
    'waist position':                  'waist position',
    'hip position':                    'hip position',
    'front body length':               'body length',
    'straight back body length':       'body length cb',
    'side length':                     'side seam',
    'front armhole curve':             'armhole curve',
    'neckline width':                  'neck width',
    'collar height':                   'collar height',
    'collar width':                    'collar width',
    'bottom width':                    'skirt sweep',
    'body zip length':                 'zip length',
    'lining length at center front':   'lining length',
    'lining length at center back':    'lining length',
    'lining bottom width along hem':   'lining hem width',
    # El bloc "Brownie positional POMs" (nomenclatura del customer BRW disfressada de sinònim
    # canònic) s'ha MIGRAT a CustomerPOMAlias (origen=MIGRACIO), migració pom 0031 (N2-4a,
    # DIAGNOSI_NOMENCLATURA_ALIES_2026-07-08). Els sinònims genèrics d'aquest diccionari es
    # conserven; el matcher llegeix els àlies com a estratègia (a) prioritària (N3 fet, veure
    # find_pom_master més avall).
}


def _nom_resolt(pom):
    """El nom llarg d'un POM per a la UI del wizard, o `None` si no n'hi ha POM.

    🔑 PUNT ÚNIC DE F1. Els tres camins del wizard —les files del pas 2, el suggeriment feble i
    els candidats del 409— servien `POMMaster.nom_client` CRU, i aquell camp és buit a **103
    dels 144 POMs actius** del schema `fhort` (mesurat el 26/08; tots 103 amb `pom_global`
    poblat). El resultat era la fila muda que la formació va veure: «B · », «BF · », «BT · ».

    `noms_de` és el resolutor únic de la llei **ÀLIES > TENANT > GLOBAL** (22/08) i ja cau al
    canònic del sector quan el tenant no ha batejat el POM. Aquí es demana sense àlies a posta:
    el wizard resol contra el CATÀLEG DE LA CASA, i la nomenclatura del client d'aquell model
    ja viatja per `codi_fitxa`/`descripcio`, que són el que el document porta escrit.

    Es torna `nom_en` —el canònic— perquè és la PRIMERA línia de la nomenclatura de la casa;
    la segona (el nom d'usuari, en gris) la resol el front amb `nomsDePom`, que ja rep
    `nom_ca` pel seu costat. `None` i no `''` per no canviar el contracte de qui ja distingeix
    «no hi ha POM» de «el POM no té nom».
    """
    if pom is None:
        return None
    from fhort.pom.nomenclatura import noms_de
    return noms_de(pom)['nom_en'] or None


def find_pom_master(code, description, customer=None):
    """
    Find the most suitable POMMaster.
    Return (pom_master, match_type, confidence)
    confidence: 'HIGH' | 'MEDIUM' | 'LOW' | 'NO_MATCH'

    ORDRE (DIAGNOSI_NOMENCLATURA_ALIES_2026-07-08, N3):
      (a) ÀLIES exacte del `customer` (CustomerPOMAlias) → HIGH. Requereix `customer`; si és None
          (context sense client) se salta. El `client_code` d'un àlies pot ser un codi posicional
          (LOS 'H.6') O el text de la descripció del client (BRW 'front armhole curve') → es prova
          contra `code` I contra `description`.
          ⚠️ Un àlies amb `pendent_revisio=True` **NO auto-vincula** (v. sota, QA-S8-R1).
      (b) descripció + sinònims canònics → HIGH/MEDIUM (nom_client, POMGlobal.nom_en).
      (c) codi numèric + 'lining' → MEDIUM.
      (c-bis) l'àlies PENDENT DE REVISIÓ, com a darrer suggeriment → LOW (mai auto-vincle).
      (d) FALLBACK TRANSITORI (deprecació — objectiu de la diagnosi: treure `codi_client` del
          matcher): `codi_client` exacte i root-prefix → LOW. Amb el llindar d'auto-vinculació
          (c2b19bd) un LOW NO auto-vincula: cau a pendents amb el suggeriment visible. Abans
          anaven PRIMER amb HIGH; ara són l'últim recurs, per sota de l'àlies i la descripció.
    """
    from fhort.pom.models import POMMaster, CustomerPOMAlias

    desc_clean = (description or '').lower().strip()
    desc_base = _re.sub(r'\s*[\(\[].*?[\)\]]', '', desc_clean).strip()

    # (a) Àlies de nomenclatura del client. Va PRIMER: un codi/descripció reclamat explícitament
    # per un àlies d'AQUEST customer mana sobre qualsevol heurística de descripció.
    #
    # ⚠️ QA-S8-R1 · LA PORTA DEL MATCHER. Un àlies marcat `pendent_revisio` és un àlies del qual
    # el sistema DESCONFIA: el guard d'aprenentatge (pom/services.py) el marca així quan el POM
    # que reclama ja el reclamava un ALTRE codi del mateix client — o sigui, quan o bé sobra, o
    # bé una de les dues mesures acabarà sobre el POM equivocat. Un àlies del qual desconfiem no
    # pot ser alhora la font de màxima confiança del matcher: seria marcar-lo per revisar i
    # continuar creient-l'hi. Aquí es DEGRADA a suggeriment (c-bis), mai a auto-vincle.
    #
    # I no s'atura la cerca: es prova la resta d'estratègies, que poden trobar-hi un vincle bo
    # de debò. L'àlies pendent només parla si no parla ningú altre.
    #
    # `pom__isnull=False` (QA-S8-R1): un àlies SENSE POM no és matchable — és vocabulari del
    # client pendent de mapar (CustomerPOMAlias.pom és nullable, migració 0037). No té destí,
    # així que no pot vincular res, i sense el filtre `alias.pom.actiu` petaria amb AttributeError.
    alias_pendent = None
    if customer is not None:
        for key in (k for k in (code, desc_clean) if k):
            # `pom__pom_global` i no només `pom`: aquest era **l'únic camí de match** que no
            # prefetchava el catàleg global, i des que la fila serveix el nom RESOLT
            # (`noms_de`, F1) cada match d'àlies hi hauria comprat una query. La resta de
            # branques d'aquesta funció ja fan `select_related('pom_global')`; aquesta s'hi
            # posa al costat.
            alias = (CustomerPOMAlias.objects
                     .filter(customer=customer, client_code__iexact=key, pom__isnull=False)
                     .select_related('pom', 'pom__pom_global').first())
            if alias and alias.pom.actiu:
                if alias.pendent_revisio:
                    if alias_pendent is None:
                        alias_pendent = alias.pom
                    continue
                return alias.pom, 'alias_match', 'HIGH'

    if desc_clean:
        # Strategy 2 — explicit synonym (curated table).
        syn = _POM_SYNONYMS.get(desc_clean) or _POM_SYNONYMS.get(desc_base)
        if syn:
            for pm in POMMaster.objects.select_related('pom_global').filter(actiu=True):
                nom = (pm.nom_client or '').lower()
                if syn in nom or nom in syn:
                    return pm, 'synonym_match', 'HIGH'
            for pm in POMMaster.objects.select_related('pom_global').filter(
                pom_global__isnull=False, actiu=True,
            ):
                nom_en = (pm.pom_global.nom_en or '').lower()
                if syn in nom_en or nom_en in syn:
                    return pm, 'synonym_global_match', 'HIGH'

        # Strategy 3 — match by nom_client (exact=HIGH, contains=MEDIUM).
        for pm in POMMaster.objects.select_related('pom_global').filter(actiu=True):
            nom = (pm.nom_client or '').lower()
            if desc_base and len(desc_base) > 3:
                if desc_base == nom:
                    return pm, 'exact_description', 'HIGH'
                if desc_base in nom or nom in desc_base:
                    return pm, 'description_match', 'MEDIUM'

        # Strategy 4 — match by POMGlobal nom_en / abbreviation.
        for pm in POMMaster.objects.select_related('pom_global').filter(
            pom_global__isnull=False, actiu=True,
        ):
            pg = pm.pom_global
            nom_en = (pg.nom_en or '').lower()
            abbrev = (pg.abbreviation or '').lower()
            if desc_base and len(desc_base) > 3:
                if desc_base == nom_en:
                    return pm, 'global_exact', 'HIGH'
                if desc_base in nom_en or nom_en in desc_base:
                    return pm, 'global_name_match', 'MEDIUM'
            if code and code.lower() == abbrev:
                return pm, 'abbreviation_match', 'HIGH'

    # (c) Strategy — pure numeric codes → lining.
    if code and code.isdigit():
        desc_lower = (description or '').lower()
        if 'lining' in desc_lower:
            for pm in POMMaster.objects.select_related('pom_global').filter(actiu=True):
                nom = (pm.nom_client or '').lower()
                if 'lining' in nom:
                    return pm, 'numeric_lining_match', 'MEDIUM'

    # (c-bis) L'ÀLIES PENDENT DE REVISIÓ (QA-S8-R1). Cap altra estratègia no ha trobat res ferm,
    # així que ara sí que val la pena dir què reclamava aquell àlies del qual desconfiem — però
    # com el que és: un SUGGERIMENT (LOW). El llindar (`_apply_match_threshold`) el deixarà a
    # pendents amb el nom visible, i una persona decidirà. Va per damunt dels fallbacks de codi
    # (d) perquè un àlies el va declarar algú d'aquest client; un root-prefix no l'ha declarat ningú.
    if alias_pendent is not None:
        return alias_pendent, 'alias_pendent_revisio', 'LOW'

    # (d) FALLBACK TRANSITORI — `codi_client` exacte. Abans era la 1a estratègia amb HIGH; ara és
    # penúltim recurs amb LOW (deprecació): l'àlies i la descripció manen. Un exacte que arriba
    # aquí no ha resolt per àlies ni per descripció → suggeriment feble, no auto-vinculació.
    if code:
        pm = POMMaster.objects.filter(codi_client__iexact=code, actiu=True).first()
        if pm:
            return pm, 'legacy_code_match', 'LOW'

    # (d) FALLBACK TRANSITORI (ÚLTIM RECURS) — root de lletres inicials per a codis posicionals
    # (D1, G2s → D, G). NO es rooteja la nomenclatura d'AGRUPACIÓ 'LLETRA.NÚMERO' (H.6, G.3, J.2):
    # la lletra és un grup del document, no un codi de mesura, i col·lapsaria a un POM d'una sola
    # lletra aliè. Confiança LOW: darrer recurs, no una vinculació segura.
    if code and not _re.match(r'^[A-Za-z]+\.\d', code):
        m = _re.match(r'^([A-Za-z]+)', code)
        if m and m.group(1) != code:
            root = m.group(1)
            pm = POMMaster.objects.filter(codi_client__iexact=root, actiu=True).first()
            if pm:
                return pm, 'root_code_match', 'LOW'

    return None, 'no_match', 'NO_MATCH'


# ─────────────────────────────────────────────────────────────────────────────
# PORTES DE VINCULACIÓ (QA-S8, DIAGNOSI_QA_S8_IMPORT)
#
# Bessones de les de `pom/size_map_views.py:29,53` (importador de la Size Library), que
# ja les tenia i que l'importador de MODELS no. La diagnosi va trobar el forat: el mateix
# mode de fallada estava protegit en un importador i despullat a l'altre. No s'extreu un
# helper compartit entre les dues apps (seria refactor fora d'abast); s'adapten aquí a les
# claus de `poms_extrets` (`pom_master_id`/`pom_codi`/`pom_nom`) i el docstring diu d'on
# vénen, perquè el dia que una de les dues canviï se sàpiga que hi ha una germana.
# ─────────────────────────────────────────────────────────────────────────────

#: Un match per sota d'això NO auto-vincula. Un LOW (codi legacy, arrel del codi) és el
#: darrer recurs del matcher, no una certesa: la fila cau a pendents amb el suggeriment
#: visible i la persona decideix. Vincular en silenci amb confiança baixa és el que va
#: fer que 'U2' i 'U3' (First/Last button) acabessin tots dos sobre el POM 'U'
#: (Width sequins piece) — un disbarat que ningú no va veure perquè no es va dir.
_POM_AUTOLINK_CONF = ('HIGH', 'MEDIUM')


def _apply_match_threshold(pom, conf):
    """El llindar: (pom, conf) → (pom_efectiu, weak_suggestion).

    Per sota del llindar es desvincula i es torna el nom suggerit, perquè la UI el mostri
    com a pendent. Mai una vinculació dubtosa en silenci.
    """
    if pom is not None and conf not in _POM_AUTOLINK_CONF:
        return None, pom
    return pom, None


def _apply_many_to_one_guard(rows):
    """Si DUES files del document resolen al MATEIX POM, **cap de les dues auto-vincula**.

    `BaseMeasurement` és únic per `(model, pom)`: dues files que hi cauen col·lapsen, i la
    segona sobreescriu la primera **en silenci** (W5, `update_or_create`). El símptoma que
    va veure QA —una mesura del document que desapareix— surt exactament d'aquí.

    ⚠️ **AQUÍ L'ÀLIES NO QUEDA EXEMPT, i la germana de `size_map_views.py:53` SÍ.** La
    divergència és deliberada i és el moll de l'os:

      · A `size_map` el destí és `GradingRule`, i que dos codis del client comparteixin un
        POM hi és tolerable (Losan H.11 sleeve opening / H.16 cuff opening).
      · Aquí el destí és `BaseMeasurement`, **únic per (model, pom)**. Dues files NO hi
        caben. Per legítim que sigui l'àlies, la segona esborra la primera. L'exempció
        importaria una premissa que en aquest destí no es compleix.

    I no és teòric: al catàleg viu, el client BRW té els àlies `F` i `FF` (Centre FRONT
    length i Centre BACK length — dues mesures distintes) tots dos cap al POM 389
    'TOTAL LENGTH', i `U2`/`U3` (First/Last button) tots dos cap al 439 'Width sequins
    piece'. Amb l'exempció posada, aquestes quatre files travessaven les dues portes amb
    confiança HIGH i dues mesures del document s'esborraven a W5 sense dir res.

    Un àlies dolent és un problema del catàleg i es resol al catàleg; el que aquesta porta
    ha de garantir és que **no acabi sent una pèrdua de dades silenciosa**.

    ── SET-2/T8-ter (16/08) · LA PREMISSA D'AQUEST GUARD HA CADUCAT DUES VEGADES ──────────
    La frase de dalt —«`BaseMeasurement` és únic per `(model, pom)`»— era certa el dia que es va
    escriure i ja no ho és: la clau porta `capa` i `instancia` des de C1/T3 i `garment` des de
    T2. El guard, però, seguia comptant per `pom_master_id` PELAT, i per això desvinculava files
    que **no col·lideixen**.

    És exactament el que va passar amb la Brumà el 16/08: «G1 · Bottom height» (faldilla) i
    «M1 · Bottom hem height» (short) resolen al mateix POM 962 i el guard va desvincular-les
    totes dues. La persona va haver de resoldre-les a mà —dos 409 al rastre de la sessió 113— i
    la sortida que va trobar va ser inventar-li una instància al short. El guard fabricava la
    feina que després obligava a fabricar la dada.

    Ara compta per **`(pom, peça proposada)`**, que és el gra que la clau realment té a
    l'abast en aquest moment del pipeline. Els altres dos eixos NO hi entren i és a posta:
    capa i instància encara no les ha dit ningú quan aquest guard corre (les decideix la persona
    al pas 2), o sigui que comptar-hi seria donar per bo un buit que ningú ha confirmat. La peça
    sí que hi és, perquè la porta el DOCUMENT (`seccio` → `_proposta_de_peca`, F2).

    Dues files del mateix POM dins de la MATEIXA peça segueixen desvinculant-se, que és el
    vermell original i ha de seguir sent-ho.

    Muta `rows` in situ.
    """
    counts = {}
    for r in rows:
        if r.get('pom_master_id'):
            clau = (r['pom_master_id'], r.get('garment_proposat') or '')
            counts[clau] = counts.get(clau, 0) + 1
    dup_ids = {clau for clau, n in counts.items() if n >= 2}
    if not dup_ids:
        return rows

    for r in rows:
        if (r.get('pom_master_id'), r.get('garment_proposat') or '') in dup_ids:
            # El suggeriment queda VISIBLE: la persona ha de poder veure a què s'assemblava.
            r['weak_suggestion'] = r.get('pom_nom')
            r['weak_suggestion_codi'] = r.get('pom_codi')
            r['pom_master_id'] = None
            r['pom_codi'] = None
            r['pom_nom'] = None
            r['many_to_one'] = True
            r['actiu'] = False
    return rows


def _match_rows(files, customer, model=None):
    """Files llegides del document → `poms_extrets`, amb les portes aplicades.

    **Font ÚNICA de matching per als DOS camins d'extracció** (parser ràpid d'Excel i visió
    Opus). Abans cadascun es muntava la seva llista i divergien: la via ràpida marcava
    `actiu=True` per a tothom i la via Opus `actiu=bool(pm)`. Ara el criteri és un i és
    aquest:

        **actiu ⇔ vincle FERM** (match per sobre del llindar i no compartit amb cap altra fila).

    `files`: [{codi_fitxa, descripcio, values, tol_minus, tol_plus, seccio}].
    Retorna (poms_extrets, stats) amb stats = {n_nomatch, n_low, n_many_to_one}.

    `seccio` (F4) travessa TAL QUAL: aquesta funció aparella POMs, no interpreta el document.
    Els dos camins d'extracció l'omplen (el parser llegint la fila de secció, la IA amb el camp
    `section` de l'esquema) i el que arriba aquí ja és el text final.

    SET-2/T8-ter — `model` hi entra perquè la PROPOSTA de peça (F2) s'ha de resoldre AQUÍ DINS,
    entre l'aparellament i el guard de many-to-one: el guard compta per `(pom, peça proposada)` i
    sense la proposta ja feta comptaria per POM pelat i desvincularia files que no col·lideixen.
    L'ordre dels tres passos és, doncs, part del contracte i no una casualitat. `model=None`
    (cap cridador de producte) degrada a la proposta buida: totes les files a la mare.
    """
    rows = []
    n_nomatch = n_low = 0

    for i, f in enumerate(files):
        codi = (f.get('codi_fitxa') or '').strip()
        descripcio = (f.get('descripcio') or '').strip()
        pm, match_type, confidence = find_pom_master(codi, descripcio, customer=customer)

        # Els comptadors es prenen del match CRU (abans del llindar): així l'avís continua
        # distingint "no s'ha trobat res" de "s'ha trobat però no és de fiar".
        if pm is None:
            n_nomatch += 1
        elif confidence == 'LOW':
            n_low += 1

        pm_efectiu, suggeriment = _apply_match_threshold(pm, confidence)

        rows.append({
            'codi_fitxa': codi,
            'descripcio': descripcio,
            'pom_master_id': pm_efectiu.id if pm_efectiu else None,
            'pom_codi': pm_efectiu.codi_client if pm_efectiu else None,
            # 🔑 F1 · EL NOM VA RESOLT, NO CRU. Era `pm_efectiu.nom_client`, i el camp del
            # tenant és buit a **103 dels 144 POMs actius** de `fhort` —tots amb `pom_global`
            # poblat—, o sigui que la fila del wizard sortia «B · » i el nom canònic, que hi
            # era, no arribava mai. `noms_de` és la font única de la llei ÀLIES > TENANT >
            # GLOBAL (22/08) i cau al global quan el tenant no bateja.
            'pom_nom': _nom_resolt(pm_efectiu),
            'match_type': match_type,
            'confidence': confidence,
            'values': f.get('values') or {},
            'tol_minus': f.get('tol_minus'),
            'tol_plus': f.get('tol_plus'),
            'seccio': f.get('seccio') or None,
            'actiu': bool(pm_efectiu),
            'ordre': i,
            # El suggeriment és el MATEIX text que hauria sortit si el vincle hagués estat
            # ferm: resoldre'l per un camí i no per l'altre faria que la mateixa mesura es
            # digués de dues maneres segons la confiança del match.
            'weak_suggestion': _nom_resolt(suggeriment),
            'weak_suggestion_codi': suggeriment.codi_client if suggeriment else None,
            'many_to_one': False,
        })

    # L'ORDRE MANA: proposta (F2) → guard (F4). Vegeu el docstring.
    proposta = _proposta_de_peca(model, rows)
    _apply_many_to_one_guard(rows)
    n_many = sum(1 for r in rows if r.get('many_to_one'))

    return rows, {'n_nomatch': n_nomatch, 'n_low': n_low, 'n_many_to_one': n_many,
                  'proposta': proposta}


def _avisos_de_matching(stats):
    """Els avisos del matching. Un per motiu, i cadascun diu QUÈ ha de fer la persona."""
    avisos = []
    if stats['n_nomatch']:
        avisos.append(
            f"{stats['n_nomatch']} POM(s) sense match al catàleg — cal revisar o "
            f"afegir manualment."
        )
    if stats['n_low']:
        avisos.append(
            f"{stats['n_low']} POM(s) amb confiança baixa: NO s'han vinculat "
            f"automàticament. Revisa'ls al pas de POMs — hi tens el suggeriment."
        )
    if stats['n_many_to_one']:
        avisos.append(
            f"{stats['n_many_to_one']} POM(s) de la fitxa apuntaven al MATEIX POM del "
            f"catàleg: cap no s'ha vinculat automàticament (dues mesures no poden compartir "
            f"un POM: la segona esborraria la primera). Resol-los un per un."
        )
    return avisos


# ═══════════════════════════ W2 — Extracció POMs ═══════════════════════════
EXTRACCIO_MODEL = 'claude-opus-4-7'
EXTRACCIO_MAX_TOKENS = 16000

EXCEL_REVISION_MODEL = 'claude-sonnet-4-6'
EXCEL_REVISION_MAX_TOKENS = 2000


def _revise_excel_poms_with_sonnet(poms_text: str, api_key: str) -> dict:
    """Revisió lleugera (Sonnet) dels POMs extrets d'un Excel. No-fatal:
    retorna SEMPRE un dict {'corrections': [...], 'warnings': [...]}."""
    import anthropic
    from fhort.models_app.extraction_utils import safe_json_parse

    default = {'corrections': [], 'warnings': []}
    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model=EXCEL_REVISION_MODEL,
            max_tokens=EXCEL_REVISION_MAX_TOKENS,
            system="""Ets un validador de fitxes tècniques tèxtils.
Reps una llista de POMs (punts de mesura) extrets d'un Excel.
Retorna NOMÉS un JSON amb aquest format exacte:
{"corrections": [{"codi": "X", "camp": "descripcio|dim", "valor_suggerit": "..."}], "warnings": ["..."]}
Si no cal cap correcció, retorna {"corrections": [], "warnings": []}
No afegeixis cap text fora del JSON.""",
            messages=[{'role': 'user', 'content': poms_text}],
        )
        raw = ''.join(
            b.text for b in response.content if getattr(b, 'type', None) == 'text'
        ).strip()
        registra_us_ia(cami='revisio', model_ia=EXCEL_REVISION_MODEL,
                       usage=getattr(response, 'usage', None))
        parsed = safe_json_parse(raw)
        if not isinstance(parsed, dict):
            return default
        return {
            'corrections': parsed.get('corrections') or [],
            'warnings': parsed.get('warnings') or [],
        }
    except Exception:
        _logging.getLogger(__name__).exception('Revisió Excel (Sonnet): error no-fatal')
        return default


def _avis_files_perdudes(n_document, n_extretes):
    """FIX D (DIAGNOSI_QA_S8_IMPORT §D1e): el document té més files de POM que les extretes.

    La fitxa del Tate té 26 POMs i la IA en va retornar 25. La que va deixar caure —`JJ`,
    '1/2 Elbow width'— era l'única fila SENSE valor a la talla base, i **cap avís ho va dir**:
    una fila del document desapareixia en silenci. Un POM sense mesura base és legítim
    (`BaseMeasurement.base_value_cm` és `null=True`), així que la fila no s'havia de perdre;
    i si es perd, s'ha de dir.

    El recompte del document el dona el parser determinista (`meta['n_files_amb_codi']`), que
    sap comptar les files encara que abdiqui de llegir-ne els valors. Només s'avisa en el sentit
    que fa mal —el document en té MÉS que les extretes—, mai al revés.
    """
    if n_document and n_extretes < n_document:
        perdudes = n_document - n_extretes
        return [f"El document té {n_document} files amb codi de POM i se n'han extret "
                f"{n_extretes}: {perdudes} fila(es) no s'han llegit. Revisa-les a mà (sovint "
                f"són files sense valor a la talla base, que són POMs igualment vàlids)."]
    return []


def _avis_fulls_multiples(fulls, full_llegit):
    """F5 · el llibre té més d'una pestanya amb POMs i només se n'ha llegit una.

    Fins ara els fulls 2..N es perdien SENCERS i sense cap avís de contingut: l'únic que
    arribava era el genèric de multi-model del cribratge, que parla de models detectats i no de
    pestanyes no llegides. Un document multi-peça amb una peça per full desapareixia en silenci.
    """
    amb_poms = [f for f in (fulls or []) if f.get('passa_porta')]
    if len(amb_poms) <= 1:
        return []
    noms = ', '.join(f"'{f['nom']}' ({f['n_files_amb_codi']} POMs)" for f in amb_poms)
    return [f"El llibre té {len(amb_poms)} fulls amb taules de POMs: {noms}. "
            f"S'ha llegit '{full_llegit}'; els altres NO s'importen. Si la fitxa porta una "
            f"peça per full, tria quin vols importar."]


def _extraccio_via_excel(session, api_key):
    """Via ràpida d'extracció per a fitxes Excel: parse determinista + revisió Sonnet,
    SENSE la crida Opus.

    Retorna `(resposta, meta)`:
      · `resposta` = Response amb la MATEIXA forma que la via PDF/imatge, o **None** si el
        parser abdica (el caller fa fallback IA via Opus, com sempre).
      · `meta` = el que el parser ha pogut saber del document **encara que abdiqui** — hi ha
        el recompte de files, que el camí IA necessita per al Fix D.
    """
    from django.conf import settings
    # 1. Bytes del document desat al Pas 1.
    try:
        session.document.open('rb')
        file_bytes = session.document.read()
    finally:
        session.document.close()

    # 2. Parse determinista. Les pistes del model (talla base i run) ajuden a reconèixer les
    # columnes de talla quan el document no declara `SAMPLE SIZE`; si el document sí que ho
    # diu, mana el document.
    # SET-2/T8 — les pistes són les EFECTIVES de la peça de destí, com al cribratge: llegir el
    # document amb l'escala de la mare quan la prenda en té una de pròpia és ancorar-lo malament.
    model = session.model
    raw_poms, talles_detectades, meta = _parse_excel_poms(
        file_bytes,
        base_hint=(_efectiu(session, 'base_size_label') if model else None),
        run_hint=[s.strip() for s in ((_efectiu(session, 'size_run_model') if model else '') or '')
                  .replace(';', '·').split('·') if s.strip()],
        # F5 · el full que el tècnic ha triat al pas 1. Viu a `run_conciliat` (JSONField que ja
        # és la casa de l'estat de W1) i NO en un camp nou: no cal migració per a un punter.
        full_seleccionat=(session.run_conciliat or {}).get('full_seleccionat'),
    )

    # 3. Sense POMs llegibles → senyal (None) perquè el caller faci fallback IA (Opus).
    if not raw_poms:
        # L'informe de fulls viatja igual: el camí IA també l'ha de poder ensenyar (el llibre
        # té els fulls que té, els hagi entès el parser o no).
        return None, meta

    # 4. Text pla per a la revisió Sonnet.
    linies = [
        f"{p['codi_fitxa']} | {p['descripcio']} | DIM:{p.get('dim', '')} | {p['values']}"
        for p in raw_poms
    ]
    poms_text = '\n'.join(linies)

    # 5. Revisió lleugera (no-fatal) — OPT-IN, per defecte APAGADA.
    #
    # Decisió Agus 2026-07-22: «un xlsx parsejable no ha de costar ni un cèntim de token».
    # Aquesta crida només retoca `descripcio`/`dim` —mai un valor de mesura— sobre files que
    # el parser determinista ja ha entès i demostrat que entenia (porta d'abdicació). És a
    # dir: paga una crida a Sonnet per a cada import d'xlsx a canvi de polir text cosmètic.
    # Queda darrere d'un setting per poder-la encendre quan es vulgui avaluar, sense que
    # sigui el comportament de tothom.
    revision = ({'corrections': [], 'warnings': []}
                if not getattr(settings, 'IMPORT_REVISIO_SONNET', False)
                else _revise_excel_poms_with_sonnet(poms_text, api_key))

    # 6. Aplica correccions (només camp descripcio/dim, codis existents).
    by_codi = {}
    for p in raw_poms:
        by_codi.setdefault(p['codi_fitxa'], p)
    for corr in (revision.get('corrections') or []):
        if not isinstance(corr, dict):
            continue
        target = by_codi.get(str(corr.get('codi') or '').strip())
        camp = corr.get('camp')
        if not target or camp not in ('descripcio', 'dim'):
            continue
        if camp == 'descripcio':
            target['descripcio'] = str(corr.get('valor_suggerit') or '').strip()
        else:  # dim
            try:
                target['dim'] = float(str(corr.get('valor_suggerit')).replace(',', '.'))
            except (ValueError, TypeError):
                pass

    # 7-8. Matching POM + format IDÈNTIC al de la via Opus: la MATEIXA funció (`_match_rows`),
    # amb les mateixes portes. Abans aquesta via marcava `actiu=True` per a totes les files,
    # inclosos els sense match; ara el criteri és únic (actiu ⇔ vincle ferm) perquè el
    # matching és literalment el mateix codi.
    # N3: customer del model per resoldre els àlies de nomenclatura del client.
    import_customer = session.model.customer if session.model_id else None
    poms_extrets, stats = _match_rows(raw_poms, import_customer, session.model)

    avisos_extraccio = list(revision.get('warnings', []))
    avisos_extraccio += _avisos_de_matching(stats)
    avisos_extraccio += _avis_fulls_multiples(meta.get('fulls'), meta.get('full'))
    if meta.get('full_seleccionat_ignorat'):
        avisos_extraccio.append(
            f"El full triat «{meta['full_seleccionat_ignorat']}» no té cap taula de POMs "
            f"llegible; s'ha llegit «{meta.get('full')}».")

    # 9. Talles, capçalera i talla BASE — els tres, llegits del document.
    #
    # PARITAT AMB LA VIA OPUS (bandera 3 de la diagnosi). Aquesta via retornava `header: {}`
    # i `base_size = sizes[0]` **encara quan funcionava**. Cap de les dues coses era innocent:
    #   · el `header` buit deixava el wizard sense marca/temporada/nom d'estil, que la via Opus
    #     sí que omplia → dues respostes amb la mateixa forma i contingut diferent;
    #   · i `sizes[0]` NO és la talla base: a la fitxa Rosalia el run és XXS·XS·S·M·L i la base
    #     és 'S', no 'XXS'. A més, sense `base_size` a `resultat['extraccio']`, la reconciliació
    #     de talles de W5 (:1426) queda desactivada del tot per aquest camí ("manca base").
    # Ara les dues surten del bloc de metadades del document (`SAMPLE SIZE`), i el camí ràpid
    # deixa de ser un ciutadà de segona.
    sizes = [str(t) for t in talles_detectades]
    header = meta.get('header') or {}
    base_size = meta.get('base_size') or (sizes[0] if sizes else None)
    # F5 · informe de fulls del llibre + quin s'ha llegit. El wizard el necessita per poder dir
    # «el llibre té N fulls amb POMs» i per oferir-ne la tria.
    fulls = meta.get('fulls') or []
    extraccio = {'via': 'excel', 'header': header, 'sizes': sizes, 'base_size': base_size,
                 'fulls': fulls, 'full': meta.get('full')}

    # 10. Persisteix. NOTA: `session.poms_extrets` és la font de veritat per als passos
    # W2-confirmació (:1216) i W3-mesures (:1415); cal desar-la (paritat amb la via Opus).
    # SET-2/T8-ter — la proposta ja l'ha resolta `_match_rows` (ha de córrer ABANS del guard
    # de many-to-one, que hi compta); aquí només se'n desa la traça.
    _proposta = stats.get('proposta') or {}
    session.resultat = {**(session.resultat or {}),
                        'extraccio': extraccio,
                        'proposta_peces': _proposta,
                        'grading_status': 'ok'}
    session.poms_extrets = poms_extrets
    session.avisos = list(session.avisos or []) + avisos_extraccio
    session.estat = 'POMS'
    session.save(update_fields=['resultat', 'poms_extrets', 'avisos', 'estat',
                                'actualitzat_at'])

    # 11. Resposta amb EXACTAMENT el mateix format que la via PDF/imatge (:1180-1188),
    # `suggested_valors_mode` inclòs: el toggle absoluts/deltes del wizard el llegeix, i sense
    # ell aquesta via el deixava sense default. Cosmètic → mai pot petar l'extracció.
    try:
        from fhort.pom.grading_utils import suggest_valors_mode
        suggested_valors_mode = suggest_valors_mode(
            {p['pom_master_id']: p['values'] for p in poms_extrets
             if p.get('pom_master_id') and p.get('values')},
            base_size, sizes)
    except Exception:
        suggested_valors_mode = 'absoluts'

    return Response({
        'estat': 'POMS',
        'poms_extrets': poms_extrets,
        'header': header,
        'base_size': base_size,
        'sizes': sizes,
        'grading_status': {'status': 'ok', 'detail': ''},
        'avisos': avisos_extraccio,
        'suggested_valors_mode': suggested_valors_mode,
        # SET-2/T8-ter — la proposta secció→peça i, sobretot, les seccions que NO tenen
        # peça: una absència que el pas 2 ha de poder DIR (crear la peça és un gest humà).
        'proposta_peces': _proposta,
        'fulls': fulls,                    # F5 · informe del llibre
        'full': meta.get('full'),          # F5 · quin s'ha llegit
    }, status=200), meta


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_session_extraccio_view(request, token):
    """
    POST /api/v1/import-sessions/<token>/extraccio/  (Pas W2 — Crida 2: extracció completa)

    Re-llegeix el document desat a la sessió i fa l'extracció completa (POMs + valors +
    grading) amb visió Opus 16k. Per cada POM crida find_pom_master i desa el matching a
    session.poms_extrets. Desa l'extracció completa a session.resultat. estat→'POMS'.
    SEMPRE retorna (mai bloqueja: salvage de Fase 1 si el JSON global falla).
    """
    import anthropic
    from django.conf import settings

    from fhort.models_app.models import ImportSession
    from fhort.models_app.extraction_prompt import TECH_SHEET_EXTRACTION_PROMPT
    from fhort.models_app.extraction_utils import safe_json_parse, salvage_measurements

    session = ImportSession.objects.filter(token=token).select_related('model').first()
    if not session:
        return Response({'error': 'Sessió no trobada'}, status=404)
    if not session.document:
        return Response({'error': 'La sessió no té document desat'}, status=400)

    api_key = getattr(settings, 'ANTHROPIC_API_KEY', '')
    if not api_key:
        return Response({'error': 'ANTHROPIC_API_KEY no configurada al backend'}, status=500)

    # Via ràpida Excel: parse determinista + revisió Sonnet, saltant Opus.
    # Si el parser ràpid no reconeix el format (None), es continua pel camí comú
    # Opus amb el full de càlcul convertit a text. PDF/imatge no canvien.
    doc_name = session.document.name or ''
    es_excel = doc_name.lower().endswith(('.xlsx', '.xls'))
    excel_meta = {}
    if es_excel:
        resposta_rapida, excel_meta = _extraccio_via_excel(session, api_key)
        if resposta_rapida is not None:
            return resposta_rapida

    # Llegeix el document desat al Pas 1.
    try:
        session.document.open('rb')
        file_bytes = session.document.read()
    finally:
        session.document.close()

    avisos = list(session.avisos or [])
    detectats = session.model_detectat or []
    if len(detectats) > 1:
        avisos.append(
            f'Document multi-model ({len(detectats)} detectats); extracció del model principal.'
        )

    if es_excel:
        content_block = {'type': 'text',
                         'text': f'Contingut del full de càlcul (fitxa Excel):\n{_excel_to_text(file_bytes)}'}
        # L'avís diu ara PER QUÈ el parser ha abdicat. Abans deia només que ho havia fet, i la
        # diagnosi va haver d'executar el parser sobre els bytes reals per esbrinar el motiu.
        motiu = (excel_meta.get('motiu') or '').strip()
        avisos.append('Format Excel no reconegut pel parser ràpid; extracció via IA.'
                      + (f' Motiu: {motiu}.' if motiu else ''))
    else:
        content_block = _cribratge_content_block(file_bytes, session.document.name, '')

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model=EXTRACCIO_MODEL,
            max_tokens=EXTRACCIO_MAX_TOKENS,
            thinking={'type': 'adaptive'},
            output_config={'effort': 'high'},
            system=[{'type': 'text', 'text': TECH_SHEET_EXTRACTION_PROMPT,
                     'cache_control': {'type': 'ephemeral'}}],
            messages=[{'role': 'user', 'content': [content_block]}],
        )
        raw = ''.join(
            b.text for b in response.content if getattr(b, 'type', None) == 'text'
        ).strip()
        registra_us_ia(cami='extraccio', model_ia=EXTRACCIO_MODEL,
                       usage=getattr(response, 'usage', None),
                       import_session=session, model=session.model,
                       created_by=session.creat_per)
    except Exception as e:
        _logging.getLogger(__name__).exception('Extracció W2: error a la crida Claude')
        registra_us_ia(cami='extraccio', model_ia=EXTRACCIO_MODEL,
                       import_session=session, model=session.model,
                       created_by=session.creat_per, ok=False, error=str(e))
        return Response({'error': f'Error a la crida d\'extracció: {e}'}, status=502)

    # Guarda de truncament: si Opus talla per límit de tokens, degradem amb gràcia
    # (no bloqueja; el JSON pot quedar incomplet i el gestiona el salvage de sota).
    if getattr(response, 'stop_reason', None) == 'max_tokens':
        avisos.append("Resposta d'extracció truncada pel límit de tokens; "
                      'resultat possiblement incomplet.')

    # Parse tolerant (Fase 1) amb salvage per fila.
    grading_status = {'status': 'ok', 'detail': ''}
    try:
        extracted = safe_json_parse(raw)
    except ValueError as e:
        salvaged = salvage_measurements(raw)
        if not salvaged:
            session.avisos = avisos + [f'Extracció: JSON il·legible ({e})']
            session.save(update_fields=['avisos', 'actualitzat_at'])
            return Response({'error': 'La IA no ha retornat dades llegibles',
                             'detail': str(e)}, status=422)
        extracted = {'measurements': salvaged}
        grading_status = {'status': 'error',
                          'detail': f'JSON global malformat; recuperats {len(salvaged)} POMs per fila. ({e})'}
        avisos.append(grading_status['detail'])

    measurements = extracted.get('measurements', []) or []

    # FIX D — la fila que la IA deixa caure en silenci. Si el document és un Excel, el parser
    # determinista n'ha comptat les files de POM encara que hagi abdicat de llegir-lo; si la IA
    # en torna menys, es diu. (Amb el Tate: 26 al document, 25 d'Opus, `JJ` perduda i cap avís.)
    avisos += _avis_files_perdudes(excel_meta.get('n_files_amb_codi') or 0, len(measurements))

    # Matching POM per fila.
    # N3 (DIAGNOSI_NOMENCLATURA_ALIES): customer del model → el matcher resol els àlies de
    # nomenclatura d'AQUEST client abans que per descripció. Si el model no en té, customer=None
    # (comportament previ: resol per descripció).
    import_customer = session.model.customer if session.model_id else None
    poms_extrets, stats = _match_rows(
        [
            {
                'codi_fitxa': msr.get('client_code') or msr.get('code') or '',
                'descripcio': msr.get('description') or '',
                # D3 · porta d'entrada del camí IA. El JSON del model pot donar el valor
                # com a número o com a cadena ("16,75"); fins ara passava verbatim i el
                # primer que el tocava era un `float()` dins de detect_grading.
                'values': {k: normalitza_cm(v)
                           for k, v in (msr.get('values') or {}).items()},
                # B2: tolerància del document (None si absent).
                'tol_minus': normalitza_cm(msr.get('tol_minus')),
                'tol_plus': normalitza_cm(msr.get('tol_plus')),
                # F4 · secció d'origen. `section` és OPCIONAL a l'esquema del prompt: les
                # fitxes d'una sola peça no en tenen, i una IA que no l'empleni no ha fallat.
                'seccio': (str(msr.get('section')).strip() or None
                           if msr.get('section') else None),
            }
            for msr in measurements
        ],
        import_customer,
        session.model,
    )
    avisos += _avisos_de_matching(stats)

    # SET-2/T8-ter — la proposta ja l'ha resolta `_match_rows` (un sol punt per als dos camins).
    _proposta = stats.get('proposta') or {}
    session.resultat = {**(session.resultat or {}), 'extraccio': extracted,
                        'proposta_peces': _proposta,
                        'grading_status': grading_status}
    session.poms_extrets = poms_extrets
    session.avisos = avisos
    session.estat = 'POMS'
    session.save(update_fields=['resultat', 'poms_extrets', 'avisos', 'estat', 'actualitzat_at'])

    # 1C-2b: suggeriment del mode dels valors (default del toggle al front). Es calcula sobre
    # els POMs amb match (identitat canònica) i sobre el run/base del DOCUMENT (extracted) —
    # mateix origen que les claus de `values`. Cosmètic → mai pot petar W2; default 'absoluts'.
    try:
        from fhort.pom.grading_utils import suggest_valors_mode
        vals_per_pom = {
            p['pom_master_id']: p['values']
            for p in poms_extrets
            if p.get('pom_master_id') and p.get('values')
        }
        suggested_valors_mode = suggest_valors_mode(
            vals_per_pom, extracted.get('base_size'), extracted.get('sizes') or [])
    except Exception:
        suggested_valors_mode = 'absoluts'

    return Response({
        'estat': session.estat,
        'poms_extrets': poms_extrets,
        'header': extracted.get('header') or {},
        'base_size': extracted.get('base_size'),
        'sizes': extracted.get('sizes') or [],
        'grading_status': grading_status,
        'avisos': avisos,
        'suggested_valors_mode': suggested_valors_mode,
        # SET-2/T8-ter — la proposta secció→peça i, sobretot, les seccions que NO tenen
        # peça: una absència que el pas 2 ha de poder DIR (crear la peça és un gest humà).
        'proposta_peces': _proposta,
        # F5 · l'informe de fulls surt del parser determinista, que sap comptar les pestanyes
        # encara que abdiqui de llegir-les. Buit per a PDF/imatge, que no tenen fulls.
        'fulls': excel_meta.get('fulls') or [],
        'full': excel_meta.get('full'),
    }, status=200)


def _candidats_de_codi(codi):
    """Els POMMaster del catàleg que es disputen un codi, serialitzats per a la UI.

    R1 · el 409 ha de portar els candidats. Fins ara la resposta deia NOMÉS el codi en
    conflicte i l'única sortida que li quedava al tècnic era sortir del wizard i anar al
    catàleg a mirar què hi havia. La vista ja feia la query per COMPTAR-los; aquí es
    serialitzen perquè la decisió (quin dels dos és el bo) es pugui prendre a la fila,
    sense perdre la governança: el backend segueix sense triar-ne cap.
    """
    from fhort.pom.models import POMMaster
    return [{
        'id': pm.id,
        'codi_client': pm.codi_client,
        # F1 · `nom_client` es queda perquè hi ha lectors que el volen CRU (saber si el TENANT
        # ha batejat el POM és una pregunta legítima), però la fila ara porta també el nom
        # RESOLT, que és el que s'ha de pintar. Sense ell, els candidats del 409 sortien amb el
        # codi pelat exactament igual que les files.
        'nom_client': pm.nom_client,
        'nom_en': _nom_resolt(pm) or '',
        'origen_import': pm.origen_import or '',
        'pendent_revisio': bool(pm.pendent_revisio),
        'actiu': bool(pm.actiu),
        # `select_related('pom_global')`: `_nom_resolt` hi entra per fila i sense això el 409
        # comprava una query per candidat.
    } for pm in POMMaster.objects.select_related('pom_global')
                                 .filter(codi_client=codi).order_by('id')]


def _capa_instancia_de(fila):
    """Els dos eixos d'identitat d'una fila, normalitzats. Absents → els de sempre.

    `''` d'instància és «la instància única» i `SLUG_DEFECTE` de capa és l'exterior: són els
    valors que l'import escrivia HARDCODEJATS abans de l'Onada 3, o sigui que una fila que no
    en digui res es comporta exactament com abans d'aquesta peça. Això és el que fa que la
    columna nova no toqui cap sessió existent ni cap camí que no la faci servir.

    Es normalitza a minúscules perquè els slugs són slugs: `'Left'` i `'left'` han de ser la
    MATEIXA instància, i si no ho fossin el detector deixaria passar una col·lisió real per
    una diferència de majúscules — que és el mode de fallada més lleig possible aquí (dues
    mesures a la mateixa cel·la, i la segona guanya en silenci al confirm).
    """
    from fhort.pom.models import MeasurementLayer
    capa = str((fila or {}).get('capa') or '').strip().lower() or MeasurementLayer.SLUG_DEFECTE
    instancia = str((fila or {}).get('instancia') or '').strip().lower()
    return capa, instancia


def _garment_de(fila, garment_sessio=''):
    """L'eix de PEÇA d'una fila: el seu si el declara, i si no el de la SESSIÓ.

    ── SET-2/T8-ter (2026-08-16) · **REOBERTURA CONSCIENT DE T8** (Agus, Patró C) ──────────
    T8 va decidir «un import = una prenda» i va posar el garment a `ImportSession`. La decisió
    era bona amb el que hi havia: llavors la fila no tenia transport per a cap eix propi. L'Onada
    3 (14/08) el va construir —capa i instància ja viatgen a la fila i sobreviuen del pas 2 al
    confirm— i amb ell la premissa de T8 ha caducat: el garment és **l'últim eix que queda a la
    sessió**, i baixa a la fila pel mateix camí que els altres dos.

    ⚠️ EL PREDICAT ÉS `is None`, NO LA FALSEDAT, i és la MATEIXA llei que `valor_efectiu`
    (`services_garment.py`) i que el `??` de `filaPas2.js`: `''` és una DECISIÓ («aquesta fila és
    de la mare») i l'absència és «no s'ha dit res» (→ la de la sessió). Amb `or` no es podria
    dir mai «mare» dins d'un import obert des d'una peça, que és exactament el gest que el
    desdoblament necessita: un document que porta la faldilla i el short s'obre des d'on sigui i
    cada fila diu de qui és.

    Una fila que no en declari cap es comporta EXACTAMENT com abans d'aquesta peça: rep el de la
    sessió, que és el que el confirm hi escrivia hardcodejat. Cap de les 16 sessions vives del
    corpus (totes amb `garment=''`) canvia de comportament.

    ── EL PRE-MARCAT HI ENTRA (16/08) · I VIU EN UN SOL LLOC ────────────────────────────────
    Entre la decisió de la persona i la de la sessió hi ha la PROPOSTA del document
    (`garment_proposat`, F2). Hi entra perquè el disseny la crida «pre-marcat» i un pre-marcat és
    un valor POSAT que es pot canviar, no una insinuació: la cel·la el mostra i enviar el pas
    l'accepta.

    ⚠️ I hi entra AQUÍ, no a cada consumidor. El 16/08 la regla només vivia al front i el
    detector del backend comparava com a mare una fila que la pantalla mostrava com a «Short»
    (captura 13:21): dues meitats de la mateixa pantalla parlant de files diferents. Amb la
    resolució en un sol punt, el detector, el confirm i el pas 3 no poden divergir.
    """
    g = (fila or {}).get('garment')
    if g is None:
        g = (fila or {}).get('garment_proposat')
    if g is None:
        return (garment_sessio or '').strip()
    return str(g).strip()


def _norm_seccio(text):
    """Forma comparable d'un rètol de secció o d'un nom de peça.

    Majúscules/minúscules i espais sobrers no distingeixen res: «SHORT», «Short» i « short »
    són el mateix rètol. Els accents SÍ que es conserven —no és el mateix «Llaçada» que
    «Llacada» per a un tècnic que escriu bé— i no cal desaccentuar per al cas real.
    """
    return ' '.join(str(text or '').strip().split()).casefold()


def _proposta_de_peca(model, poms):
    """Marca `garment_proposat` a cada fila segons la SECCIÓ del document. Torna la traça.

    ── SET-2/T8-ter · EL SUGGERIMENT NEIX D'ON JA ESTAVA ────────────────────────────────
    `seccio` no és cap camp nou: els DOS camins d'extracció ja la porten a la fila —el parser
    determinista la treu del rètol de bloc (`:530`) i la via IA del `section` de l'esquema
    (`:1745`)— i el confirm ja la persisteix a `BaseMeasurement.seccio` des de F3. L'única
    cosa que faltava era aparellar-la amb les peces del model.

    ⚠️ PROPOSA, NO DECIDEIX. `garment_proposat` és una clau SEPARADA de `garment`: la fila
    proposada segueix sense declarar eix propi i, si ningú la confirma, el confirm hi escriu
    el de la sessió com sempre. La decisió és de la persona (la columna del pas 2, F3) i el
    dia que el rètol del document canviï de nom, el pitjor que pot passar és que no es
    proposi res — mai que la fila aterri sola a una peça que ningú no ha triat.

    L'aparellament és per NOM o per CODI de la peça, en forma comparable, i té una segona
    passada per PARAULA SENCERA («SHORT MEASUREMENTS» → la peça «Short») perquè els documents
    reals titulen les seccions, no les etiqueten. Es demana ≥3 caràcters per no aparellar una
    peça dita «A» amb qualsevol secció que porti una a.

    Les seccions que NO troben peça es tornen a part: **és una absència que s'ha de DIR**. Un
    document amb una secció SHORT sobre un model sense peça Short vol dir que falta crear-la, i
    un silenci aquí deixaria set files aterrant a la mare sense que ningú se n'adonés.
    """
    from fhort.models_app.models import ModelGarment

    peces = list(ModelGarment.objects.filter(model=model).order_by('ordre', 'codi')) if model else []
    per_nom = {}
    for g in peces:
        for etiqueta in (g.nom, g.codi):
            clau = _norm_seccio(etiqueta)
            if clau:
                per_nom.setdefault(clau, g.codi)

    mapa, sense_peca = {}, []
    for p in (poms or []):
        seccio = _norm_seccio(p.get('seccio'))
        if not seccio:
            continue
        if seccio in mapa:
            p['garment_proposat'] = mapa[seccio]
            continue
        codi = per_nom.get(seccio)
        if codi is None:
            # Segona passada: el nom de la peça com a PARAULA SENCERA dins del rètol.
            mots = set(seccio.split())
            for clau, c in per_nom.items():
                if len(clau) >= 3 and clau in mots:
                    codi = c
                    break
        if codi is None:
            if p.get('seccio') not in sense_peca:
                sense_peca.append(p.get('seccio'))
            continue
        mapa[seccio] = codi
        p['garment_proposat'] = codi

    return {'mapa': mapa, 'seccions_sense_peca': sense_peca,
            'n_proposades': sum(1 for p in (poms or []) if p.get('garment_proposat'))}


def _identitat(fila, garment_sessio=''):
    """`(pom_master_id, capa, instancia, garment)` d'una fila resolta. La clau del detector.

    SET-2/T8-ter — el `garment` hi entra com a QUART element. La clau de `BaseMeasurement` ja
    és aquesta des de T2 (`('model','pom','capa','instancia','garment')`): el detector del
    wizard es limitava a mirar-ne tres perquè el quart era constant dins d'una tramesa. Ara que
    no ho és, mirar-ne tres tornaria a fabricar el mode de fallada que l'Onada 3 va tancar —dues
    files que voldrien ocupar cel·les DIFERENTS declarades com a col·lisió, i la resolució
    barrada sense motiu.

    ⚠️ `garment_sessio` és OBLIGATORI de pensar a cada cridador, no un defecte còmode: amb el
    defecte `''` i una sessió oberta a la 02, la identitat de lectura diria «mare» i la
    d'escriptura «02», i les dues cadenes deixarien de trobar-se. Tots els cridadors el passen
    (cens del 16/08: 8 punts, tots amb la sessió o el seu garment a l'abast).
    """
    capa, instancia = _capa_instancia_de(fila)
    return (fila.get('pom_master_id'), capa, instancia, _garment_de(fila, garment_sessio))


def _files_per_ordre(poms):
    """`{ordre: fila}` de `poms_extrets`.

    L'`ordre` és la identitat de FILA del wizard: neix a l'extracció (`:1342`), sobreviu a
    les resolucions del pas 2 (`_pla_de_resolucions` hi indexa) i és l'única clau que no
    col·lapsa quan dues files parlen del mateix POM. Les sessions velles i les muntades a mà
    poden no portar-lo; per a aquestes val la posició, que és el mateix nombre.
    """
    per_ordre = {}
    for i, p in enumerate(poms or []):
        ordre = p.get('ordre')
        per_ordre[i if ordre is None else ordre] = p
    return per_ordre


def _valors_de_les_mesures(session):
    """Les mesures del pas 3 indexades per IDENTITAT SENCERA `(pom, capa, instancia)`.

    🔑 ONADA 3 (14/08) · AQUESTA ERA LA MEITAT QUE FALTAVA. El pas 2 ja sabia dir que tres
    files són el mateix POM en tres instàncies, però la cadena de mesures seguia parlant per
    `pom_master_id` PELAT: `{pom_id: {talla: valor}}`. Amb la fitxa de la Brumà (B «at the
    top» 30 · BB «at the bottom» 31 · B1 «stretched out» 40) el diccionari en retenia UNA i
    les tres files s'escrivien amb el MATEIX valor, sense error i sense avís.

    **Les mesures VELLES no es perden ni canvien de sentit.** Una mesura desada abans
    d'aquesta peça (o enviada pel front d'avui) no porta identitat, i aquí se li dona la de
    les FILES que reclamen el seu POM: si n'hi ha una, la seva; si n'hi ha diverses, la de
    TOTES —que és exactament el que passava abans, quan cada fila llegia `valors[pid]`—.
    Així una sessió a mig fer es confirma igual que ahir i el guard de no-regressió és el
    primer que passa, no l'últim.
    """
    _gs = (getattr(session, 'garment', '') or '')
    poms = list(session.poms_extrets or [])
    idents_per_pom = {}
    for p in poms:
        pid = p.get('pom_master_id')
        if pid:
            idents_per_pom.setdefault(int(pid), []).append(_identitat(p, _gs))

    valors = {}
    for m in (session.resultat or {}).get('mesures', []):
        talla = m.get('talla_label')
        if talla in (None, ''):
            continue
        try:
            pid = int(m['pom_master_id'])
        except (KeyError, TypeError, ValueError):
            continue
        if m.get('capa') is not None or m.get('instancia') is not None:
            # SET-2/T8-ter — la mesura porta la seva peça quan el pas 3 la declara; si no, la de
            # la sessió, que és el que la fila d'on ve també haurà rebut.
            idents = [(pid,) + _capa_instancia_de(m) + (_garment_de(m, _gs),)]
        else:
            idents = (idents_per_pom.get(pid)
                      or [(pid,) + _capa_instancia_de(None) + (_gs.strip(),)])
        for ident in idents:
            valors.setdefault(ident, {})[talla] = m.get('valor')
    return valors


def _valors_per_pom(valors, avisos=None):
    """La vista PER POM de la taula, per als consumidors que graduen.

    El grading és del POM: `derive_rules_from_fitxa` deriva UNA regla per `pom_id` i el
    contenidor del client tampoc no coneix instàncies. Quan un POM es reparteix en més d'una
    fila, la regla surt de la fila CANÒNICA (capa per defecte, instància única) i, si no
    n'hi ha cap, de la primera de la taula. Les germanes NO deriven regla pròpia: inventar-ne
    una per instància seria fabricar catàleg que ningú no ha decidit — i el valor de cada
    germana sí que arriba sencer al seu `BaseMeasurement`, que és on viu.
    """
    from fhort.pom.models import MeasurementLayer
    tria = {}      # pid → (és_canònica, files)
    n_files = {}   # pid → quantes identitats el mesuren
    # SET-2/T8-ter — la identitat ja porta el `garment` i aquí encara es col·lapsa per POM sol.
    # És el gra que aquesta vista declara des de sempre (el catàleg del client no coneix
    # instàncies) i NO canvia amb aquesta peça; el que canvia és que ara el col·lapse també es
    # menja la frontera de peça, i això sí que és una pèrdua nova. Es tanca a F6, amb acta i
    # paritat; aquí es DIU al desempaquetar perquè qui hi passi ho vegi.
    for (pid, capa, instancia, _garment), files in valors.items():
        canonica = (capa == MeasurementLayer.SLUG_DEFECTE and instancia == '')
        n_files[pid] = n_files.get(pid, 0) + 1
        if pid not in tria or (canonica and not tria[pid][0]):
            tria[pid] = (canonica, files)
    if avisos is not None:
        for pid, n in n_files.items():
            if n > 1:
                avisos.append(
                    f"El POM #{pid} es mesura en {n} files (capes/instàncies diferents): la "
                    "graduació es deriva d'una sola —la canònica— i les germanes desen el seu "
                    "valor sense regla pròpia.")
    return {pid: files for pid, (_canonica, files) in tria.items()}


def _descriu_fila(fila, pm, capa, instancia, garment):
    """Com s'anomena una fila del pas 2 quan cal DIR-LA en un conflicte.

    El text que llegeix la persona és «M · Short · exterior · única», no un `ordre` pelat: el
    conflicte s'ha de poder resoldre mirant les dues files, i per mirar-les cal saber quines són.
    """
    fila = fila or {}
    return {
        'ordre': fila.get('ordre'),
        'codi': fila.get('codi_fitxa') or (pm.codi_client if pm is not None else '') or '',
        'nom': (fila.get('pom_nom') or (pm.nom_client if pm is not None else '')
                or fila.get('descripcio') or ''),
        'capa': capa if capa is not None else _capa_instancia_de(fila)[0],
        'instancia': instancia if instancia is not None else _capa_instancia_de(fila)[1],
        'garment': garment,
    }


def _pla_de_resolucions(poms, brut, garment_sessio=''):
    """R2 · valida les `resolucions` de fila i en retorna el PLA, o els errors per fila.

    Una resolució és la decisió humana sobre UNA fila del pas 2:
      · `{'ordre': n, 'accio': 'vincula', 'pom_master_id': id}` → la fila pren aquest POM
        del catàleg (ha d'existir i ser actiu, i no el pot tenir ja una altra fila activa).
      · `{'ordre': n, 'accio': 'crea', 'codi': 'E', 'nom': '...'}` → POMMaster tenant-only
        amb el codi i el nom DONATS. Si el codi ja existeix al catàleg NO es crea res: es
        torna l'error amb els candidats (mateixa forma que el 409 de R1). El wizard mai pot
        fabricar un duplicat nou — que és l'estat que va provocar l'incident.

    Res s'escriu aquí: validar i escriure van separats a propòsit, perquè una sola resolució
    dolenta no deixi les bones a mitges (la resposta d'error és per fila i el frontend les
    reenvia senceres).
    """
    from fhort.pom.models import POMMaster

    per_ordre = {p.get('ordre'): p for p in poms if p.get('ordre') is not None}
    ordres_resolts = {r.get('ordre') for r in brut}
    #: IDENTITAT → fila que ja la té. Les files que aquesta tramesa resol no compten: si es
    #: re-vincula una fila, la identitat que tenia queda lliure.
    #:
    #: 🔑 LA CLAU ÉS LA IDENTITAT SENCERA, NO EL POM PELAT (Onada 3, 14/08). Era
    #: `{pom_master_id: ordre}`, i amb això la fitxa BROWNIE BRUMA/RUFFLES no es podia
    #: importar: B «at the top», BB «at the bottom» i B1 «stretched out» són el MATEIX POM B
    #: mesurat en tres instàncies, i el detector les llegia com dues col·lisions. Una mesura
    #: d'aquesta casa s'identifica per `(pom, capa, instancia, garment)` —el motor ja hi
    #: indexa (`:2087`)— i el `garment` no entra a la clau perquè és de la SESSIÓ: un import
    #: és una prenda (T8), o sigui que totes les files d'aquesta tramesa el comparteixen i
    #: afegir-lo no distingiria res.
    #:
    #: La col·lisió no desapareix: es fa PRECISA. Dues files amb la mateixa identitat sencera
    #: segueixen sent l'error de sempre, i han de ser-ho — són dues mesures que voldrien
    #: ocupar la mateixa cel·la i el confirm en perdria una en silenci.
    presos = {_identitat(p, garment_sessio): p for p in poms
              if p.get('pom_master_id') and p.get('actiu')
              and p.get('ordre') not in ordres_resolts}

    pla, errors, codis_nous = [], [], {}
    for r in brut:
        ordre = r.get('ordre')
        accio = str(r.get('accio') or '').strip()
        fila = per_ordre.get(ordre)
        if fila is None:
            errors.append({'ordre': ordre, 'error': 'fila_no_trobada'})
            continue
        if accio == 'vincula':
            pid = r.get('pom_master_id')
            pm = (POMMaster.objects.filter(id=int(pid), actiu=True).first()
                  if str(pid).isdigit() else None)
            if pm is None:
                errors.append({'ordre': ordre, 'error': 'pom_no_valid', 'pom_master_id': pid})
                continue
            # La identitat que la PERSONA ha triat per a aquesta fila. Absent = la de sempre
            # (exterior, instància única): una resolució que no parli d'instàncies es comporta
            # exactament com abans d'aquesta peça.
            capa, instancia = _capa_instancia_de(r)
            # ⚠️ LA PEÇA DE LA FILA MANA SOBRE LA DE LA SESSIÓ (QA Agus 16/08, captura 13:21).
            # Una resolució diu QUIN POM és aquesta fila; de QUI és ho diu la COLUMNA, que ja ha
            # desat la decisió a la fila unes línies més amunt (`files_garment`). Fent el fallback
            # a la sessió, el detector comparava com a mare una fila que la pantalla mostrava com
            # a «Short», i la Brumà donava un conflicte que no es podia verificar mirant-la.
            garment = _garment_de(r, _garment_de(fila, garment_sessio))
            ident = (pm.id, capa, instancia, garment)
            if ident in presos:
                ocupada = presos[ident]
                errors.append({'ordre': ordre, 'error': 'pom_ja_usat', 'pom_master_id': pm.id,
                               'capa': capa, 'instancia': instancia, 'garment': garment,
                               'ordre_ocupat': ocupada.get('ordre'),
                               # EL CONFLICTE ES RESOL VEIENT-LO, NO ENDEVINANT-LO. «Un POM no pot
                               # ser dues files» és una acusació sense judici: no diu per què
                               # AQUESTA vegada, i amb la peça i la instància a la identitat les
                               # dues files poden diferir en tres eixos. Van les DUES, amb nom.
                               'aquesta': _descriu_fila(fila, pm, capa, instancia, garment),
                               'ocupada': _descriu_fila(ocupada, None, None, None,
                                                        _garment_de(ocupada, garment_sessio))})
                continue
            presos[ident] = dict(fila or {}, ordre=ordre, capa=capa, instancia=instancia,
                                 garment=garment, pom_master_id=pm.id, pom_codi=pm.codi_client)
            pla.append({'fila': fila, 'accio': 'vincula', 'pom': pm,
                        'capa': capa, 'instancia': instancia, 'garment': garment})
        elif accio == 'crea':
            codi = str(r.get('codi') or '').strip()
            if not codi:
                errors.append({'ordre': ordre, 'error': 'codi_buit'})
                continue
            if POMMaster.objects.filter(codi_client=codi).exists():
                errors.append({'ordre': ordre, 'error': 'codi_existent', 'codi': codi,
                               'candidats': _candidats_de_codi(codi)})
                continue
            if codi.upper() in codis_nous:
                errors.append({'ordre': ordre, 'error': 'codi_repetit', 'codi': codi,
                               'ordre_ocupat': codis_nous[codi.upper()]})
                continue
            codis_nous[codi.upper()] = ordre
            # La identitat viatja també amb el `crea`: un POM nou pot néixer ja partit (la
            # fitxa el porta esquerre i dret). El detector de `crea` segueix sent per CODI i
            # no per identitat a posta — dues files que creen el mateix codi són un conflicte
            # de CATÀLEG, i això no depèn de quina instància mesuri cadascuna.
            capa, instancia = _capa_instancia_de(r)
            pla.append({'fila': fila, 'accio': 'crea', 'codi': codi,
                        'nom': str(r.get('nom') or '').strip() or codi,
                        'capa': capa, 'instancia': instancia,
                        'garment': _garment_de(r, garment_sessio)})
        else:
            errors.append({'ordre': ordre, 'error': 'accio_desconeguda', 'accio': accio})
    return pla, errors


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def import_session_poms_view(request, token):
    """
    PATCH /api/v1/import-sessions/<token>/poms/  (Pas W2 — confirmació de POMs)

    Rep poms_confirmats (llista de pom_master_id actius). Marca actiu per cada POM extret;
    els pom_master_id confirmats que no hi siguin (afegits manualment del catàleg) s'incorporen.
    P2 · `files_confirmades` (opcional, llista d'`ordre`) fa la mateixa pregunta PER FILA i, si
    hi és, mana sobre `poms_confirmats` per a l'estat `actiu`: amb germanes, un POM no pot
    respondre per totes les seves files.
    Rep també poms_tenant_only (llista d'ordres de files NO_MATCH que el tècnic vol crear com
    a POMMaster tenant-only: pom_global=None, codi_client=codi_fitxa). estat→'MESURES'.

    R2 · `resolucions` (opcional) — la decisió de fila. `[{ordre, accio, pom_master_id?, codi?,
    nom?}]` amb accio 'vincula' (la fila pren un POM del catàleg) o 'crea' (POMMaster tenant-only
    amb el codi i el nom DONATS, mai el codi_fitxa a cegues). S'apliquen dins del MATEIX atomic i
    manen sobre `poms_tenant_only` per a la mateixa fila. El contracte antic (poms_confirmats +
    poms_tenant_only, sense `resolucions`) segueix funcionant idèntic.

    409 `codi_duplicat`: el catàleg de tenant NO té cap constraint d'unicitat sobre
    (pom_global, codi_client) — `pom/models.py:183-185` no en declara cap. Dos POMMaster
    tenant-only amb el mateix codi són, doncs, un estat LEGAL de la BD, i el `get_or_create`
    que hi havia aquí hi petava amb `MultipleObjectsReturned` → 500 i sessió descartada (QA
    real a PROD, 27/07/2026). Ara la vista ho detecta ABANS d'escriure res i ho diu; resoldre
    el duplicat del catàleg és una decisió humana, no una que el wizard pugui prendre sola.
    """
    from django.db import transaction

    from fhort.models_app.models import ImportSession
    from fhort.pom.models import POMMaster, POMCategory

    session = ImportSession.objects.filter(token=token).first()
    if not session:
        return Response({'error': 'Sessió no trobada'}, status=404)

    confirmats = [int(x) for x in (request.data.get('poms_confirmats') or []) if str(x).isdigit()]
    confirmats_set = set(confirmats)
    tenant_only_ordres = {
        int(x) for x in (request.data.get('poms_tenant_only') or [])
        if str(x).lstrip('-').isdigit()
    }

    poms = list(session.poms_extrets or [])

    # R2 · resolucions de fila (opcional). Manen sobre `poms_tenant_only`: una fila que el
    # tècnic ha resolt explícitament ja no passa pel camí a cegues del codi_fitxa.
    resolucions_brut = [r for r in (request.data.get('resolucions') or []) if isinstance(r, dict)]
    tenant_only_ordres -= {r.get('ordre') for r in resolucions_brut}

    # ── PORTA 409, ABANS DE TOCAR RES. Les files que el tècnic vol crear com a tenant-only,
    # amb el seu codi. Es recullen TOTS els codis que ja tenen 2+ POMMaster tenant-only al
    # catàleg: petar al primer obligaria a descobrir-los d'un en un, un import per duplicat.
    files_tenant_only = [
        p for p in poms
        if not p.get('pom_master_id')
        and p.get('ordre') in tenant_only_ordres
        and (p.get('codi_fitxa') or '').strip()
    ] if tenant_only_ordres else []

    duplicats = sorted({
        codi for codi in {(p.get('codi_fitxa') or '').strip() for p in files_tenant_only}
        if POMMaster.objects.filter(pom_global=None, codi_client=codi).count() > 1
    })
    if duplicats:
        # R1 · el 409 porta els CANDIDATS de cada codi. Sense ells la UI només podia enviar
        # el tècnic al catàleg; amb ells el conflicte es resol a la fila (PATCH `resolucions`).
        return Response({'error': 'codi_duplicat', 'codis': duplicats,
                         'candidats': {c: _candidats_de_codi(c) for c in duplicats}}, status=409)

    existents = {p.get('pom_master_id') for p in poms if p.get('pom_master_id')}
    # P2 · QUINES FILES ENTREN A LA TAULA — la pregunta és per FILA quan el client la sap fer.
    #
    # `poms_confirmats` és una llista d'IDs de POM, i mentre un POM no podia ocupar més d'una
    # fila això era la mateixa cosa. Des de l'Onada 3 no ho és: amb tres germanes, desmarcar-ne
    # una les desmarcava TOTES TRES. `files_confirmades` (llista d'`ordre`) fa la mateixa
    # pregunta a la fila. Absent = el camí de sempre, byte a byte; `[]` és una decisió («cap»),
    # que no és el mateix que no dir-ne res.
    #
    # No substitueix `poms_confirmats`: aquell té una segona feina —incorporar POMs del catàleg
    # que el document no menciona— i aquells encara no tenen fila amb què demanar-se.
    # ── SET-2/T8-ter · LA PEÇA DE CADA FILA, DECIDIDA PER LA PERSONA ────────────────────────
    # `files_garment` és `[{ordre, garment}]` i és la porta de la COLUMNA del pas 2. Va a part
    # de `resolucions` a posta: aquelles resolen QUIN POM és una fila que no en tenia, i la
    # immensa majoria de files ja el tenen —el que la persona hi canvia és de qui SÓN—. Fer-ho
    # passar per `resolucions` obligaria a re-declarar el vincle per moure una fila de peça.
    #
    # ⚠️ Es valida contra les peces REALS del model, com la porta del cribratge: un codi que no
    # és cap peça d'aquest model és un 400 i no un silenci. Escriure a la mare quan la persona
    # ha dit «02» és exactament el dany que aquest tram tanca, fabricat des de l'altra banda.
    # `''` és sempre vàlid: és la mare.
    files_garment = request.data.get('files_garment')
    if isinstance(files_garment, list) and files_garment:
        from fhort.models_app.models import ModelGarment
        _codis_valids = set(
            ModelGarment.objects.filter(model=session.model).values_list('codi', flat=True))
        _per_ordre = _files_per_ordre(poms)
        _desconeguts, _ordres_no_trobats = set(), []
        for entrada in files_garment:
            if not isinstance(entrada, dict):
                continue
            g = str(entrada.get('garment') or '').strip()
            if g and g not in _codis_valids:
                _desconeguts.add(g)
                continue
            fila = _per_ordre.get(entrada.get('ordre'))
            if fila is None:
                _ordres_no_trobats.append(entrada.get('ordre'))
                continue
            fila['garment'] = g
        if _desconeguts:
            return Response({'error': 'garment_desconegut', 'codis': sorted(_desconeguts),
                             'message': (f"El model {session.model_id} no té cap prenda "
                                         f"«{', '.join(sorted(_desconeguts))}».")}, status=400)
        if _ordres_no_trobats:
            return Response({'error': 'files_inexistents', 'ordres': _ordres_no_trobats},
                            status=400)

    files_confirmades = request.data.get('files_confirmades')
    ordres_confirmats = ({int(x) for x in files_confirmades if str(x).lstrip('-').isdigit()}
                         if isinstance(files_confirmades, list) else None)
    for i, p in enumerate(poms):
        if p.get('pom_master_id'):
            p['actiu'] = ((p.get('ordre', i) in ordres_confirmats)
                          if ordres_confirmats is not None
                          else p['pom_master_id'] in confirmats_set)

    # PORTA de les resolucions, també abans de tocar res: si una sola falla, no n'entra CAP.
    # Els errors van per `ordre` perquè el wizard els pugui pintar a la fila que toca i
    # reenviar la tramesa sencera sense que el tècnic hagi de refer les que ja eren bones.
    pla_resolucions, errors_resolucions = _pla_de_resolucions(
        poms, resolucions_brut, session.garment or '')
    if errors_resolucions:
        return Response({'error': 'resolucions_invalides', 'errors': errors_resolucions},
                        status=409)

    # Cos MUTADOR: o hi entra tot, o no hi entra res. Sense l'atomic, un POMMaster creat i una
    # sessió no desada deixaven catàleg brut sense cap fila que hi apuntés.
    with transaction.atomic():
        categoria_default = ((POMCategory.objects.filter(actiu=True)
                              .order_by('display_order', 'codi').first())
                             if (files_tenant_only or pla_resolucions) else None)

        # R2 · les decisions de fila. Dins del MATEIX atomic que la resta.
        for r in pla_resolucions:
            fila = r['fila']
            if r['accio'] == 'crea':
                pm = POMMaster.objects.create(
                    pom_global=None,
                    codi_client=r['codi'],
                    nom_client=r['nom'],
                    actiu=True,
                    categoria=categoria_default,
                    pendent_revisio=True,
                    origen_import=str(session.token),
                    notes=f'Creat des del pas 2 de l\'import, fitxa {session.token}',
                )
            else:
                pm = r['pom']
            fila['pom_master_id'] = pm.id
            fila['pom_codi'] = pm.codi_client
            fila['pom_nom'] = pm.nom_client
            fila['match_type'] = 'tenant_only' if r['accio'] == 'crea' else 'manual'
            fila['confidence'] = 'TENANT_ONLY' if r['accio'] == 'crea' else 'HIGH'
            fila['actiu'] = True
            # ONADA 3 · LA IDENTITAT QUEDA A LA FILA. Sense aquestes dues línies el detector
            # ja no col·lisionaria però el confirm tornaria a escriure les tres files de la
            # Brumà al mateix `(pom, exterior, '')`: l'`update_or_create` en desaria una i les
            # altres dues es perdrien sense dir res. La decisió de la persona ha de sobreviure
            # del pas 2 al confirm, i és aquí on hi puja.
            fila['capa'] = r['capa']
            fila['instancia'] = r['instancia']
            # SET-2/T8-ter — i la PEÇA, pel mateix argument literal que les altres dues: la
            # decisió de la persona ha de sobreviure del pas 2 al confirm. Sense aquesta línia
            # el confirm tornaria a llegir el garment de la sessió i les files de les altres
            # peces aterrarien totes al mateix contenidor.
            fila['garment'] = r['garment']
            existents.add(pm.id)

        # POMs sense match triats pel tècnic → crear (o reutilitzar) POMMaster tenant-only.
        if files_tenant_only:
            for p in files_tenant_only:
                codi = (p.get('codi_fitxa') or '').strip()
                descripcio = (p.get('descripcio') or '').strip()
                # count==1 → reutilitzar · count==0 → crear. El cas count>1 ja ha sortit per
                # la porta 409 de dalt, i per això aquí NO hi ha get_or_create.
                pm = POMMaster.objects.filter(pom_global=None, codi_client=codi).first()
                if pm is None:
                    pm = POMMaster.objects.create(
                        pom_global=None,
                        codi_client=codi,
                        nom_client=descripcio or codi,
                        actiu=True,
                        categoria=categoria_default,
                        pendent_revisio=True,
                        origen_import=str(session.token),
                        notes=f'Creat automàticament per import, fitxa {session.token}',
                    )
                p['pom_master_id'] = pm.id
                p['pom_codi'] = pm.codi_client
                p['pom_nom'] = pm.nom_client
                p['match_type'] = 'tenant_only'
                p['confidence'] = 'TENANT_ONLY'
                p['actiu'] = True
                existents.add(pm.id)

        # Afegir POMs confirmats que no eren a l'extracció (afegits manualment).
        for pid in confirmats_set - existents:
            pm = POMMaster.objects.filter(id=pid, actiu=True).first()
            if not pm:
                continue
            poms.append({
                'codi_fitxa': '',
                'descripcio': pm.nom_client or '',
                'pom_master_id': pm.id,
                'pom_codi': pm.codi_client,
                'pom_nom': pm.nom_client,
                'match_type': 'manual',
                'confidence': 'HIGH',
                'values': {},
                'actiu': True,
                'ordre': len(poms),
            })

        session.poms_extrets = poms
        session.estat = 'MESURES'
        session.save(update_fields=['poms_extrets', 'estat', 'actualitzat_at'])

    actius = [p for p in poms if p.get('actiu')]
    return Response({'ok': True, 'estat': session.estat,
                     'poms_actius': len(actius), 'poms_extrets': poms}, status=200)


# ═══════════════════════════ W3 — Mesures ═══════════════════════════
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_session_grading_preview_view(request, token):
    """
    POST /api/v1/import-sessions/<token>/grading-preview/  (Pas W3 — preview de grading)

    Calcula el grading SENSE persistir (reutilitza el motor via preview_graded_specs) per
    omplir talles buides a la taula del wizard. NO crea SizeFitting/GradedSpec — això és
    feina del desament definitiu (W5). Rep base_values {pom_master_id: valor}.
    """
    from fhort.models_app.models import ImportSession
    from fhort.pom.services import preview_graded_specs

    session = ImportSession.objects.filter(token=token).select_related('model').first()
    if not session:
        return Response({'error': 'Sessió no trobada'}, status=404)
    model = session.model
    if not model:
        return Response({'error': 'La sessió no té model associat'}, status=400)
    if _efectiu(session, 'grading_rule_set') is None:
        return Response({'error': 'El model no té GradingRuleSet configurat', 'grading': {}}, status=400)

    # SET-2/T8 — l'eix entra a la CLAU, que és per on el motor el llegeix. `_regla_de` resol
    # la regla de la peça amb herència de la mare; una clau escalar (la d'abans) hauria
    # ensenyat sempre la regla de la mare, i el preview ha de dir el mateix que el generador.
    garment = session.garment or ''
    raw = request.data.get('base_values') or {}
    base_values = {}
    # ONADA 3 · LA TERCERA PORTA. `base_values` accepta les DUES formes:
    #   · `{pom_master_id: valor}` — el contracte d'avui, intacte. Els eixos són els de sempre.
    #   · `[{ordre, valor}]` — la taula parlant per FILA. Un objecte JSON no pot tenir una clau
    #     composta, i per això la forma nova és una LLISTA: és l'única manera que tres files
    #     del mateix POM hi càpiguen. Els eixos, com al pas 3, es HEREDEN de la fila.
    # La resposta es torna amb la MATEIXA clau amb què s'ha preguntat, i ho declara a `clau`.
    clau = 'ordre' if isinstance(raw, list) else 'pom_master_id'
    ordre_de_ident = {}
    if clau == 'ordre':
        per_ordre = _files_per_ordre(session.poms_extrets)
        for entrada in raw:
            if not isinstance(entrada, dict):
                continue
            fila = per_ordre.get(entrada.get('ordre'))
            v = entrada.get('valor')
            if fila is None or not fila.get('pom_master_id') or v in (None, ''):
                continue
            capa, instancia = _capa_instancia_de(fila)
            try:
                base_values[(int(fila['pom_master_id']), capa, instancia, garment)] = float(v)
            except (TypeError, ValueError):
                continue
            ordre_de_ident[(int(fila['pom_master_id']), capa, instancia)] = entrada.get('ordre')
    else:
        for k, v in raw.items():
            if not str(k).isdigit() or v in (None, ''):
                continue
            try:
                base_values[(int(k), MeasurementLayer.SLUG_DEFECTE, '', garment)] = float(v)
            except (TypeError, ValueError):
                continue

    grading_avisos: list[str] = []
    grading = preview_graded_specs(model, base_values, warnings=grading_avisos)
    # Claus a string per a JSON consistent al frontend.
    #
    # C3/B5 — el motor ja indexa per la identitat sencera `(pom_id, capa, instancia)`, però
    # AQUEST CONTRACTE NO ES TOCA: el cos que arriba és un objecte JSON `{pom_id: valor}` i un
    # objecte JSON no pot expressar una clau composta, ni a l'entrada ni a la sortida. El
    # `base_values` que hi entra és escalar i `preview_graded_specs` el normalitza a la mesura
    # única del POM ('exterior', ''); aquí es desfà la identitat per tornar al `pom_id` que el
    # frontend espera. Avui és una equivalència exacta —les comportes de C1/C1-ins garanteixen
    # una sola germana per POM— i per això el col·lapse no perd res.
    # Fer créixer aquest payload és INTERFÍCIE i va a C4 amb maqueta (llei 3c.5).
    # SET-2/T6a — la clau que arriba té QUATRE trams (el `garment` darrere de la instància).
    # SET-2/T8 — i el col·lapse a `pom_id` SEGUEIX SENT EXACTE, ara per una altra raó: **un
    # import = una prenda**, o sigui que totes les claus d'aquesta resposta porten el MATEIX
    # garment (el de la sessió) i cap parella no pot xocar. El que abans el garantia era la
    # comporta; ara ho garanteix el disseny. Fer créixer el payload seguiria sent INTERFÍCIE
    # (llei 3c.5), i aquest camí ja no la necessita.
    # ONADA 3 — …i quan la pregunta ha vingut per FILA, la resposta hi torna per FILA: el
    # col·lapse a `pom_id` és precisament el que faria que les tres files de la Brumà
    # s'omplissin totes amb la graduació d'una sola.
    if clau == 'ordre':
        grading = {str(ordre_de_ident[(pom_id, capa, instancia)]): row
                   for (pom_id, capa, instancia, _garment), row in grading.items()
                   if (pom_id, capa, instancia) in ordre_de_ident}
    else:
        grading = {str(pom_id): row
                   for (pom_id, _capa, _instancia, _garment), row in grading.items()}
    return Response({'grading': grading, 'clau': clau,
                     'base_size': _efectiu(session, 'base_size_label'),
                     'size_run': (_efectiu(session, 'size_run_model') or '').split('·'),
                     'avisos': grading_avisos}, status=200)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def import_session_mesures_view(request, token):
    """
    PATCH /api/v1/import-sessions/<token>/mesures/  (Pas W3 — desa valors de la taula)

    Rep mesures `[{ordre, talla_label, valor}]` — o `[{pom_master_id, …}]`, el contracte
    d'avui. Desa a session.resultat['mesures']. estat→'MESURES_OK'.

    🔑 ONADA 3 · LA TAULA PARLA PER FILA, NO PER POM. `ordre` és la identitat de fila del
    wizard i és l'única clau que no col·lapsa quan tres files són el mateix POM en tres
    instàncies (la fitxa de la Brumà). Els EIXOS NO ELS DIU AQUEST PAYLOAD: es HEREDEN de la
    fila, que és on el pas 2 va desar la decisió de la persona. Si el pas 3 els pogués tornar
    a declarar hi hauria dues fonts de veritat per a la mateixa cosa i, el dia que
    discrepessin, guanyaria l'última a escriure —el mode de fallada més lleig d'aquí.

    Sense `ordre` el comportament és EXACTAMENT el d'avui: la mesura es desa amb el
    `pom_master_id` pelat i els lectors li donen la identitat de les files que reclamen aquell
    POM (`_valors_de_les_mesures`). Una sessió a mig fer amb el front vell es confirma igual.
    """
    from fhort.models_app.models import ImportSession

    session = ImportSession.objects.filter(token=token).first()
    if not session:
        return Response({'error': 'Sessió no trobada'}, status=404)

    per_ordre = _files_per_ordre(session.poms_extrets)
    mesures = request.data.get('mesures') or []
    # Normalitza a llista neta de {pom_master_id, talla_label, valor} (+ identitat de fila
    # quan la tramesa parla per `ordre`).
    net = []
    for m in mesures:
        talla = m.get('talla_label')
        valor = m.get('valor')
        if talla in (None, ''):
            continue
        ordre = m.get('ordre')
        fila = per_ordre.get(ordre) if ordre is not None else None
        if ordre is not None and fila is None:
            # Un `ordre` que no és de cap fila no pot inventar-se una mesura sense identitat:
            # es descarta aquí i no al pas 5, on ja no se sabria de qui parlava.
            continue
        pid = fila.get('pom_master_id') if fila is not None else m.get('pom_master_id')
        if pid is None:
            continue
        # D3 · porta d'entrada del camí ENGANXAT. Arribava el que el navegador hagués
        # posat al JSON (el front fa `parseFloat`, que davant d'un "3,5" enganxat d'un
        # Excel europeu dona 3). `normalitza_cm` entén la coma i talla a 0,1 mm.
        entrada = {'pom_master_id': pid, 'talla_label': talla,
                   'valor': normalitza_cm(valor)}
        if fila is not None:
            capa, instancia = _capa_instancia_de(fila)
            # SET-2/T8-ter — i la PEÇA de la fila, pel mateix argument que els altres dos eixos:
            # els EIXOS NO ELS DIU AQUEST PAYLOAD, s'hereten de la fila (que és on el pas 2 va
            # desar la decisió). Sense això una mesura d'una fila decidida a la 02 es desava amb
            # el garment de la SESSIÓ i el confirm no la trobava: la cel·la quedava buida sense
            # que ningú petés — el mode de fallada que aquest tram persegueix, fabricat al pas 3.
            entrada.update({'ordre': ordre, 'capa': capa, 'instancia': instancia,
                            'garment': _garment_de(fila, session.garment or '')})
        net.append(entrada)

    session.resultat = {**(session.resultat or {}), 'mesures': net}
    # 1C-2a: si el wizard declara el mode dels valors (absoluts/deltes), desar-lo per al W5.
    valors_mode = request.data.get('valors_mode')
    if valors_mode in ('absoluts', 'deltes'):
        session.resultat = {**session.resultat, 'valors_mode': valors_mode}
    session.estat = 'MESURES_OK'
    session.save(update_fields=['resultat', 'estat', 'actualitzat_at'])

    return Response({'ok': True, 'estat': session.estat, 'n_valors': len(net)}, status=200)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_session_library_prefill_view(request, token):
    """
    POST /api/v1/import-sessions/<token>/library-prefill/  (1C-3 — pont fitxa → Size Library)

    Construeix el prefill ENRIQUIT per a la Size Library des de l'extracció ja feta: run + base
    + target + POMs amb els seus valors per talla, en ABSOLUTS. Si la fitxa era en mode 'deltes',
    converteix amb deltes_a_absoluts (1C-2a) abans d'enviar (el camí Library deriva amb
    detect_grading, que espera absoluts). NO crea res: només llegeix la sessió. Robust: degrada.
    """
    from fhort.models_app.models import ImportSession

    session = ImportSession.objects.filter(token=token).select_related('model').first()
    if not session:
        return Response({'error': 'Sessió no trobada'}, status=404)
    model = session.model

    # valors {pid:{talla:valor}} des de les mesures desades (els valors EDITATS al W3).
    #
    # ONADA 3 — la taula ja parla per identitat sencera, però la Size Library NO: el seu
    # prefill viatja per `pom_codi` (`find_pom_master` matcheja `codi_client`) i no té eix de
    # capa ni d'instància. La taula hi entra, doncs, PER POM, i és el mateix col·lapse que
    # feia abans aquest bucle: un POM repartit en tres files hi porta la canònica i les
    # germanes no hi són. Créixer aquest contracte és INTERFÍCIE de la Library (llei 3c.5) i
    # té la seva pròpia decisió; el que aquí importa és que el col·lapse ara és EXPLÍCIT.
    valors = _valors_per_pom(_valors_de_les_mesures(session))

    # SET-2/T8 — run i base de la PEÇA de destí (efectius): és la seva taula la que viatja
    # cap a la Size Library, i amb els de la mare el run proposat seria d'una altra prenda.
    base_size = ((_efectiu(session, 'base_size_label') if model else '') or '').strip()
    run = [s.strip() for s in ((_efectiu(session, 'size_run_model') if model else '') or '')
           .replace(';', '·').split('·') if s.strip()]

    # Mode deltes → absoluts ABANS d'enviar (reusa 1C-2a; una sola font de conversió).
    if ((session.resultat or {}).get('valors_mode') or 'absoluts') == 'deltes' and run and base_size:
        from fhort.pom.grading_utils import deltes_a_absoluts
        valors = deltes_a_absoluts(valors, base_size, run)

    # codi_client per pom des del CATÀLEG (font autoritativa per pid), no de la còpia
    # serialitzada a poms_extrets. find_pom_master al camí Library matcheja codi_client__iexact
    # (Strategy 1, exact_code HIGH) → round-trip garantit al MATEIX POMMaster.
    from fhort.pom.models import POMMaster
    codi_by_pid = {
        pm.id: (pm.codi_client or '')
        for pm in POMMaster.objects.filter(id__in=list(valors.keys()))
    }

    poms = []
    for pid, vals in valors.items():
        net = {k: v for k, v in (vals or {}).items() if v not in (None, '')}
        if net:
            poms.append({'pom_codi': codi_by_pid.get(pid) or '', 'valors': net})

    target_codi = (model.target if model else '') or ''
    _ss = _efectiu(session, 'size_system') if model else None
    if not target_codi and _ss is not None:
        _t = _ss.targets.first()
        if _t:
            target_codi = _t.codi

    # Classificació del model resolta a IDs (com 1B, codi__iexact) perquè el drawer pugui crear
    # el SizingProfile (target+construction+fit+garment_type). garment_type ja és FK al model.
    from fhort.pom.models import ConstructionType, FitType
    rs_constr = (ConstructionType.objects.filter(codi__iexact=model.construction).first()
                 if (model and model.construction) else None)
    rs_fit = (FitType.objects.filter(codi__iexact=model.fit_type).first()
              if (model and model.fit_type) else None)

    return Response({
        'target_codi': target_codi or None,
        'labels': run,
        'base_size': base_size or None,
        'poms': poms,
        'construction_id': rs_constr.id if rs_constr else None,
        'fit_type_id': rs_fit.id if rs_fit else None,
        'garment_type_id': (model.garment_type_id if model else None),
        'import_session_token': str(session.token),
        'model_id': model.id if model else None,
    }, status=200)


# ═══════════════════════════ W4 — Teixit ═══════════════════════════
_TEIXIT_FIELDS = ['fabric_main', 'fabric_composition', 'shrinkage_type', 'shrinkage_warp',
                  'shrinkage_weft', 'shrinkage_pct', 'shrinkage_iso_key', 'fabric_notes']


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def import_session_teixit_view(request, token):
    """
    PATCH /api/v1/import-sessions/<token>/teixit/  (Pas W4 — desa el teixit a la sessió)

    Desa els camps de teixit a session.resultat['teixit'] (no toca el model fins a W5).
    Opcional (es pot ometre amb skip).
    """
    from fhort.models_app.models import ImportSession

    session = ImportSession.objects.filter(token=token).first()
    if not session:
        return Response({'error': 'Sessió no trobada'}, status=404)

    teixit = {f: request.data.get(f) for f in _TEIXIT_FIELDS if f in request.data}
    session.resultat = {**(session.resultat or {}), 'teixit': teixit}
    session.save(update_fields=['resultat', 'actualitzat_at'])
    return Response({'ok': True, 'teixit': teixit}, status=200)


# ═══════════════════════════ W5 — Confirmar i guardar ═══════════════════════════
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_session_confirmar_view(request, token):
    """
    POST /api/v1/import-sessions/<token>/confirmar/  (Pas W5 — desament definitiu)

    NORMES INAMOVIBLES:
      1. Mana el document: crea NOMÉS BaseMeasurement dels POMs confirmats (Pas 2). NO
         materialitza la plantilla de l'item (no crida materialize_poms_view) i elimina les
         files buides de plantilla preexistents (base_value_cm=None).
      2. SizeFitting contenidor (sense GradingVersion/GradedSpec): el grading PROPAGAT no es
         reté; es projecta conscientment des de la regla del model (deltes+breaks), D-10.
      3. NO sessions de fitting (cap FittingSession).
      4. PDF → ModelFitxer(tipus='DOCUMENT', versio NNN, naming {codi}_DOCUMENT_{NNN});
         re-import → versio_anterior apunta a l'anterior.
      5. session.estat='CONFIRMAT'.

    ── SET-2/T8 · L'EIX DE LA PRENDA TRAVESSA TOTA L'ESCRIPTURA ────────────────────────
    **Un import = una prenda.** El `garment` de la sessió (fixat a la iniciació, mai
    preguntat) entra per PARÀMETRE a cada escriptura i a cada consulta prèvia:

      · `BaseMeasurement` neix amb `garment=<el de la sessió>` (i el signal el COPIA al
        `MeasurementChangeLog`, o sigui que el log ja diu la veritat sense tocar-lo).
      · `ModelGradingOverride` i `ModelGradingRule`, igual.
      · I LES TRES CONSULTES QUE MIREN «QUÈ HI HA JA AL MODEL» s'hi acoten: la poda de
        POMs no mencionats, la neteja de files buides i el pre-flight de les MANUAL. És
        la MATEIXA llei que #12b va aplicar a `_poda_mesures` —una llista de files no és
        una ordre d'esborrar la feina d'una altra prenda— per una TERCERA porta que la
        té pròpia: aquesta funció no passa per `_poda_mesures`, poda ella mateixa.
      · Run, talla base, sistema i joc de regles són els EFECTIUS de la peça
        (`services_garment.valor_efectiu`), i el guard dur de la talla base valida
        contra ells. Amb els de la mare, una peça amb run propi seria rebutjada —o pitjor,
        acceptada— per una escala que no és la seva.
      · La metadata reconciliada (sistema · base · run) aterra a la FILA DE LA PEÇA quan
        no som a la mare: el camí de fallback escrivia `model.save(...)` i hauria canviat
        el run de la MARE des d'un import a la 02.
    """
    import os
    from django.db import transaction
    from django.core.files.base import ContentFile

    from fhort.models_app.models import ImportSession, BaseMeasurement, ModelFitxer
    from fhort.accounts.models import UserProfile
    from fhort.pom.models import POMMaster
    from fhort.fitting.models import SizeFitting
    from fhort.models_app.matching import match_size_system

    session = ImportSession.objects.filter(token=token).select_related('model').first()
    if not session:
        return Response({'error': 'Sessió no trobada'}, status=404)
    model = session.model
    if not model:
        return Response({'error': 'La sessió no té model associat'}, status=400)

    # ── L'EIX DE LA SESSIÓ. Es llegeix UNA vegada i viatja per paràmetre; cap escriptura
    # d'aquesta funció torna a preguntar-lo ni el declara com a literal.
    from fhort.models_app.services_garment import valor_efectiu
    garment = session.garment or ''
    peca = _peca_de(session)
    destinatari = peca if peca is not None else model

    def _ef(camp):
        """El valor efectiu d'un camp heretable, llegit dels objectes EN MEMÒRIA.

        No pot ser `_efectiu(session, …)`: aquesta funció RECONCILIA metadata (sistema,
        base, run) i la desa diferida al final del pre-flight, o sigui que una relectura
        de la BD entremig tornaria el valor RANCI i el guard de la talla base validaria
        contra una escala que aquest mateix import ja ha canviat.
        """
        return valor_efectiu(model, peca, camp)

    size_system = _ef('size_system')

    user_profile = UserProfile.objects.filter(user=request.user).first()

    poms = [p for p in (session.poms_extrets or []) if p.get('actiu') and p.get('pom_master_id')]
    if not poms:
        return Response({'error': 'No hi ha POMs confirmats per importar'}, status=400)

    # mesures (Pas 3) → {(pom, capa, instancia): {talla: valor}}
    #
    # 🔑 ONADA 3 · LA CLAU ÉS LA IDENTITAT SENCERA. Era `{pom_id: {talla: valor}}`, i amb
    # tres files al mateix POM (la Brumà: 30 · 31 · 40) el diccionari en retenia UNA i les
    # tres files s'escrivien amb el mateix valor, sense error i sense avís. Els TRES
    # consumidors d'aquest mapa —el guard de la base, l'escriptura de `BaseMeasurement` i el
    # grading— el llegeixen des d'aquí, i per això es canvia la clau en origen i no a cada
    # lector. El que graduava per POM segueix graduant per POM (`_valors_per_pom`).
    valors = _valors_de_les_mesures(session)

    with transaction.atomic():
        # ── B1 · APARELLAMENT = LLEI DE LA SESSIÓ. Si el pas 1 va fixar `talla_mapping`, el confirm
        # el consumeix EN EXCLUSIVA (document→model tenant) i NO re-deriva res. Si no hi és (sessions
        # anteriors al canvi), fallback al remap canònic C1b + avís. El save de model es difereix.
        meta_update_fields = []
        _to_tenant = None
        doc_to_model = {}
        extraccio = (session.resultat or {}).get('extraccio') or {}
        run_detectat = extraccio.get('sizes') or []
        base_detectada = extraccio.get('base_size')
        talla_mapping = (session.run_conciliat or {}).get('talla_mapping')

        if talla_mapping:
            for _p in talla_mapping:
                _d, _m = (_p or {}).get('document'), (_p or {}).get('model')
                if _d and _m:
                    doc_to_model[_d] = _m
            valors = {pid: {doc_to_model.get(k, k): v for k, v in d.items()}
                      for pid, d in valors.items()}
            # El run i la base del model ja parlen tenant (el pas 1 no els toca); res a re-derivar.
        else:
            session.avisos = (session.avisos or []) + [
                "Sessió sense taula d'aparellament de talles (anterior al canvi): s'aplica el remap "
                "canònic automàtic. Reobre el pas 1 per fixar l'aparellament si cal."]
            target_codi = model.target or ''
            if not target_codi and size_system is not None:
                _ss_target = size_system.targets.first()
                if _ss_target:
                    target_codi = _ss_target.codi
            if run_detectat and base_detectada and target_codi:
                mr = match_size_system(target_codi, run_detectat, base_detectada)
                if mr.ok and mr.score == 1.0 and mr.base_ok:
                    # SET-2/T8 — la metadata reconciliada és de la PEÇA de destí. A la mare,
                    # `destinatari` ÉS el model i això és el camí de sempre, byte a byte.
                    destinatari.size_system = mr.size_system
                    size_system = mr.size_system
                    meta_update_fields = ['size_system', 'base_size_label', 'size_run_model']
                else:
                    session.avisos = (session.avisos or []) + [
                        f"Size system no reconciliat automàticament (match {mr.score:.0%} per target "
                        f"'{target_codi}'): es manté la classificació manual."]
            if size_system is not None:
                from fhort.pom.models import SizeDefinition
                _tenant_labels = list(SizeDefinition.objects.filter(size_system=size_system)
                                      .values_list('etiqueta', flat=True))
                _canon_to_tenant, _canon_ambig = {}, set()
                for _e in _tenant_labels:
                    _c = canonical_size_label(_e)
                    if _c in _canon_to_tenant and _canon_to_tenant[_c] != _e:
                        _canon_ambig.add(_c)
                    _canon_to_tenant[_c] = _e
                _no_resol = set()

                def _to_tenant(lbl):
                    _c = canonical_size_label(lbl)
                    if _c in _canon_ambig:
                        _no_resol.add(lbl)
                        return lbl
                    _t = _canon_to_tenant.get(_c)
                    if _t is None:
                        _no_resol.add(lbl)
                        return lbl
                    return _t

                valors = {pid: {_to_tenant(k): v for k, v in d.items()} for pid, d in valors.items()}
                if base_detectada:
                    base_detectada = _to_tenant(base_detectada)
                if meta_update_fields:
                    destinatari.base_size_label = (base_detectada
                                                   or _ef('base_size_label'))
                    # PORTA ÚNICA DEL RUN (llei S24b): l'ordre del DOCUMENT no mana sobre el run
                    # de la PEÇA. Les etiquetes ja vénen traduïdes a llengua-tenant per
                    # `_to_tenant`; aquí només se n'imposa l'ordre del SizeSystem.
                    from fhort.pom.grading_utils import run_del_model
                    _run_ordenat, _run_fora = run_del_model(
                        [_to_tenant(l) for l in run_detectat], size_system,
                    )
                    destinatari.size_run_model = '·'.join(_run_ordenat)
                    if _run_fora:
                        # Abans, una etiqueta sense equivalència al sistema es desava CRUA dins el
                        # run (només avisava `_no_resol`), i el model acabava amb una talla que el
                        # seu propi sistema no coneix. Ara no hi entra i es diu en clar — mateixa
                        # direcció que el check (d) de la S24.
                        session.avisos = (session.avisos or []) + [
                            f"Talles del document fora del sistema '{size_system.codi}' "
                            f"i excloses del run del model: {', '.join(_run_fora)}."]
                if _no_resol:
                    session.avisos = (session.avisos or []) + [
                        "Etiquetes del document sense equivalència única al sistema "
                        f"'{size_system.codi}': {', '.join(sorted(_no_resol))} (no traduïdes)."]

        # base_size = etiqueta tenant de la PEÇA DE DESTÍ (mai del document, mai de la mare
        # si la peça en té de pròpia).
        base_size = (_ef('base_size_label') or '').strip()
        if _to_tenant is not None and base_size:
            base_size = _to_tenant(base_size)

        # ── 1C-2a. Si la fitxa portava INCREMENTS (deltes) en comptes de mesures absolutes,
        # convertir-los a absoluts AQUÍ — abans dels TRES consumidors de `valors`
        # (BaseMeasurement a 1693, chain de GradedSpec a 1733, derive_grading_rule_set) —
        # perquè tots tres rebin absoluts i detect_grading/derive quedin INTACTES. Default
        # 'absoluts' = camí d'avui sense canvi (la conversió només s'activa si s'ha declarat).
        valors_mode = (session.resultat or {}).get('valors_mode') or 'absoluts'
        if valors_mode == 'deltes':
            from fhort.pom.grading_utils import deltes_a_absoluts
            run_ordenat_conv = [
                s.strip() for s in (_ef('size_run_model') or '').replace(';', '·').split('·')
                if s.strip()
            ]
            valors = deltes_a_absoluts(valors, base_size, run_ordenat_conv)

        # ── C1c (D2, guard DUR). La talla base de la PEÇA DE DESTÍ ha de tenir valor a la fitxa
        # (després del remap). Si no hi és entre les etiquetes dels valors → 422 ABANS de cap
        # escriptura: mai més base_value_cm=None silenciós. set_rollback per desfer el save de
        # metadata diferit (si n'hi hagués). Els valors ja parlen la llengua-tenant per C1b.
        #
        # SET-2/T8 — «la talla base» d'aquest guard és la EFECTIVA de la peça (`_ef`), no la del
        # model: importar a una prenda amb base pròpia i validar contra la de la mare rebutjaria
        # fitxes bones i n'acceptaria de dolentes, que és el mateix error amb els dos signes.
        _val_labels = set()
        for _d in valors.values():
            _val_labels |= {k for k, v in _d.items() if v is not None}
        if base_size and base_size not in _val_labels:
            transaction.set_rollback(True)
            return Response({
                'error': ("La talla base «%s» no té valor a la fitxa (etiquetes disponibles: %s)."
                          % (base_size, ', '.join(sorted(_val_labels)) or '—')),
                'tipus': 'base_size_absent',
                'base_size': base_size,
                'etiquetes': sorted(_val_labels),
            }, status=422)

        # ── POMs confirmats resolts (pur, sense escriure): necessari per a la detecció de grading.
        resolved = []
        confirmed_pom_ids = []
        for i, p in enumerate(poms):
            pm = POMMaster.objects.filter(id=int(p['pom_master_id'])).first()
            if not pm:
                continue
            resolved.append((i, p, pm))
            confirmed_pom_ids.append(int(p['pom_master_id']))

        # ══ PRE-FLIGHT SOROLL (B1, LLEI DEL SOROLL 2026-07-22) ═══════════════════════════
        # «El model s'alimenta de realitat: tot element sense contingut real és soroll i ES
        # PROPOSA eliminar, amb confirmació.»
        #
        # La norma 1 («mana el document») ja existia, però NOMÉS de cara al que el document
        # SÍ porta. La contrapartida faltava: els POMs vius del model que el document NO
        # menciona sobrevivien actius EN SILENCI, i la fitxa importada quedava contaminada
        # amb mesures que el client no demanava (§B3.3 de la DIAGNOSI_GTI_PLANTILLA).
        #
        # Ara es PROPOSEN, mai s'esborren sols: 409 amb la llista → el tècnic tria
        # `poda_choice`. Mateix mecanisme que `container_choice` (no cal pantalla nova).
        # Sempre soft (is_active=False), mai DELETE dur.
        #
        # ⚠️ SET-2/T8 · I L'ABAST ÉS EL DE LA PRENDA QUE S'IMPORTA. Aquesta és la TERCERA
        # porta de poda de la casa —les dues de `_poda_mesures` les va acotar el #12b— i té
        # la poda pròpia, o sigui que la llei s'hi ha de tornar a escriure: un document que
        # parla de la Llaçada no diu RES de les mesures del Pantaló, i sense el filtre les
        # hauria desactivades totes per no sortir a la seva llista. Un import a la 02 amb
        # `garment=''` al filtre proposava podar la mare sencera.
        #
        # ⚠️ SET-2/T8-ter (16/08) · I ARA L'ABAST SÓN LES PECES QUE LA TRAMESA ANOMENA, EN
        # PLURAL. Amb el garment a la fila un sol import en toca N, i el filtre escalar tornava
        # a ser el dany de sempre girat: podar la mare mirava només la mare —correcte— però un
        # import que porta faldilla i short deixava el short sencer FORA de la poda, o sigui que
        # les seves mesures ràncies sobrevivien en silenci a un document que ja no les diu.
        #
        # I la pregunta es respon PER PEÇA: els POMs confirmats del short no protegeixen les
        # files de la faldilla. Amb un `confirmed_pom_ids` pla, importar el short hauria
        # «confirmat» el POM 962 i la fila G1 de la faldilla, que el document ja no menciona,
        # no hauria sortit mai a la llista de poda. És la MATEIXA llei que `_poda_mesures` va
        # aprendre al #12b —qui no surt a la llista cau només si la seva peça és una de les que
        # la llista anomena—, escrita aquí per tercera vegada perquè aquesta porta és pròpia.
        poda_choice = (request.data.get('poda_choice') or '').strip().lower()  # 'desactivar'|'conservar'
        confirmats_per_peca = {}
        for _i, _p, _pm in resolved:
            confirmats_per_peca.setdefault(_garment_de(_p, garment), set()).add(_pm.id)
        orfes = []
        for _g, _pids in confirmats_per_peca.items():
            orfes.extend(
                BaseMeasurement.objects
                .filter(model=model, garment=_g, is_active=True, base_value_cm__isnull=False)
                .exclude(pom_id__in=_pids)
                .select_related('pom')
            )
        if orfes and poda_choice not in ('desactivar', 'conservar'):
            return Response({
                'conflict': True,
                'tipus': 'poms_no_mencionats',
                'poms': [{
                    'pom_id': bm.pom_id,
                    'codi': bm.pom.codi_client or '',
                    'nom': bm.nom_fitxa or getattr(bm.pom, 'nom_ca', '') or '',
                    'base_value_cm': bm.base_value_cm,
                    'origen': bm.origen,
                    # SET-2/T8-ter — DE QUINA PEÇA és cada òrfena. Amb un import que en toca
                    # més d'una, una llista que no ho digués faria decidir a cegues: dues files
                    # del mateix POM, una de la faldilla i una del short, es llegirien com una
                    # repetició i el tècnic no sabria quina està a punt de donar de baixa.
                    'garment': bm.garment,
                    'garment_nom': _nom_de_peca(model, bm.garment),
                } for bm in orfes],
                'n': len(orfes),
                'garment': garment,
                'garment_nom': _nom_de_la_peca(session),
                'garments': sorted(confirmats_per_peca),
                'message': ("Aquesta prenda té mesures vives que el document no menciona. "
                            "Vols desactivar-les (el model s'alimenta de la realitat del "
                            "document) o conservar-les?"),
            }, status=409)

        # ══ PRE-FLIGHT PRECEDÈNCIA MANUAL (B2) ══════════════════════════════════════════
        # Precedència d'orígens MÍNIMA — només la que aquesta llei necessita, no el mapa
        # complet (§B3.3: `ORIGEN_CHOICES` segueix sent una llista plana i això NO ho canvia).
        #
        # Una fila origen='MANUAL' amb valor és patrimoni escrit a mà per un tècnic. Fins ara
        # l'`update_or_create` de sota la trepitjava sense mirar-se l'origen previ — l'única
        # comparació d'origen de tot el repo vivia a `models_app/views.py:827`. El patró de
        # guard que se segueix és `pom/dictionary_service.py:158` (`preserve_manual`).
        #
        # No es decideix per ell en cap direcció: es PROPOSA, com el soroll.
        manual_choice = (request.data.get('manual_choice') or '').strip().lower()  # 'sobreescriure'|'respectar'
        # ONADA 3 — la pregunta segueix sent PER POM (la consulta de sota filtra per `pom_id`),
        # però «porta valor?» es respon per la IDENTITAT de la fila: amb el POM pelat, una
        # germana sense valor de base n'hauria arrossegat una altra que sí que en té.
        _doc_pom_ids = {int(p['pom_master_id']) for _i, p, _pm in resolved
                        if (valors.get(_identitat(p, garment)) or {}).get(base_size) is not None}
        # SET-2/T8 — acotat a la prenda, com el soroll: el que un tècnic va teclejar al
        # Pantaló no el trepitja un document de la Llaçada, i per tant tampoc n'ha de sortir
        # una pregunta que faria decidir sobre files que aquest import no tocarà mai.
        manuals = list(
            BaseMeasurement.objects
            # SET-2/T8-ter — les peces que la tramesa anomena, no la de la sessió: un import de
            # dues peces ha de preguntar pel patrimoni MANUAL de totes dues.
            .filter(model=model, garment__in=sorted(confirmats_per_peca) or [garment],
                    is_active=True, origen='MANUAL',
                    base_value_cm__isnull=False, pom_id__in=_doc_pom_ids)
            .select_related('pom')
        ) if _doc_pom_ids else []
        if manuals and manual_choice not in ('sobreescriure', 'respectar'):
            return Response({
                'conflict': True,
                'tipus': 'manual_trepitjat',
                'poms': [{
                    'pom_id': bm.pom_id,
                    'codi': bm.pom.codi_client or '',
                    'nom': bm.nom_fitxa or getattr(bm.pom, 'nom_ca', '') or '',
                    'valor_manual': bm.base_value_cm,
                    # La fila MANUAL que es trepitjaria té els seus propis eixos: el valor que
                    # se li compara ha de sortir de la SEVA identitat, no d'una germana.
                    'valor_document': (valors.get((bm.pom_id, bm.capa, bm.instancia,
                                                   bm.garment)) or {}).get(base_size),
                    'garment': bm.garment,
                    'garment_nom': _nom_de_peca(model, bm.garment),
                } for bm in manuals],
                'n': len(manuals),
                'message': ("El document porta valor per a mesures introduïdes MANUALMENT. "
                            "Vols que mani el document o que es respecti el valor manual?"),
            }, status=409)
        # Les FILES on el valor manual guanya: l'escriptura de sota les salta. ONADA 3 — per
        # IDENTITAT, no per POM: amb la clau curta, respectar una MANUAL de l'exterior hauria
        # deixat sense escriure la germana del folre, que ningú no havia decidit respectar.
        respectats_idents = ({(bm.pom_id, bm.capa, bm.instancia, bm.garment) for bm in manuals}
                             if manual_choice == 'respectar' else set())

        # ══ PRE-FLIGHT GRADING (D1) — detecció (pura) + matcher + GATES 409, TOT abans d'escriure.
        # El contenidor de client (GradingRuleSet origen=CLIENT_RUN) és ÚNIC per (customer +
        # size_system + garment_type_item + fit). Les decisions que exigeixen tria del tècnic surten
        # amb 409 SENSE haver tocat cap fila → les mesures (sobirania del model) ja no es fan
        # rollback per una decisió de grading. derive_rules_from_fitxa és pur (no persisteix).
        from fhort.pom.grading_utils import (
            derive_rules_from_fitxa, resolve_grading_container, classifica_fitxa_vs_contenidor)
        from fhort.pom.models import FitType, Target, ConstructionType, GradingRuleSet
        from fhort.models_app.services import (
            materialize_model_grading_rules_from_specs, afegeix_regles_al_contenidor,
            proposta_promocio, resum_proposta_promocio)
        from fhort.models_app.models import Watchpoint
        from fhort.pom.services import maybe_learn_customer_alias

        watchpoint_promocio = None
        grading_avisos = []
        grading_bloqueigs = []
        grading_descartades = []   # LLEI BEACH: columnes del document fora del sistema (no-base)
        # REFERENT (llei S24): el run del DOCUMENT en llengua-tenant. Es passa pel MATEIX
        # aparellament que ja s'ha aplicat a `valors` més amunt (taula de la sessió si n'hi ha,
        # remap canònic si no) perquè referent i valors parlin la mateixa llengua. Abans aquí
        # s'hi passava `model.size_run_model`: amb un run de model més ESTRET que el document,
        # els deltes es calculaven entre veïns falsos i el break sortia fabricat (bug 166).
        run_document = list(run_detectat or [])
        if talla_mapping:
            run_document = [doc_to_model.get(l, l) for l in run_document]
        elif _to_tenant is not None:
            run_document = [_to_tenant(l) for l in run_document]
        # (a) DETECCIÓ de les regles de la fitxa (pur, sense persistència; reusa detect_grading).
        #
        # ONADA 3 — el motor de detecció gradua PER POM (una regla per `pom_id`, i el
        # contenidor del client tampoc no coneix instàncies), o sigui que aquí la taula entra
        # per la vista col·lapsada. No és una pèrdua: el valor de cada germana ja ha arribat
        # sencer al seu `BaseMeasurement`; el que no es fabrica és una regla de catàleg per
        # instància, que ningú no ha decidit. `_valors_per_pom` ho DIU en un avís quan passa.
        valors_pom = _valors_per_pom(valors, avisos=grading_avisos)
        fitxa_specs = derive_rules_from_fitxa(
            run_document=run_document, base_size=base_size, valors=valors_pom,
            confirmed_pom_ids=confirmed_pom_ids, size_system=size_system,
            avisos=grading_avisos, bloqueigs=grading_bloqueigs,
            descartades=grading_descartades)
        # LLEI BEACH (2026-07-26): les columnes fora del sistema (no-base) es descarten també de
        # `valors`, perquè cap escriptura de sota en depengui — en particular el bucle de
        # ModelGradingOverride (que itera TOTES les etiquetes de `valors`) no ha de persistir un
        # override per a una talla d'un altre sistema. `run_ordenat` ja les excloïa del grading;
        # aquí es tanca el forat de l'escriptura d'overrides.
        if grading_descartades:
            _desc_canon = {canonical_size_label(x) for x in grading_descartades}
            valors = {ident: {k: v for k, v in d.items()
                              if canonical_size_label(k) not in _desc_canon}
                      for ident, d in valors.items()}
            # I la vista per POM es refà de la taula ja podada: és la que el bucle
            # d'overrides itera, i llegir-la de la d'abans hi tornaria a colar la columna.
            valors_pom = _valors_per_pom(valors)
        # BLOQUEIG d'integritat (llei 2026-07-08): cap regla d'una taula incompleta. Abans
        # aquest camí només avisava i persistia igualment — el forat del bug 166. 422 ABANS de
        # cap escriptura de grading; `set_rollback` perquè som dins de l'atomic i les mesures
        # ja escrites més amunt no poden quedar confirmades amb un error a la mà.
        if grading_bloqueigs:
            incompletes = [b for b in grading_bloqueigs if b['tipus'] == 'fila_incompleta']
            desconegudes = [e for b in grading_bloqueigs
                            if b['tipus'] == 'talles_desconegudes' for e in b['etiquetes']]
            if desconegudes:
                msg = ("El document porta talles que el sistema de talles del model no coneix: "
                       f"{', '.join(desconegudes)}.")
            else:
                msg = (f"{len(incompletes)} mesura/es no tenen valor per a totes les talles del "
                       "document; no se'n pot derivar cap regla sense inventar-ne el trencament. "
                       "Completa-les al pas de mesures o desmarca-les.")
            transaction.set_rollback(True)
            return Response({
                'error': msg + ' (cap regla desada)',
                'tipus': 'grading_taula_incompleta',
                'bloqueigs': grading_bloqueigs,
                'run_document': run_document,
                'avisos': grading_avisos,
            }, status=422)
        base_def_id = fitxa_specs[0]['talla_base_id'] if fitxa_specs else None

        # (b) MATCHER ÚNIC (M1): resol fit (codi→FK) i EL contenidor per la llei del contenidor
        # (N1 identitat exacta · N2 ampli item-NULL del mateix client · N3 cap).
        rs_fit = FitType.objects.filter(codi__iexact=model.fit_type).first() if model.fit_type else None
        gti = model.garment_type_item
        grp_codi = model.garment_group.codi if model.garment_group_id else None
        # El contenidor és del CLIENT (customer + sistema + item + fit) i per tant no porta eix
        # de peça: el que sí que hi entra per la peça és el SISTEMA DE TALLES efectiu, perquè
        # una prenda amb sistema propi ha de buscar el contenidor del SEU sistema.
        res_cont = resolve_grading_container(
            model.customer, size_system, model.target, model.construction,
            rs_fit, grp_codi, garment_type_item=gti)
        container = res_cont['container']
        container_choice = (request.data.get('container_choice') or '').strip().lower()  # 'create'|'no_container'

        cls = None
        if fitxa_specs:
            # (c) DECISIONS que exigeixen tria conscient → 409 SENSE cap escriptura (cap set_rollback:
            # res s'ha tocat encara; la metadata reconciliada tampoc, es desa al bloc d'escriptura).
            if res_cont['motiu'] == 'ambiguous':
                return Response({
                    'conflict': True,
                    'tipus': 'container_ambigu',
                    'candidats': [{'id': c.id, 'nom': c.nom} for c in res_cont['candidats']],
                    'message': ("Hi ha més d'un contenidor de graduació possible per a aquesta "
                                "combinació. Cal triar-ne un abans de continuar."),
                }, status=409)
            if container is None and container_choice not in ('create', 'no_container'):
                return Response({
                    'conflict': True,
                    'tipus': 'container_absent',
                    'customer_nom': str(getattr(model.customer, 'nom', '') or model.customer or ''),
                    'garment_type_item': (getattr(gti, 'name', '') if gti else ''),
                    'size_system': str(getattr(size_system, 'nom', '') or size_system or ''),
                    'fit': (rs_fit.codi if rs_fit else (model.fit_type or '')),
                    'n_regles': len(fitxa_specs),
                    'message': ("Aquest client no té graduació per a aquesta combinació "
                                "(peça + sistema de talles + fit). Vols crear-ne el contenidor?"),
                }, status=409)
            if container is not None:
                # M3 — llei del contenidor INTOCABLE (classificació pura; s'aplica a l'escriptura).
                cls = classifica_fitxa_vs_contenidor(fitxa_specs, container)

        # ════════════════════════════════ ESCRIPTURA ════════════════════════════════
        # Totes les decisions que podien retornar 409/422 ja s'han pres. Persistim la metadata
        # reconciliada de la PEÇA DE DESTÍ (diferida del pre-flight) i escrivim les mesures
        # (sobirania). A la mare, `destinatari` ÉS el model: `model.save(update_fields=…)` de
        # sempre. A una prenda, els tres camps són overrides seus i la mare no es toca.
        if meta_update_fields:
            destinatari.save(update_fields=meta_update_fields)

        # ── 1. Mana el document: neteja files buides i crea NOMÉS els confirmats.
        #
        # B1 (LLEI DEL SOROLL) — CRITERI TRIAT per a les files sense valor:
        #   · origen TEMPLATE/ITEM_STANDARD → **DELETE dur**. Són bastida de plantilla que
        #     mai va ser realitat: ningú les va mesurar, no hi ha res a auditar i deixar-les
        #     com a inactives només acumularia runa que un segon import tornaria a trobar.
        #   · qualsevol altre origen (MANUAL, IMPORTED, FITTED…) → **SOFT** (is_active=False)
        #     + entrada al MeasurementChangeLog. Algú les va crear conscientment encara que
        #     ara no portin valor; la seva desaparició ha de deixar rastre.
        #
        # SET-2/T8 — acotat a la prenda (la mateixa llei de la poda de dalt: aquesta neteja
        # ÉS una poda, i esborrar files de plantilla d'una altra peça és esborrar la seva
        # bastida sense que ningú l'hagi anomenada).
        _TEMPLATE_ORIGENS = ('TEMPLATE', 'ITEM_STANDARD')
        # SET-2/T8-ter — i aquesta neteja també és per PECES ANOMENADES: la bastida de plantilla
        # d'una peça que aquest import no toca no és soroll seu i no li pertoca esborrar-la.
        _buides = BaseMeasurement.objects.filter(
            model=model, garment__in=sorted(confirmats_per_peca) or [garment],
            base_value_cm__isnull=True)
        _buides.filter(origen__in=_TEMPLATE_ORIGENS).delete()
        n_buides_soft = 0
        for bm in _buides.exclude(origen__in=_TEMPLATE_ORIGENS).exclude(is_active=False):
            bm.is_active = False
            bm._desactivat = True
            bm._changed_by = request.user
            bm._motiu = 'import: fila sense valor (soroll)'
            bm.save(update_fields=['is_active'])
            n_buides_soft += 1

        n_bm = 0
        n_bm_valors = 0
        n_manual_respectats = 0
        for i, p, pm in resolved:
            # ONADA 3 — el valor és el de LA FILA, no el del POM: tres files del mateix POM
            # són tres mesures, i amb la clau curta les tres es desaven amb el mateix número.
            base_val = (valors.get(_identitat(p, garment)) or {}).get(base_size)
            # B2 — el tècnic ha decidit que el valor manual mana: el document no el trepitja.
            # La fila queda tal com està (valor, origen MANUAL i tot); només és patrimoni que
            # sobreviu a l'import, no una fila nova.
            if _identitat(p, garment) in respectats_idents:
                n_manual_respectats += 1
                n_bm += 1
                n_bm_valors += 1
                # El vincle codi↔POM SÍ s'aprèn: la tria és sobre el VALOR, no sobre el
                # vocabulari. El document ha anomenat aquest POM i això és realitat.
                maybe_learn_customer_alias(
                    model.customer, p.get('codi_fitxa'), p.get('descripcio'), pm,
                    origen='IMPORT', nomes_si_manual=False)
                continue
            _defaults = {
                'base_value_cm': base_val,
                'nom_fitxa': p.get('codi_fitxa') or '',
                'origen': 'IMPORTED',
                'is_active': True,
                'ordre': i,
                'notes': p.get('descripcio') or '',
                # F3 — la secció d'origen. Els dos camins d'extracció ja la portaven a
                # `poms_extrets` i aquí es perdia perquè no hi havia columna on desar-la.
                'seccio': p.get('seccio') or '',
            }
            # B2: només escrivim tolerància si el document en porta (asimètrica, contracte Size Check).
            if p.get('tol_minus') is not None:
                _defaults['tolerancia_minus'] = p['tol_minus']
            if p.get('tol_plus') is not None:
                _defaults['tolerancia_plus'] = p['tol_plus']
            # ✅ ONADA 3 · TANCADA EL 14/08. Aquí hi havia `capa=SLUG_DEFECTE, instancia=''`
            # literals, amb l'acta que deia: «el document importat encara no diu de quina capa
            # ni de quina instància parla cada fila; el lèxic multilingüe que ho sabrà llegir
            # és l'Onada 3, que té UI i maqueta pendent». Ja no és pendent, i el que ha
            # resolt el forat NO és el lèxic: és que **ho decideix la persona**. El pas de
            # POMs porta el columnat d'instància de la definició manual, la decisió es desa a
            # la fila i aquí es llegeix. El lèxic automàtic (LINING/FORRO→folre, «stretched
            # out»→extended) queda per a quan hi hagi corpus d'imports reals que l'ensenyi;
            # decisió d'Agus, i és la llei de l'import: el que no se sap segur, ho decideix
            # l'humà. Sense declarar res, `_capa_instancia_de` torna exactament els dos
            # literals d'abans: cap sessió existent canvia de comportament.
            _capa, _instancia = _capa_instancia_de(p)
            # …i la comporta `instancia_exigeix_nom` (migració 0074) segueix manant: una fila
            # amb instància SENSE `nom_fitxa` és un IntegrityError, no un 422 educat. Passa
            # amb els POMs afegits a mà al pas 2, que neixen amb `codi_fitxa` buit: el codi
            # del catàleg és el nom que la fitxa hauria portat si el document l'hagués dit.
            if _instancia and not _defaults['nom_fitxa']:
                _defaults['nom_fitxa'] = (pm.codi_client or '')[:20]
            # ── DECISIÓ 7 · F3 — L'IMPORT DEIXA DE REBATEJAR ────────────────────────────
            #
            # 🚨 EL DEFECTE, mesurat el 28/08: `nom_fitxa` anava als `defaults` d'aquest
            # `update_or_create` i la clau NO el conté, o sigui que un RE-IMPORT del mateix
            # document substituïa la nomenclatura que el tècnic hagués posat, en silenci i
            # sense entrada al `MeasurementChangeLog` (que només registra `base_value_cm`).
            #
            # LA LLEI (Agus, 28/08): sobre fila VERGE mana el DOCUMENT; sobre fila EXISTENT
            # mana la Montse. Els dos casos ja els distingeix la clau, i per això això no
            # necessita cap flag: `create_defaults` s'aplica només quan la fila neix, i
            # `defaults` només quan ja hi era. El valor, l'origen i la resta segueixen
            # actualitzant-se igual — l'import segueix portant les MESURES; el que deixa de
            # portar és el BATEIG.
            #
            # ⚠️ `nom_fitxa` ha de quedar als DOS diccionaris i no només a `create_defaults`:
            # Django aplica `defaults` a la creació quan no hi ha `create_defaults`, però si
            # se'n dona un, la creació passa a fer servir NOMÉS aquest. Deixar-lo fora de
            # `create_defaults` faria que una fila nova amb instància nasqués sense
            # nomenclatura → IntegrityError contra `instancia_exigeix_nom`.
            _create_defaults = dict(_defaults)
            _defaults.pop('nom_fitxa', None)
            BaseMeasurement.objects.update_or_create(
                model=model, pom=pm,
                # SET-2/T5 — el garment es declara igual que els altres dos eixos i entra a la
                # CLAU, no als defaults: és identitat de fila, i sense ell l'update_or_create
                # petaria amb MultipleObjectsReturned amb dues peces vives.
                # SET-2/T8 — i el valor ja no és el literal `''`: és el de la SESSIÓ. Aquesta
                # és la línia on l'import deixa d'escriure sempre a la mare.
                # SET-2/T8-ter — i ara és el de LA FILA, amb el de la sessió de recanvi. Aquesta
                # és la línia on un sol import deixa d'escriure sempre a la MATEIXA peça: la
                # faldilla i el short de la Brumà surten del mateix document i aterren a dos
                # contenidors, cadascuna amb la seva.
                capa=_capa, instancia=_instancia, garment=_garment_de(p, garment),
                defaults=_defaults, create_defaults=_create_defaults)
            n_bm += 1
            if base_val is not None:
                n_bm_valors += 1
            # Biblioteca del client (QA-S8-R1): aprèn de tot vincle ferm confirmat (idempotent). El
            # guard de pom/services.py aplica: si un ALTRE codi ja reclama el POM, l'àlies neix
            # pendent_revisio=True i find_pom_master no l'auto-vincula.
            maybe_learn_customer_alias(
                model.customer, p.get('codi_fitxa'), p.get('descripcio'), pm,
                origen='IMPORT', nomes_si_manual=False)

        # ── 1b. PODA CONFIRMADA (B1). Els POMs vius que el document no menciona: el tècnic
        # ja ha triat al pre-flight. SOFT sempre (is_active=False) + MeasurementChangeLog;
        # cap DELETE dur — la mesura va existir i el model n'ha de guardar memòria.
        n_podats = 0
        if orfes and poda_choice == 'desactivar':
            for bm in orfes:
                bm.is_active = False
                bm._desactivat = True
                bm._changed_by = request.user
                bm._motiu = 'import: POM no mencionat pel document (poda confirmada)'
                bm.save(update_fields=['is_active'])
                n_podats += 1
            grading_avisos.append(
                f"Poda confirmada: {n_podats} POM(s) que el document no menciona s'han "
                f"desactivat (soft, amb registre al log de mesures).")
        elif orfes and poda_choice == 'conservar':
            grading_avisos.append(
                f"{len(orfes)} POM(s) vius que el document NO menciona s'han CONSERVAT per "
                f"decisió del tècnic: la fitxa del model els segueix incloent.")
        if n_manual_respectats:
            grading_avisos.append(
                f"{n_manual_respectats} mesura/es d'origen MANUAL s'han RESPECTAT per decisió "
                f"del tècnic: el valor del document no les ha trepitjat.")
        elif manuals and manual_choice == 'sobreescriure':
            grading_avisos.append(
                f"{len(manuals)} mesura/es d'origen MANUAL s'han sobreescrit amb el valor del "
                f"document per decisió del tècnic.")
        if n_buides_soft:
            grading_avisos.append(
                f"{n_buides_soft} fila/es sense valor i d'origen no-plantilla s'han "
                f"desactivat (soft) en lloc d'esborrar-se.")

        # ── 2. Identificador del contenidor SF.
        next_num = 1
        while SizeFitting.objects.filter(model=model, numero=next_num).exists():
            next_num += 1
        sf_codi = f"IMP-{model.id}-{next_num}"
        while SizeFitting.objects.filter(codi=sf_codi).exists():
            next_num += 1
            sf_codi = f"IMP-{model.id}-{next_num}"

        # (d) APLICAR (escriptures) — savepoint intern amb degradació amb gràcia.
        #
        # SET-2/T8 — el JOC és un camp heretable: el que aquest import decideixi és de la PEÇA
        # DE DESTÍ i s'escriu a `destinatari`. A la mare, `destinatari` ÉS el model i el camí
        # és el de sempre; a una prenda, l'override seu, i el joc de la mare no es mou.
        new_rule_set = _ef('grading_rule_set')
        resident_specs = None
        prev_grs_id = getattr(destinatari, 'grading_rule_set_id', None)
        # 6.1 — el joc que la PEÇA tenia ABANS d'aquest import, com a OBJECTE: cal llegir-ne
        # l'`origen` per decidir si les seves `MANUAL` són autoria. Capturat aquí, abans de
        # qualsevol `save()` d'aquesta funció.
        prev_grs_obj = new_rule_set
        if fitxa_specs:
            try:
                with transaction.atomic():
                    if container is None:
                        if container_choice == 'no_container':
                            # SOBIRANIA: la peça queda amb regles residents pròpies, sense contenidor.
                            destinatari.grading_rule_set = None
                            destinatari.save(update_fields=['grading_rule_set'])
                            new_rule_set = None
                            resident_specs = fitxa_specs
                            grading_avisos.append(
                                "Contenidor no creat (decisió del tècnic): el model queda amb "
                                "regles residents pròpies (sobirania de dades).")
                        else:  # 'create' — M3: CREAR contenidor AMPLI (item=NULL) per defecte.
                            rs_target = (Target.objects.filter(codi__iexact=model.target).first()
                                         if model.target else None)
                            rs_constr = (ConstructionType.objects.filter(codi__iexact=model.construction).first()
                                         if model.construction else None)
                            nom_cont = " · ".join(p for p in [
                                str(getattr(model.customer, 'nom', '') or model.customer or ''),
                                str(getattr(model.garment_group, 'nom', '') or model.garment_group or ''),
                                str(getattr(size_system, 'nom', '') or size_system or ''),
                            ] if p)[:120] or f"Contenidor client · {model.codi_intern}"
                            # AMPLI: garment_type_item=NULL (abast per garment_group FK → el troba M1
                            # nivell 2 la propera vegada). NO és la identitat fina (item), és de món.
                            container = GradingRuleSet.objects.create(
                                nom=nom_cont, size_system=size_system,
                                garment_group=model.garment_group, garment_type_item=None,
                                construction=rs_constr, fit_type=rs_fit,
                                is_system_default=False, actiu=True,
                                origen=GradingRuleSet.ORIGEN_CLIENT_RUN, customer=model.customer)
                            if rs_target:
                                container.targets.add(rs_target)
                            afegeix_regles_al_contenidor(container, fitxa_specs, base_def_id)
                            destinatari.grading_rule_set = container
                            destinatari.save(update_fields=['grading_rule_set'])
                            new_rule_set = container
                            resident_specs = fitxa_specs
                            grading_avisos.append(
                                f"Contenidor de client AMPLI NOU creat #{container.id} '{container.nom}' "
                                f"(el client estrenava aquesta combinació) amb {len(fitxa_specs)} regla(es).")
                    elif not container.regles.exists():
                        # CONTENIDOR ESQUELET (0 regles) → sembrar-lo des de la fitxa és LEGÍTIM (M3).
                        # Amb 0 regles, cls['amplia'] == totes les specs (res per coincidir/divergir).
                        if cls['amplia']:
                            afegeix_regles_al_contenidor(container, cls['amplia'], base_def_id)
                        destinatari.grading_rule_set = container
                        destinatari.save(update_fields=['grading_rule_set'])
                        new_rule_set = container
                        resident_specs = fitxa_specs
                        grading_avisos.append(
                            f"Contenidor esquelet #{container.id} '{container.nom}' sembrat des de la "
                            f"fitxa ({len(fitxa_specs)} regla(es)).")
                    else:
                        # CONTENIDOR AMB REGLES → INTOCABLE (llei M3): el catàleg del client NO es toca.
                        #   coincideix (sembra) → res: el model hereta la regla del contenidor.
                        #   divergeix (conflicte) + POM nou (amplia) → ModelGradingOverride per-talla
                        #     (valors de la fitxa a les talles no-base) + WATCHPOINT. El motor llegeix
                        #     l'override amb prioritat sobre la projecció del contenidor
                        #     (services._load_model_overrides); base i talla-base van a BaseMeasurement.
                        from fhort.pom.grading_utils import _norm as _norm_label
                        from fhort.models_app.models import ModelGradingOverride
                        destinatari.grading_rule_set = container
                        destinatari.save(update_fields=['grading_rule_set'])
                        new_rule_set = container
                        # SENSE residents: el contenidor mana (all-or-nothing de _load_grading_rules).
                        # Neteja residents ranços perquè l'herència del contenidor no quedi tapada.
                        #
                        # 🔴 6.1 · AQUESTA BRANCA NO PRESERVA LES `MANUAL`, I ÉS A POSTA. Salvar-ne
                        # una sola deixaria el model amb residents, i `_load_grading_rules` és
                        # ALL-OR-NOTHING: amb una resident viva el contenidor deixaria de graduar
                        # els altres 24 POMs. Preservar aquí no salvaria una regla, desactivaria
                        # el contenidor sencer en silenci. És l'únic camí de wipe que la decisió
                        # 6.1 NO tanca, i el tancarà la política fina (decisió de l'Agus).
                        #
                        # ⚠️ SET-2/T8 — I EL WIPE S'ACOTA A LA PRENDA. `model.grading_rules.all()`
                        # era el mateix dany del #12d amb una altra taula: un import a la Llaçada
                        # esborrava les regles residents del Pantaló, que no surten enlloc del
                        # document i que ningú no ha anomenat. La germana d'aquesta línia
                        # (`materialize_model_grading_rules_from_specs`) ja fa el wipe per
                        # `(model, garment)` des de T3; aquesta se li iguala.
                        model.grading_rules.filter(garment=garment).delete()
                        resident_specs = None
                        base_norm = _norm_label(base_size)
                        pom_divergents = ([c['pom_id'] for c in cls['conflicte']]
                                          + [s['pom_id'] for s in cls['amplia']])
                        n_ovr = 0
                        for pom_id in pom_divergents:
                            # Per POM: l'override és la contrapartida d'una regla de catàleg,
                            # i el catàleg no coneix instàncies (mateixa llei que la detecció).
                            for label, val in (valors_pom.get(pom_id) or {}).items():
                                if val is None or _norm_label(label) == base_norm:
                                    continue
                                # SET-2/T8 — `garment` a la CLAU (és identitat de fila, com a
                                # BaseMeasurement): l'override d'una prenda no trepitja el de
                                # la germana ni el de la mare.
                                ModelGradingOverride.objects.update_or_create(
                                    model=model, pom_id=pom_id, size_label=label,
                                    capa=MeasurementLayer.SLUG_DEFECTE, instancia='',
                                    garment=garment,
                                    defaults={'value_cm': float(val), 'created_by': user_profile,
                                              'motiu': ("Import W5 — divergència vs catàleg del "
                                                        "contenidor (INTOCABLE)")})
                                n_ovr += 1
                        # D1 — PROPOSTA DE PROMOCIÓ. Aquí hi havia dos avisos de text lliure que
                        # ningú no llegia mai: anaven a `grading_avisos`, que el front descarta, i a
                        # `session.avisos`, que cap serialitzador exposa. Els POMs que no entraven al
                        # catàleg quedaven registrats enlloc visible — la pèrdua silenciosa que
                        # aquest sprint tanca. Ara en surt un Watchpoint ESTRUCTURAT i accionable
                        # (mateix mecanisme viu que l'import de config: `dades` + WatchpointsPanel).
                        proposta = proposta_promocio(cls, container, base_def_id)
                        if proposta:
                            watchpoint_promocio = Watchpoint.objects.create(
                                model=model, task=None, created_by=user_profile,
                                text=resum_proposta_promocio(proposta),
                                dades=proposta,
                            )
                        if cls['sembra']:
                            grading_avisos.append(
                                f"{len(cls['sembra'])} POM(s) coincideixen amb el catàleg: el model "
                                f"els hereta del contenidor #{container.id} (sense override).")
                    # SEMBRA SELECTIVA de residents (origen=IMPORTED); el motor les llegeix amb prioritat.
                    if resident_specs is not None:
                        materialize_model_grading_rules_from_specs(
                            model, resident_specs, origen='IMPORTED',
                            joc_anterior=prev_grs_obj, garment=garment)
            except Exception as e:
                # La reversió és del MATEIX objecte que hauria escrit: a la mare el model, a una
                # prenda la seva fila. Restaurar `model.grading_rule_set_id` des d'un import a la
                # 02 hauria mogut el joc de la mare a l'hora d'anar malament.
                destinatari.grading_rule_set_id = prev_grs_id
                new_rule_set = _ef('grading_rule_set')
                grading_avisos.append(
                    f"Grading no aplicat (error en desar: {e}); es manté el ruleset previ de la peça.")
        n_rules = model.grading_rules.filter(garment=garment).count()

        # ── SET-2/T8-ter (16/08) · AQUÍ HI HAVIA UN AVÍS + WATCHPOINT `residents_de_peca` I
        # S'HA RETIRAT, perquè el forat que denunciava ja no existeix.
        #
        # Deia: «mentre el motor decideixi "hi ha residents?" per MODEL i no per peça, les altres
        # prendes poden deixar de graduar pel contenidor del client», i obria un Watchpoint a
        # cada import a una peça. **PRED-3 el va tancar el 12/08**: `_load_grading_rules_per_garment`
        # (`pom/services.py:876`) ja no pregunta `rules.exists()` sobre tot el model sinó
        # `te_residents_la_mare` —«en té LA MARE?»—, que és el subjecte bo, i amb ell una peça
        # amb residents propis ja NO fa baixar el catàleg per a les seves germanes.
        #
        # Un avís que descriu un estat que ja no es dona no és prudència: és soroll que fa
        # desconfiar del que sí que avisa. I amb T8-ter costava més que ahir —un sol import pot
        # tocar N peces, o sigui que n'hauria obert un per import i per peça.
        # Verificat contra el codi viu el 16/08 abans de retirar-lo.

        # ── 3. SizeFitting CONTENIDOR per a la projecció CONSCIENT (D-10) — només quan no hi ha
        # conflicte pendent. L'import reté base + deltes + breaks (ModelGradingRule); el grading
        # PROPAGAT no es reté: el projecta el motor després, des de la regla vigent del model
        # (generate_grading_view crea/omple la versió sobre AQUEST SF). Estat/segellat (D-1) NO aquí.
        size_fitting = SizeFitting.objects.create(
            model=model, numero=next_num, codi=sf_codi, tipus='SizeSet',
            estat='Tancat', base_tancada=True, creat_per=user_profile,
            notes="Importació guiada (wizard). Contenidor; grading propagat NO retingut "
                  "(es projecta conscientment des de la regla del model, D-10).",
        )
        n_specs = 0   # cap valor propagat persistit a l'import

        # ── C3 (D3, defensa en profunditat). Amb el guard C1c això no hauria de passar; si tot i
        # així cap POM ha rebut valor de talla base, avís destacat (mai un "OK" enganyós amb 0 valors).
        if n_bm and not n_bm_valors:
            grading_avisos.append(
                f"⚠️ S'han desat {n_bm} POM(s) SENSE cap valor de talla base "
                f"(base '{base_size}'): revisa l'alineació d'etiquetes de la fitxa.")

        # ── BEACH (2026-07-26). Columnes del document fora del sistema de talles (no-base): NO és
        # error — la fitxa portava una talla d'un altre sistema (p.ex. una BABY en un model
        # NEWBORN). S'han importat les conegudes; aquí en queda constància DURABLE (Watchpoint,
        # visible al WatchpointsPanel del model) + avís al resum del pas 5.
        columnes_descartades = list(dict.fromkeys(grading_descartades))
        if columnes_descartades:
            _ss_codi = size_system.codi if size_system is not None else ''
            _etq = ', '.join(columnes_descartades)
            grading_avisos.append(
                f"{len(columnes_descartades)} columna/es descartada/es: {_etq} "
                f"(fora del sistema '{_ss_codi}'). S'han importat les talles conegudes.")
            Watchpoint.objects.create(
                model=model, task=None, created_by=user_profile,
                text=(f"Import: {len(columnes_descartades)} columna/es del document fora del "
                      f"sistema de talles '{_ss_codi}' descartada/es: {_etq}."),
                dades={'tipus': 'talles_descartades', 'etiquetes': columnes_descartades,
                       'size_system': _ss_codi},
            )

        if grading_avisos:
            session.avisos = (session.avisos or []) + grading_avisos

        # ── 4. PDF/document → ModelFitxer(tipus='DOCUMENT') amb versionat (re-import = v2).
        #     Delega la invariant a save_model_file (B2): re-import encadena (versio_anterior)
        #     i deixa is_current correcte. El naming {codi}_DOCUMENT_{NNN} es passa explícit.
        doc_fitxer = None
        if session.document:
            from .services_fitxers import save_model_file
            anterior = ModelFitxer.objects.filter(
                model=model, tipus='DOCUMENT',
            ).order_by('-id').first()
            num = (anterior.versio + 1) if anterior else 1
            ext = os.path.splitext(session.document.name)[1] or '.pdf'
            nom = f"{model.codi_intern}_DOCUMENT_{num:03d}{ext}"
            try:
                session.document.open('rb')
                doc_bytes = session.document.read()
            finally:
                session.document.close()
            doc_fitxer = save_model_file(
                model, ContentFile(doc_bytes),
                versio_anterior=anterior, tipus='DOCUMENT',
                origen='upload', nom=nom,
            )
            doc_fitxer.pujat_per = user_profile
            doc_fitxer.descripcio = 'Document origen de la importació guiada.'
            doc_fitxer.save(update_fields=['pujat_per', 'descripcio'])

        # ── 5. Teixit (si informat al Pas 4) → camps del model.
        teixit = (session.resultat or {}).get('teixit') or {}
        teixit_aplicat = False
        for f in _TEIXIT_FIELDS:
            if f in teixit and teixit[f] not in (None, ''):
                setattr(model, f, teixit[f])
                teixit_aplicat = True
        if teixit_aplicat:
            model.save()

        # ── 6. Tanca la sessió. L'EIX NO S'HI TOCA: la fila d'`ImportSession` que queda amb
        # `estat='CONFIRMAT'` porta el document, el model i la PRENDA on ha aterrat, i és el
        # registre import→peça que el brief demana (v. l'acta al camp `garment` del model).
        session.estat = 'CONFIRMAT'
        session.save(update_fields=['estat', 'avisos', 'actualitzat_at'])

    return Response({
        'ok': True,
        'estat': session.estat,
        'model_id': model.id,
        'model_codi': model.codi_intern,
        # SET-2/T8 — de quina prenda parla tot aquest resum.
        'garment': garment,
        'garment_nom': _nom_de_la_peca(session),
        # L'AVÍS INFORMATIU, MAI BLOQUEJANT: el cribratge va veure més d'un patró al
        # document i l'import ha anat sencer a UNA prenda. No demana cap clic —el desat ja
        # s'ha fet— i es pot ignorar amb raó legítima; el que no pot és no dir-se.
        'avis_multiprenda': ({
            'garment': garment,
            'garment_nom': _nom_de_la_peca(session),
        } if (session.resultat or {}).get('mes_duna_prenda') else None),
        'base_measurements': n_bm,
        'base_measurements_amb_valor': n_bm_valors,
        # B1 — el resultat de la poda mai és silenciós.
        'poms_podats': n_podats,
        'poms_conservats': (len(orfes) if poda_choice == 'conservar' else 0),
        'files_buides_desactivades': n_buides_soft,
        'manual_respectats': n_manual_respectats,
        'graded_specs': n_specs,
        'size_fitting': size_fitting.codi,
        'document_fitxer': (doc_fitxer.nom_fitxer if doc_fitxer else None),
        'teixit_aplicat': teixit_aplicat,
        'grading_rule_set': (new_rule_set.nom if new_rule_set else None),
        'grading_rules': n_rules,
        'grading_avisos': grading_avisos,
        # BEACH — columnes del document fora del sistema, descartades (no bloqueja; watchpoint creat).
        'columnes_descartades': columnes_descartades,
        'size_system_codi': (size_system.codi if size_system is not None else None),
        # D1 — la proposta viatja també a la resposta perquè el wizard la pugui ensenyar
        # a l'acte; la font persistent, però, és el Watchpoint (sobreviu al tancament del
        # wizard, que és justament on abans es perdia tot).
        'proposta_promocio': (watchpoint_promocio.dades if watchpoint_promocio else None),
        'proposta_promocio_watchpoint_id': (watchpoint_promocio.id if watchpoint_promocio else None),
        'message': f'Importació confirmada: {n_bm} POMs ({n_bm_valors} amb valor de base), regla '
                   f'(deltes+breaks) retinguda al model; grading propagat pendent de projecció conscient.',
    }, status=201)


# ═══════════════════════════════════════════════════════════════════════════════
# D1 · PROPOSTA DE PROMOCIÓ — aplicar la decisió humana, POM a POM.
#
# L'import deixa els POMs `amplia`/`conflicte` NOMÉS al model i n'obre un Watchpoint
# estructurat (v. `proposta_promocio`). Aquest endpoint és l'altra meitat: el moment en què
# algú decideix que un POM concret ha d'entrar al catàleg del client.
#
# Res s'hi escriu automàticament, ni tan sols aquí: només entren al contenidor els pom_id
# que arriben explícitament a `promocions`. Els que no s'hi esmenten es queden com estan.
# ═══════════════════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def promocionar_poms_view(request, model_id):
    """Promociona al contenidor de client els POMs triats d'una proposta d'import.

    Body: {'watchpoint_id': int, 'promocions': [pom_id, ...]}

    Per a cada POM promocionat:
      1. la seva regla entra al contenidor (`afegeix_regles_al_contenidor` — el mateix camí
         que ja fa servir l'import per al contenidor nou o esquelet: no n'estrenem cap);
      2. se n'esborren els `ModelGradingOverride`. Aquest pas NO és cosmètic: l'override té
         prioritat sobre qualsevol regla (`generate_graded_specs`), o sigui que sense
         esborrar-lo el POM continuaria graduant pel valor congelat del model i la regla
         acabada de promocionar no s'aplicaria mai. Promocionar i no heretar seria pitjor
         que no promocionar: semblaria fet i no ho estaria.

    Els POMs no esmentats queden en `nomes_model`. El Watchpoint es resol tot sol quan ja
    no queda ningú per decidir.
    """
    from django.db import transaction
    from fhort.models_app.models import Model, ModelGradingOverride, Watchpoint
    from fhort.models_app.services import (
        afegeix_regles_al_contenidor, PROMOCIO_CODI, PROMOCIO_NOMES_MODEL,
        PROMOCIO_PROMOCIONAT)
    from fhort.pom.models import GradingRuleSet

    model = Model.objects.filter(pk=model_id).first()
    if model is None:
        return Response({'error': 'model no trobat'}, status=404)

    wp = Watchpoint.objects.filter(pk=request.data.get('watchpoint_id'),
                                   model_id=model_id).first()
    if wp is None or not isinstance(wp.dades, dict) or wp.dades.get('codi') != PROMOCIO_CODI:
        return Response({'error': 'proposta_no_trobada',
                         'message': "No hi ha cap proposta de promoció per a aquest model."},
                        status=404)

    demanats = {int(p) for p in (request.data.get('promocions') or [])}
    items = wp.dades.get('items') or []
    pendents = {i['pom_id'] for i in items if i['estat'] == PROMOCIO_NOMES_MODEL}
    a_promocionar = demanats & pendents
    if demanats - pendents:
        # No és un error: pot ser un doble clic o una pestanya ranci. Es diu i s'ignora.
        _logging.getLogger(__name__).info(
            f"promocionar-poms model {model_id}: {sorted(demanats - pendents)} "
            f"ja decidits o forans; ignorats.")

    container = GradingRuleSet.objects.filter(pk=wp.dades.get('contenidor_id')).first()
    if container is None:
        return Response({'error': 'contenidor_absent',
                         'message': "El contenidor de la proposta ja no existeix."}, status=409)

    n_regles = n_ovr = 0
    if a_promocionar:
        specs = [i['spec'] for i in items if i['pom_id'] in a_promocionar]
        with transaction.atomic():
            n_regles = afegeix_regles_al_contenidor(
                container, specs, wp.dades.get('base_def_id'))
            n_ovr, _ = ModelGradingOverride.objects.filter(
                model_id=model_id, pom_id__in=a_promocionar).delete()
            for i in items:
                if i['pom_id'] in a_promocionar:
                    i['estat'] = PROMOCIO_PROMOCIONAT
            # `dades` és un JSONField mutat en memòria: cal reassignar-lo perquè Django
            # el vegi brut.
            wp.dades = {**wp.dades, 'items': items}
            if all(i['estat'] != PROMOCIO_NOMES_MODEL for i in items):
                wp.estat = 'resolved'
                wp.resolved_at = _dt.datetime.now(_dt.timezone.utc)
                wp.resolution_note = (
                    f"{sum(1 for i in items if i['estat'] == PROMOCIO_PROMOCIONAT)} POM(s) "
                    f"promocionats al contenidor #{container.id}.")
                perfil = getattr(request.user, 'profile', None) or getattr(
                    request.user, 'userprofile', None)
                if perfil is not None:
                    wp.resolved_by = perfil
            wp.save(update_fields=['dades', 'estat', 'resolved_at', 'resolved_by',
                                   'resolution_note'])

    return Response({
        'ok': True,
        'promocionats': sorted(a_promocionar),
        'regles_al_contenidor': n_regles,
        'overrides_esborrats': n_ovr,
        'watchpoint_estat': wp.estat,
        'items': items,
    }, status=200)
