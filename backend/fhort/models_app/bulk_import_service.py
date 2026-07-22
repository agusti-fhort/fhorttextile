"""models_app/bulk_import_service.py — pipeline d'import massiu de models per Excel.

Generació de plantilla (amb desplegables), parse del fitxer pujat, normalització,
validació per fila i per camp, dedup (intern + contra BD), preview i commit parcial
(Model + SizeFitting en bulk dins una sola transacció). Tota la lògica és Python; no hi
ha esquema nou (els models de staging viuen a models_app.models, Commit 1).

Missatges d'error: llegibles pel CLIENT final → "Columna '{nom}': {missatge humà}".
"""
import datetime
import io
import re

# Columnes del full Plantilla (en ordre). Les marcades a OBLIGATORIES s'exigeixen.
COLUMNS = [
    'nom_prenda', 'familia', 'tipus', 'any', 'temporada', 'target', 'construccio',
    'run_talles', 'talla_base', 'codi_client', 'col·leccio', 'color_referencia',
    'es_conjunt', 'referencia_conjunt', 'piece_number',
]
# Mínim viable per crear un Model (F5): la resta és opcional (informatiu o configuració, no
# bloquejant). El matching de talles és condicional i la config es completa després al model.
OBLIGATORIES = ['nom_prenda', 'any', 'temporada']
DROPDOWN_COLS = ['familia', 'tipus', 'any', 'temporada', 'target', 'construccio']
META_SHEET = '_meta'
PLANTILLA_SHEET = 'Plantilla'


def _norm(s):
    return str(s).strip() if s is not None else ''


def _key(s):
    return _norm(s).lower()


# Falsos textuals: el que el client escriu a una columna de sí/no quan vol dir NO.
# `bool('NO')` és True — una cel·la amb 'NO' es llegia com un sí i la fila petava demanant
# 'referencia_conjunt'. Tot el que no sigui un fals explícit (ni buit) compta com a sí.
FALSOS = {'', 'no', 'false', 'fals', '0', 'n', 'f', 'nan', 'none'}


def _as_bool(raw):
    """Booleà d'una cel·la d'Excel escrita per un humà: 'NO'/'FALSE'/'0'/buit → False."""
    return _norm(raw).lower() not in FALSOS


def _split_sizes(raw):
    """Separa un string de run de talles per comes/;/·/espais → llista neta."""
    parts = re.split(r'[,;·\n\t]+|\s{2,}', _norm(raw))
    out = []
    for p in parts:
        p = p.strip()
        if p:
            out.append(p)
    return out


# ───────────────────────────── Catàleg del tenant ─────────────────────────────

def build_catalog():
    """Diccionaris de lookup (label→objecte/codi) per a normalitzar els desplegables,
    + llistes de valors acceptats per als missatges d'error."""
    from fhort.pom.models import GarmentType, Target, ConstructionType
    from fhort.tasks.models import GarmentTypeItem

    fams = list(GarmentType.objects.filter(actiu=True).order_by('nom_client'))
    fam_by_key = {}
    for f in fams:
        fam_by_key[_key(f.nom_client)] = f
        if f.codi_client:
            fam_by_key.setdefault(_key(f.codi_client), f)

    items = list(GarmentTypeItem.objects.filter(active=True)
                 .select_related('garment_type').order_by('garment_type__nom_client', 'name'))
    item_by_key = {}
    item_labels = []
    for it in items:
        label = f"{it.garment_type.nom_client} / {it.name}"
        item_labels.append(label)
        item_by_key[_key(label)] = it
        item_by_key.setdefault(_key(it.name), it)   # tolerància: només el nom de l'ítem

    targets = list(Target.objects.order_by('display_order'))
    target_by_key, target_labels = {}, []
    for t in targets:
        target_labels.append(t.nom_en)
        for v in (t.nom_en, t.nom_cat, t.nom_es, t.codi):
            if v:
                target_by_key[_key(v)] = t.codi

    constrs = list(ConstructionType.objects.order_by('display_order'))
    constr_by_key, constr_labels = {}, []
    for c in constrs:
        constr_labels.append(c.nom_en)
        for v in (c.nom_en, c.nom_cat, c.nom_es, c.codi):
            if v:
                constr_by_key[_key(v)] = c.codi

    return {
        'fam_by_key': fam_by_key, 'fam_labels': [f.nom_client for f in fams],
        'item_by_key': item_by_key, 'item_labels': item_labels,
        'target_by_key': target_by_key, 'target_labels': target_labels,
        'constr_by_key': constr_by_key, 'constr_labels': constr_labels,
        # Codi → nom real del tenant: la conciliació ha d'ensenyar contra QUÈ ha casat una
        # cel·la ("Woman"), no el codi intern amb què es guarda ("WOMAN").
        'target_nom_by_codi': {t.codi: t.nom_en for t in targets},
        'constr_nom_by_codi': {c.codi: c.nom_en for c in constrs},
    }


# ───────────────────────────── Generació de plantilla ─────────────────────────────

def generate_template_bytes(customer):
    """Construeix el workbook de plantilla per a un Customer i el retorna com a bytes."""
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter, quote_sheetname
    from openpyxl.styles import Font
    from fhort.models_app.models import Model

    cat = build_catalog()
    wb = Workbook()

    # Full Plantilla (capçalera + files buides). Els obligatoris van en NEGRETA (no canvia el TEXT
    # de la capçalera → el parse, que llegeix per nom de columna, no es trenca).
    ws = wb.active
    ws.title = PLANTILLA_SHEET
    bold = Font(bold=True)
    for ci, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=name)
        if name in OBLIGATORIES:
            cell.font = bold

    # Full Instruccions com a FULL 1 (index=0): el primer que veu l'usuari. Reflecteix les 3 capes
    # de camp (obligatori mínim · opcional informatiu · opcional configuració "com més millor").
    inst = wb.create_sheet('Instruccions', 0)
    inst['A1'] = f"Aquesta plantilla és per a: {customer.codi} — {customer.nom}"
    inst['A3'] = "Omple una fila per model. Com més camps omplis, més complet entrarà el model; només 3 són obligatoris."
    inst['A5'] = "OBLIGATORI (mínim per importar): nom_prenda, any, temporada."
    inst['A6'] = "OPCIONAL informatiu: familia, tipus, codi_client, col·leccio, color_referencia."
    inst['A7'] = "OPCIONAL configuració (recomanat, NO bloquejant): target, construccio, run_talles, talla_base, es_conjunt, referencia_conjunt, piece_number."
    inst['A9'] = "Els camps obligatoris van en NEGRETA a la capçalera. Les columnes amb desplegable només accepten valors de la llista."
    inst['A11'] = "Formats de run de talles (separa per comes):"
    inst['A12'] = "Dona alpha: S,M,L,XL   ·   Dona numèric EU: 34,36,38,40"
    inst['A13'] = "Nen edat: 2,4,6,8,10   ·   Bebè mesos: 0M-1M,1M-3M,3M-6M"
    inst['A15'] = "La 'talla_base' ha de ser una de les talles del 'run_talles'."
    inst['A17'] = "Conjunts (combo): omple 'referencia_conjunt' igual a totes les peces i 'piece_number' (1,2,...)."
    # L'usuari obre directament sobre Instruccions.
    wb.active = 0

    # Fulls ocults de fonts per a les DataValidation.
    YEARS = [datetime.date.today().year + d for d in (-2, -1, 0, 1, 2)]
    sources = {
        '_families': cat['fam_labels'],
        '_items': cat['item_labels'],
        '_targets': cat['target_labels'],
        '_construccions': cat['constr_labels'],
        '_seasons': [c for c, _ in Model.TEMPORADA_CHOICES],
        '_years': [str(y) for y in YEARS],
    }
    refs = {}
    for sheet_name, values in sources.items():
        s = wb.create_sheet(sheet_name)
        s.sheet_state = 'hidden'
        for i, v in enumerate(values or [''], start=1):
            s.cell(row=i, column=1, value=v)
        refs[sheet_name] = f"{quote_sheetname(sheet_name)}!$A$1:$A${max(len(values), 1)}"

    # Full meta ocult: identificació del Customer (detecció de mismatch al pipeline).
    meta = wb.create_sheet(META_SHEET)
    meta.sheet_state = 'hidden'
    meta['A1'] = customer.codi
    meta['A2'] = customer.id

    # DataValidation per columna desplegable (files 2..601).
    NROWS = 600
    col_source = {
        'familia': '_families', 'tipus': '_items', 'any': '_years',
        'temporada': '_seasons', 'target': '_targets', 'construccio': '_construccions',
    }
    for col_name, sheet_name in col_source.items():
        col_idx = COLUMNS.index(col_name) + 1
        letter = get_column_letter(col_idx)
        dv = DataValidation(type='list', formula1=f"={refs[sheet_name]}", allow_blank=True)
        dv.add(f"{letter}2:{letter}{NROWS + 1}")
        ws.add_data_validation(dv)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ───────────────────────────── Parse del fitxer pujat ─────────────────────────────

def parse_upload(file_bytes):
    """Llegeix el xlsx pujat. Retorna (detected_customer_codi, raw_rows) on raw_rows és
    una llista de dicts {columna: valor} amb el seu row_num (índex Excel real)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    detected = None
    if META_SHEET in wb.sheetnames:
        detected = _norm(wb[META_SHEET]['A1'].value) or None

    ws = wb[PLANTILLA_SHEET] if PLANTILLA_SHEET in wb.sheetnames else wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not all_rows:
        return detected, []
    header = [_norm(c) for c in all_rows[0]]

    raw_rows = []
    for excel_idx, row in enumerate(all_rows[1:], start=2):
        cells = {}
        any_val = False
        for ci, col in enumerate(header):
            if not col:
                continue
            val = _norm(row[ci]) if ci < len(row) else ''
            cells[col] = val
            if val:
                any_val = True
        if any_val:   # ignora files completament buides
            raw_rows.append({'row_num': excel_idx, 'cells': cells})
    return detected, raw_rows


# ───────────────────────────── Resolució + validació d'una fila ─────────────────────────────

def resolve_row(cat, cells):
    """Normalitza i valida una fila. Retorna (resolved, errors, warnings).
    `resolved` porta els objectes/codis llestos per a crear el Model (si no hi ha errors durs)."""
    from fhort.models_app.matching import match_size_system

    errors, warnings = [], []
    g = lambda k: _norm(cells.get(k, ''))

    # Obligatoris buits
    for col in OBLIGATORIES:
        if not g(col):
            errors.append({'camp': col, 'missatge_client': f"Columna '{col}': camp obligatori."})

    resolved = {}

    # família
    fam = cat['fam_by_key'].get(_key(g('familia'))) if g('familia') else None
    if g('familia') and not fam:
        errors.append({'camp': 'familia',
                       'missatge_client': f"Columna 'familia': '{g('familia')}' no és vàlid. "
                       f"Valors acceptats: {', '.join(cat['fam_labels'][:12])}…"})
    # tipus
    item = cat['item_by_key'].get(_key(g('tipus'))) if g('tipus') else None
    if g('tipus') and not item:
        errors.append({'camp': 'tipus',
                       'missatge_client': f"Columna 'tipus': '{g('tipus')}' no és vàlid. "
                       f"Valors acceptats: {', '.join(cat['item_labels'][:12])}…"})
    # coherència família/tipus
    if fam and item and item.garment_type_id != fam.id:
        errors.append({'camp': 'tipus',
                       'missatge_client': f"Columna 'tipus': '{g('tipus')}' no pertany a la "
                       f"família '{g('familia')}'."})

    # any
    year = None
    if g('any'):
        try:
            year = int(float(g('any')))
        except (TypeError, ValueError):
            errors.append({'camp': 'any', 'missatge_client': f"Columna 'any': '{g('any')}' no és un any vàlid."})

    # temporada
    from fhort.models_app.models import Model
    seasons = {c for c, _ in Model.TEMPORADA_CHOICES}
    season = g('temporada').upper()
    if g('temporada') and season not in seasons:
        errors.append({'camp': 'temporada',
                       'missatge_client': f"Columna 'temporada': '{g('temporada')}' no és vàlid. "
                       f"Valors acceptats: {', '.join(sorted(seasons))}."})

    # target
    target_codi = cat['target_by_key'].get(_key(g('target'))) if g('target') else None
    if g('target') and not target_codi:
        errors.append({'camp': 'target',
                       'missatge_client': f"Columna 'target': '{g('target')}' no és vàlid. "
                       f"Valors acceptats: {', '.join(cat['target_labels'])}."})

    # construcció
    constr_codi = cat['constr_by_key'].get(_key(g('construccio'))) if g('construccio') else None
    if g('construccio') and not constr_codi:
        errors.append({'camp': 'construccio',
                       'missatge_client': f"Columna 'construccio': '{g('construccio')}' no és vàlid. "
                       f"Valors acceptats: {', '.join(cat['constr_labels'])}."})

    # run_talles + matching
    labels = _split_sizes(g('run_talles'))
    base_size = g('talla_base')
    size_system = None
    if target_codi and labels:
        mr = match_size_system(target_codi, labels, base_size)
        size_system = mr.size_system
        if mr.error:
            # El motiu ha d'anar a la cel·la que el causa: match_size_system només emet
            # l'error de base quan el run SÍ que ha casat (score>=0.5) i la base no hi és.
            camp = 'talla_base' if (mr.score >= 0.5 and not mr.base_ok) else 'run_talles'
            errors.append({'camp': camp, 'missatge_client': mr.error})
        elif mr.warning:
            warnings.append({'camp': 'run_talles', 'missatge_client': mr.warning})

    # conjunts
    es_conjunt = _as_bool(g('es_conjunt'))
    ref_conjunt = g('referencia_conjunt')
    piece_number = None
    if es_conjunt and not ref_conjunt:
        errors.append({'camp': 'referencia_conjunt',
                       'missatge_client': "Columna 'referencia_conjunt': obligatòria quan 'es_conjunt' està marcat."})
    if ref_conjunt:
        if not g('piece_number'):
            errors.append({'camp': 'piece_number',
                           'missatge_client': "Columna 'piece_number': obligatòria quan hi ha 'referencia_conjunt'."})
        else:
            try:
                piece_number = int(float(g('piece_number')))
            except (TypeError, ValueError):
                errors.append({'camp': 'piece_number',
                               'missatge_client': f"Columna 'piece_number': '{g('piece_number')}' no és un número."})

    # PORTA ÚNICA DEL RUN (llei S24b): l'ordre de la cel·la d'Excel no mana. S'ordena aquí, en
    # el punt on `run_labels` es construeix, perquè les dues escriptures que en pengen
    # (`_build_model` i `_complement_existing`) rebin ja el run canònic sense duplicar la crida.
    if labels and size_system is not None:
        from fhort.pom.grading_utils import run_del_model
        labels, _desconegudes = run_del_model(labels, size_system)
        if _desconegudes:
            errors.append({
                'camp': 'run_talles',
                'missatge_client': (
                    f"Columna 'run_talles': aquestes talles no pertanyen al sistema "
                    f"'{size_system.codi}': {', '.join(_desconegudes)}."
                ),
            })

    resolved.update({
        'nom_prenda': g('nom_prenda'), 'garment_type': fam, 'garment_type_item': item,
        'any': year, 'temporada': season, 'target': target_codi, 'construction': constr_codi,
        'run_labels': labels, 'base_size': base_size, 'size_system': size_system,
        'codi_client': g('codi_client'), 'collection': g('col·leccio'),
        'color_referencia': g('color_referencia'),
        'es_conjunt': es_conjunt or bool(ref_conjunt), 'ref_conjunt': ref_conjunt,
        'piece_number': piece_number,
    })
    return resolved, errors, warnings


# ───────────────────────────── Validació del lot (preview) ─────────────────────────────

def validate_rows(customer, raw_rows):
    """Valida totes les files: per-camp + dedup intern + dedup contra BD. Retorna
    (results, resum). `results` = llista de dicts per crear BulkCollectionRow."""
    from fhort.models_app.models import Model

    cat = build_catalog()
    results = []
    seen = {}   # clau dedup intern → row_num original
    counts = {'OK': 0, 'ERROR': 0, 'AVIS': 0, 'DUPLICAT': 0}

    for rr in raw_rows:
        cells = rr['cells']
        resolved, errors, warnings = resolve_row(cat, cells)

        estat = None
        # dedup intern
        dedup_key = (
            _key(resolved.get('nom_prenda')), resolved.get('any'), resolved.get('temporada'),
            _key(resolved.get('codi_client')) or _key(resolved.get('collection')),
        )
        if not errors and all(dedup_key[:3]):
            if dedup_key in seen:
                estat = 'DUPLICAT'
                errors.append({'camp': '_fila',
                               'missatge_client': f"Fila duplicada a la pujada (duplica la fila {seen[dedup_key]})."})
            else:
                seen[dedup_key] = rr['row_num']

        # dedup contra BD (semi-dur)
        complement = False
        if not errors and estat is None and all(dedup_key[:3]):
            existing = Model.objects.filter(
                customer=customer, nom_prenda=resolved['nom_prenda'],
                any=resolved['any'], temporada=resolved['temporada']).first()
            if existing:
                complement = True
                warnings.append({'camp': '_fila',
                                 'missatge_client': "Ja existeix un model amb aquest nom/any/temporada; "
                                 "s'actualitzaran els camps buits."})

        if estat is None:
            estat = 'ERROR' if errors else ('AVIS' if warnings else 'OK')
        counts[estat] += 1

        results.append({
            'row_num': rr['row_num'], 'raw_data': cells, 'estat': estat,
            'errors': errors + warnings, '_complement': complement,
        })

    conjunts = len({_key(r['raw_data'].get('referencia_conjunt'))
                    for r in results if r['estat'] in ('OK', 'AVIS')
                    and _key(r['raw_data'].get('referencia_conjunt'))})
    resum = {
        'total': len(results), 'ok': counts['OK'], 'errors': counts['ERROR'],
        'avisos': counts['AVIS'], 'duplicats': counts['DUPLICAT'], 'conjunts': conjunts,
    }
    return results, resum


# ───────────────────── Pla de codis (compartit: conciliació ↔ commit) ─────────────────────
#
# El que el tècnic VEU a la conciliació ha de ser exactament el que s'escriurà al commit.
# Això només es garanteix si les dues bandes calculen el pla amb LA MATEIXA llei; per això
# la classificació i l'assignació de codis viuen aquí i les criden totes dues. L'única cosa
# que canvia és qui reparteix els números: el commit RESERVA (escriu), la conciliació només
# fa una ULLADA (no escriu). Un segon rellotge de codis seria repetir el bug que vam matar.

def _classify(cat, customer, rows):
    """Re-resol cada fila amb el catàleg actual i la classifica:
    complement (el model ja existeix) / peça de conjunt / simple. Les files amb errors cauen."""
    from fhort.models_app.models import Model

    simples, set_groups, complements = [], {}, []   # set_groups: ref → [(row, resolved), ...]
    for row in rows:
        resolved, errors, _w = resolve_row(cat, row.raw_data)
        if errors:
            continue   # el catàleg ha canviat des del preview → saltar (no peta el commit)
        existing = Model.objects.filter(
            customer=customer, nom_prenda=resolved['nom_prenda'],
            any=resolved['any'], temporada=resolved['temporada']).first()
        if existing and not resolved['ref_conjunt']:
            complements.append((row, resolved, existing))
        elif resolved['ref_conjunt']:
            set_groups.setdefault(resolved['ref_conjunt'], []).append((row, resolved))
        else:
            simples.append((row, resolved))
    return simples, set_groups, complements


def _group_by_season(simples, set_groups):
    """(year, season) → {simples, sets}. Ordre determinista: el codi d'una fila no pot
    dependre de l'atzar d'un diccionari."""
    groups = {}
    for row, r in simples:
        groups.setdefault((r['any'], r['temporada']), {'simples': [], 'sets': []})['simples'].append((row, r))
    for ref, pieces in set_groups.items():
        r0 = pieces[0][1]
        groups.setdefault((r0['any'], r0['temporada']), {'simples': [], 'sets': []})['sets'].append((ref, pieces))
    return groups


def _plan_codes(customer, groups, allocate):
    """Assigna els codi_intern que ocuparà aquest import.

    `allocate(year, season, n) -> primer seqüencial`. El commit li passa la reserva atòmica;
    la conciliació, una ullada al comptador que no escriu res.
    Retorna (plan_simples, plan_sets): [(row, resolved, codi, seq)] i [(ref, pieces, codi_base)].
    """
    plan_simples, plan_sets = [], []
    for (year, season), grp in groups.items():
        n = len(grp['simples']) + len(grp['sets'])
        if n <= 0:
            continue
        seq = allocate(year, season, n)
        yy = str(year)[-2:].zfill(2)
        for row, r in grp['simples']:
            plan_simples.append((row, r, f"{customer.codi}-{season}{yy}-{str(seq).zfill(4)}", seq))
            seq += 1
        for ref, pieces in grp['sets']:
            plan_sets.append((ref, pieces, f"{customer.codi}-{season}{yy}-{str(seq).zfill(4)}"))
            seq += 1
    return plan_simples, plan_sets


# ───────────────────────────── Commit parcial ─────────────────────────────

def commit_import(imp, creat_per_profile):
    """Crea els Models + SizeFittings de les files OK/AVIS dins una sola transacció.
    Genera codi_intern al pipeline (bulk_create bypassa el signal). Retorna stats."""
    from django.db import transaction
    from fhort.models_app.models import Model, GarmentSet, BulkCollectionRow, Watchpoint
    from fhort.fitting.models import SizeFitting
    from fhort.models_app.services import (
        reserve_sequence_range, model_config_missing, config_missing_text)

    cat = build_catalog()
    customer = imp.customer
    rows = list(BulkCollectionRow.objects.filter(importacio=imp, estat__in=['OK', 'AVIS'])
                .order_by('row_num'))

    simples, set_groups, complements = _classify(cat, customer, rows)
    groups = _group_by_season(simples, set_groups)

    simple_models = []          # (row, Model)
    set_plan = []               # (GarmentSet, [(row, resolved), ...])

    with transaction.atomic():
        # 1) Assignar codis amb la MATEIXA llei que la conciliació, però reservant de debò.
        def reserva(year, season, n):
            first, _last = reserve_sequence_range(customer, year, season, n)
            return first

        plan_simples, plan_sets = _plan_codes(customer, groups, reserva)
        for row, r, codi, seq in plan_simples:
            simple_models.append((row, _build_model(customer, codi, seq, r, creat_per_profile)))
        for _ref, pieces, codi_base in plan_sets:
            set_plan.append((GarmentSet(codi_base=codi_base,
                                        nom_comercial=pieces[0][1]['nom_prenda'] or '',
                                        num_pieces=len(pieces)), pieces))

        # 2) Crear GarmentSets (per tenir pk abans de les peces).
        if set_plan:
            GarmentSet.objects.bulk_create([g for g, _p in set_plan])

        # 3) Models de les peces dels conjunts.
        piece_models = []
        for gset, pieces in set_plan:
            for row, r in pieces:
                pn = r['piece_number'] or 1
                m = _build_model(customer, f"{gset.codi_base}-{str(pn).zfill(2)}", seq=None,
                                 r=r, creat_per_profile=creat_per_profile)
                m.garment_set = gset
                m.piece_number = pn
                piece_models.append((row, m))

        # 4) bulk_create de tots els Models (simples + peces) — bypassa signals.
        all_models = simple_models + piece_models
        if all_models:
            Model.objects.bulk_create([m for _row, m in all_models])

        # 5) SizeFittings (segon bulk, mateixa transacció). creat_per = el tècnic.
        sfs = [SizeFitting(model=m, numero=1, codi=f"{m.codi_intern}-SF1", tipus='Proto',
                           estat='Pendent', base_tancada=False, creat_per=creat_per_profile)
               for _row, m in all_models]
        if sfs:
            SizeFitting.objects.bulk_create(sfs)

        # 5b) F3 — Watchpoint d'import VIU: per cada model amb config incompleta, un avís
        # estructurat (task=None → origen sistema; dades = claus que falten de model_config_missing).
        # Es recalcularà/resoldrà sol via post_save (signals.py) en omplir-se els camps. bulk_create
        # bypassa signals, per això la creació es fa aquí explícitament.
        wps = [Watchpoint(model=m, task=None, dades=missing,
                          text=config_missing_text(missing), estat='open')
               for _row, m in all_models
               for missing in [model_config_missing(m)] if missing]
        if wps:
            Watchpoint.objects.bulk_create(wps)

        # 6) Enllaçar BulkCollectionRow.model_creat.
        for row, m in all_models:
            row.model_creat = m
        if all_models:
            BulkCollectionRow.objects.bulk_update([row for row, _m in all_models], ['model_creat'])

        # 7) Complements: omplir camps buits dels models existents.
        n_compl = 0
        for row, r, existing in complements:
            changed = _complement_existing(existing, r)
            if changed:
                existing.save(update_fields=changed)
            row.model_creat = existing
            row.save(update_fields=['model_creat'])
            n_compl += 1

        imp.estat = 'IMPORTAT'
        imp.resum = {**(imp.resum or {}), 'creats': len(all_models),
                     'conjunts_creats': len(set_plan), 'complementats': n_compl}
        imp.save(update_fields=['estat', 'resum'])

    return {'models': len(all_models), 'sets': len(set_plan),
            'size_fittings': len(sfs), 'complementats': n_compl}


def _build_model(customer, codi_intern, seq, r, creat_per_profile):
    from fhort.models_app.models import Model
    return Model(
        codi_intern=codi_intern,
        codi_client=r['codi_client'] or '',
        customer=customer,
        codi_tenant=customer.codi,
        any=r['any'], temporada=r['temporada'],
        sequencial=seq or 1,
        nom_prenda=r['nom_prenda'] or None,
        color_referencia=r['color_referencia'] or None,
        collection=r['collection'] or '',
        garment_type=r['garment_type'],
        garment_type_item=r['garment_type_item'],
        target=r['target'], construction=r['construction'],
        size_system=r['size_system'],
        size_run_model='·'.join(r['run_labels']) if r['run_labels'] else None,
        base_size_label=r['base_size'] or None,
        responsable=creat_per_profile,
        estat='Nou',
    )


def _complement_existing(existing, r):
    """Omple només els camps buits del model existent amb els valors de la fila nova."""
    changed = []
    mapping = {
        'nom_prenda': r['nom_prenda'], 'color_referencia': r['color_referencia'],
        'collection': r['collection'], 'base_size_label': r['base_size'],
        'size_run_model': '·'.join(r['run_labels']) if r['run_labels'] else None,
    }
    for field_name, new_val in mapping.items():
        if new_val and not getattr(existing, field_name, None):
            setattr(existing, field_name, new_val)
            changed.append(field_name)
    if r['size_system'] and not existing.size_system_id:
        existing.size_system = r['size_system']
        changed.append('size_system')
    if r['garment_type_item'] and not existing.garment_type_item_id:
        existing.garment_type_item = r['garment_type_item']
        changed.append('garment_type_item')
    return changed


# ───────────────────────────── Conciliació (dry-run enriquit) ─────────────────────────────
#
# La llei de la casa: el sistema ENSENYA el que ha entès i espera confirmació — mai endevina
# en silenci. El preview antic validava FORMAT ("20 files OK") i callava l'ENCAIX; el commit
# petava després. Aquí, per cada fila i cada camp mapat, es diu: què deia el fitxer, contra
# què ha casat al catàleg del tenant, i si això ha calgut transformar-ho.
#
# Read-only i idempotent: ni escriu, ni reserva números, ni canvia l'estat de la importació.

# Camps que es concilien (els que es resolen contra el catàleg o es normalitzen).
RECONCILED_FIELDS = ['familia', 'tipus', 'target', 'construccio', 'temporada', 'any',
                     'run_talles', 'talla_base', 'es_conjunt']


def _resolved_display(cat, camp, r):
    """(valor_resolt, candidat) d'un camp ja resolt. El candidat porta l'id i el NOM REAL del
    tenant: el tècnic ha de veure contra QUÈ ha casat la cel·la, no un codi intern."""
    if camp == 'familia':
        f = r.get('garment_type')
        return (f.nom_client, {'id': f.id, 'nom': f.nom_client}) if f else ('', None)
    if camp == 'tipus':
        it = r.get('garment_type_item')
        if not it:
            return ('', None)
        nom = f"{it.garment_type.nom_client} / {it.name}"
        return (nom, {'id': it.id, 'nom': nom})
    if camp == 'target':
        codi = r.get('target')
        if not codi:
            return ('', None)
        nom = cat['target_nom_by_codi'].get(codi, codi)
        return (nom, {'id': codi, 'nom': nom})
    if camp == 'construccio':
        codi = r.get('construction')
        if not codi:
            return ('', None)
        nom = cat['constr_nom_by_codi'].get(codi, codi)
        return (nom, {'id': codi, 'nom': nom})
    if camp == 'temporada':
        return (r.get('temporada') or '', None)
    if camp == 'any':
        return (str(r['any']) if r.get('any') else '', None)
    if camp == 'run_talles':
        labels = r.get('run_labels') or []
        ss = r.get('size_system')
        return ('·'.join(labels), {'id': ss.id, 'nom': ss.codi} if ss else None)
    if camp == 'talla_base':
        return (r.get('base_size') or '', None)
    if camp == 'es_conjunt':
        return ('SI' if r.get('es_conjunt') else 'NO', None)
    return ('', None)


def _reconcile_fields(cat, cells, resolved, errors):
    """Per cada camp mapat: valor del fitxer, valor resolt, i en quin dels quatre estats cau.

    MATCH        — el fitxer ja deia exactament el que el catàleg diu.
    NORMALITZAT  — s'ha transformat de manera determinista (trim, 'Woman'→Woman, comes→·).
                   Es fa I ES MOSTRA: el tècnic no ho ha de descobrir després.
    NO_MATCH     — no casa amb res del catàleg. Bloqueja LA FILA, mai tot l'import.
    BUIT         — la cel·la és buida i el camp és opcional. Ni encert ni error: no s'ha dit res.
    """
    err_by_camp = {}
    for e in errors:
        err_by_camp.setdefault(e['camp'], e['missatge_client'])

    camps = []
    for camp in RECONCILED_FIELDS:
        raw = _norm(cells.get(camp, ''))
        val, candidat = _resolved_display(cat, camp, resolved)
        motiu = err_by_camp.get(camp)
        if motiu:
            estat = 'NO_MATCH'
        elif not raw:
            estat = 'BUIT'
        elif val and val != raw:
            estat = 'NORMALITZAT'
        else:
            estat = 'MATCH'
        camps.append({'camp': camp, 'valor_fitxer': raw, 'valor_resolt': val,
                      'estat': estat, 'candidat': candidat, 'motiu': motiu})
    return camps


def reconcile(imp):
    """Conciliació completa d'una importació previsada: files × camps + els codis que ocuparà.

    Els codis previstos surten del MATEIX pla que farà servir el commit (_plan_codes), però amb
    una ullada al comptador en comptes d'una reserva. Es verifiquen un a un contra la BD: el
    "20 files OK" passa a ser "20 files OK i 20 codis lliures".
    """
    from fhort.models_app.models import BulkCollectionRow, Model
    from fhort.models_app.services import sequence_floor

    cat = build_catalog()
    customer = imp.customer
    rows = list(BulkCollectionRow.objects.filter(importacio=imp).order_by('row_num'))
    importables = [r for r in rows if r.estat in ('OK', 'AVIS')]

    # Pla de codis — ullada, no reserva (idempotent: cridar-ho dos cops dona el mateix).
    simples, set_groups, complements = _classify(cat, customer, importables)
    groups = _group_by_season(simples, set_groups)
    plan_simples, plan_sets = _plan_codes(
        customer, groups, lambda year, season, _n: sequence_floor(customer, year, season) + 1)

    codi_by_row = {row.row_num: codi for row, _r, codi, _seq in plan_simples}
    for _ref, pieces, codi_base in plan_sets:
        for row, r in pieces:
            codi_by_row[row.row_num] = f"{codi_base}-{str(r['piece_number'] or 1).zfill(2)}"
    compl_by_row = {row.row_num: ex.codi_intern for row, _r, ex in complements}

    # Anti-col·lisió VISIBLE: cap codi previst pot ser ja a la BD.
    previstos = list(codi_by_row.values())
    ocupats = set(Model.objects.filter(codi_intern__in=previstos)
                  .values_list('codi_intern', flat=True)) if previstos else set()

    files = []
    for row in rows:
        resolved, errors, warnings = resolve_row(cat, row.raw_data or {})
        codi = codi_by_row.get(row.row_num)
        files.append({
            'row_num': row.row_num,
            'estat': row.estat,
            'nom': _norm((row.raw_data or {}).get('nom_prenda')),
            'camps': _reconcile_fields(cat, row.raw_data or {}, resolved, errors),
            'codi_previst': codi,
            'codi_lliure': (codi not in ocupats) if codi else None,
            'complementa': compl_by_row.get(row.row_num),
            'motius': [e['missatge_client'] for e in errors + warnings],
        })

    resum = {
        'total': len(files),
        'netes': sum(1 for f in files if f['estat'] == 'OK'),
        'avisos': sum(1 for f in files if f['estat'] == 'AVIS'),
        'bloquejades': sum(1 for f in files if f['estat'] in ('ERROR', 'DUPLICAT')),
        'importables': len(importables),
        'codis_previstos': len(previstos),
        'codis_ocupats': len(ocupats),
        'complements': len(compl_by_row),
    }
    return {'import_id': imp.id, 'estat': imp.estat, 'resum': resum, 'files': files}


# ───────────────────────────── Informe d'errors (xlsx) ─────────────────────────────

def errors_report_bytes(imp):
    """Genera un xlsx amb les files en estat ERROR/DUPLICAT: columnes originals + una
    columna ERRORS amb tots els missatges (llegibles pel client)."""
    from openpyxl import Workbook
    from fhort.models_app.models import BulkCollectionRow

    wb = Workbook()
    ws = wb.active
    ws.title = 'Errors'
    for ci, name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=ci, value=name)
    ws.cell(row=1, column=len(COLUMNS) + 1, value='ERRORS')

    rows = BulkCollectionRow.objects.filter(
        importacio=imp, estat__in=['ERROR', 'DUPLICAT']).order_by('row_num')
    r = 2
    for row in rows:
        for ci, col in enumerate(COLUMNS, start=1):
            ws.cell(row=r, column=ci, value=(row.raw_data or {}).get(col, ''))
        msgs = ' | '.join(e.get('missatge_client', '') for e in (row.errors or []))
        ws.cell(row=r, column=len(COLUMNS) + 1, value=msgs)
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
